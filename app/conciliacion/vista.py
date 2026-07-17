"""Pestaña 'Conciliaciones Bancarias' (Flet).

Flujo: elegir rango de fechas -> subir Excel del banco -> normalizar (openpyxl en
hilo) -> traer movimientos del sistema (BigQuery en hilo) -> conciliar -> mostrar
los 4 grupos en tablas. Todo el I/O va en asyncio.to_thread para no bloquear la UI.

Devuelve (tab, contenido) igual que construir_tab_dashboard, para que main.py lo
inserte en TabBar/TabBarView en la misma posición.
"""

import asyncio
import dataclasses
import os
import re
import tempfile
import traceback
from datetime import date, datetime, timedelta

import flet as ft
from flet_datatable2 import DataColumn2, DataColumnSize, DataTable2

from .conciliador import conciliar
from .ingresos_diversos import cargar_ingresos_diversos
from .lector_banco import nombres_bancos, normalizar_banco
from .modelo import MovimientoConciliacion, ResultadoConciliacion
from ..parsers.lectura import EXTENSIONES
from ..services.bigquery_repository import BigQueryRepository

# Color por grupo (acento de la tarjeta/tabla).
_COLOR_CONCILIADOS = "#1baf7a"
_COLOR_SOLO_BANCO = "#eda100"
_COLOR_SOLO_SISTEMA = "#2a78d6"
_COLOR_CHEQUES = "#e34948"
_COLOR_REPETIDOS = "#7e57c2"
_COLOR_FUERA_RANGO = "#607d8b"


_MESES_ES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]


def _fmt_fecha(f: date | None) -> str:
    return f.strftime("%d/%m/%Y") if f else ""


def _fmt_fecha_larga(f: date) -> str:
    """01 jul 2026 — con mes en español (strftime %b depende del locale del SO)."""
    return f"{f.day:02d} {_MESES_ES[f.month - 1]} {f.year}"


def _fmt_importe(v: float) -> str:
    return f"${v:,.2f}"


def _hex_argb(color: str) -> str:
    """'#1baf7a' -> 'FF1BAF7A' (ARGB que espera openpyxl)."""
    return "FF" + color.lstrip("#").upper()


def _nombre_hoja(nombre: str) -> str:
    """Sanea el nombre para una hoja de Excel (máx 31 chars, sin \\ / ? * [ ] :)."""
    limpio = re.sub(r"[\\/?*\[\]:]", " ", nombre).strip()
    return (limpio or "Hoja")[:31]


def _construir_workbook(res: "ResultadoConciliacion", secciones: list[dict], generado: str):
    """Arma el .xlsx: una hoja 'Resumen' + una hoja por sección con sus movimientos.

    `secciones` es la lista de dicts que produce la UI (titulo/hoja/color/columnas/
    filas/total) para que Excel y pantalla muestren exactamente lo mismo."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    borde = Side(style="thin", color="FFD0D0D0")
    marco = Border(left=borde, right=borde, top=borde, bottom=borde)

    wb = openpyxl.Workbook()

    # --- Hoja Resumen ---------------------------------------------------------
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Conciliación Bancaria"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generado: {generado}"
    ws["A2"].font = Font(italic=True, size=10, color="FF666666")
    if res.ventana is not None:
        ini, fin = res.ventana
        ws["A3"] = f"Ventana de fechas conciliada: {_fmt_fecha(ini)} – {_fmt_fecha(fin)}"
    else:
        ws["A3"] = "Ventana de fechas: sin filtro (algún archivo no trae fechas)."
    ws["A3"].font = Font(size=10, color="FF666666")

    encabezados = ["Sección", "Movimientos", "Importe total"]
    fila0 = 5
    for j, h in enumerate(encabezados, start=1):
        c = ws.cell(row=fila0, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill(fill_type="solid", fgColor="FF37474F")
        c.border = marco
    for i, s in enumerate(secciones, start=fila0 + 1):
        ws.cell(row=i, column=1, value=s["titulo"]).border = marco
        ws.cell(row=i, column=2, value=len(s["filas"])).border = marco
        cimp = ws.cell(row=i, column=3, value=round(s["total"], 2))
        cimp.number_format = '"$"#,##0.00'
        cimp.border = marco
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    # --- Una hoja por sección -------------------------------------------------
    usados: set[str] = set()
    for s in secciones:
        nombre = _nombre_hoja(s["hoja"])
        base, k = nombre, 2
        while nombre in usados:  # evita choque de nombres de hoja
            nombre = f"{base[:28]} {k}"
            k += 1
        usados.add(nombre)
        hoja = wb.create_sheet(nombre)

        etiquetas = [col[0] for col in s["columnas"]]
        col_importe = {i for i, e in enumerate(etiquetas) if "Importe" in e}
        fill = PatternFill(fill_type="solid", fgColor=_hex_argb(s["color"]))
        for j, e in enumerate(etiquetas, start=1):
            c = hoja.cell(row=1, column=j, value=e)
            c.font = Font(bold=True, color="FFFFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
            c.border = marco
        if not s["filas"]:
            hoja.cell(row=2, column=1, value="Sin movimientos en este grupo.").font = Font(italic=True, color="FF888888")
        for r, fila in enumerate(s["filas"], start=2):
            for j, valor in enumerate(fila):
                c = hoja.cell(row=r, column=j + 1, value=valor)
                c.border = marco
                if j in col_importe:  # los importes van a la derecha
                    c.alignment = Alignment(horizontal="right")
        # Anchos de columna: proporcional a la etiqueta, con mínimos razonables.
        for j, e in enumerate(etiquetas, start=1):
            ancho = 40 if "Descripción" in e else max(14, len(e) + 4)
            hoja.column_dimensions[get_column_letter(j)].width = ancho
        hoja.freeze_panes = "A2"

    return wb


def construir_tab_conciliaciones(page: ft.Page) -> tuple[ft.Tab, ft.Control]:
    # El diálogo compacto (INPUT) del DateRangePicker trae textos (encabezado, ayuda
    # y etiquetas) demasiado grandes para su tamaño ajustado. El estilo del picker
    # solo se configura a nivel tema (no por instancia), así que se afina aquí para
    # claro y oscuro. Esta pestaña se construye después del Dashboard, por lo que su
    # ajuste es el que prevalece (unifica ambos selectores).
    _tema_dp = ft.DatePickerTheme(
        range_picker_header_headline_text_style=ft.TextStyle(size=15, weight=ft.FontWeight.W_600),
        range_picker_header_help_text_style=ft.TextStyle(size=11),
        header_headline_text_style=ft.TextStyle(size=15, weight=ft.FontWeight.W_600),
        header_help_text_style=ft.TextStyle(size=11),
        weekday_text_style=ft.TextStyle(size=12),
        day_text_style=ft.TextStyle(size=12),
        year_text_style=ft.TextStyle(size=13),
    )
    if page.theme is not None:
        page.theme = dataclasses.replace(page.theme, date_picker_theme=_tema_dp)
    if page.dark_theme is not None:
        page.dark_theme = dataclasses.replace(page.dark_theme, date_picker_theme=_tema_dp)

    hoy = date.today()
    # Rango por defecto: el día anterior (la conciliación normalmente se hace sobre
    # el movimiento del día previo, no el de hoy que aún no cierra).
    ayer = hoy - timedelta(days=1)
    rango_sel: list[tuple[date, date]] = [(ayer, ayer)]
    # Archivos de banco cargados: uno o varios, cada uno con su selector de banco.
    # Cada entrada: {"path","nombre","dropdown","fila"}.
    archivos_banco: list[dict] = []
    lista_archivos = ft.Column(spacing=6)         # UI: un renglón por archivo cargado
    archivo_sistema: list[str | None] = [None]    # ruta temporal del Excel de Ingresos Diversos
    nombre_sistema = ft.Text("", size=12, color=ft.Colors.ON_SURFACE_VARIANT)
    # Última conciliación calculada (para exportarla a Excel sin recalcular).
    ultimo_resultado: list[ResultadoConciliacion | None] = [None]

    # El repositorio se crea perezosamente (necesita credenciales de BigQuery); así
    # la pestaña se construye aunque BigQuery no esté configurado en ese momento.
    repo_holder: list[BigQueryRepository | None] = [None]

    def _repo() -> BigQueryRepository:
        if repo_holder[0] is None:
            repo_holder[0] = BigQueryRepository()
        return repo_holder[0]

    # En Flet 0.85 el FilePicker es un servicio: se crea y se usa directamente (NO
    # se agrega a page.overlay; hacerlo provoca "Unknown control: FilePicker").
    file_picker = ft.FilePicker()

    # SnackBar propio para avisos (Flet 0.85 no tiene page.open(); se usa el patrón
    # del resto de la app: control en overlay + .open = True + page.update()).
    snackbar = ft.SnackBar(content=ft.Text(""))
    page.overlay.append(snackbar)

    # --- Encabezado y barra de herramientas -------------------------------------
    titulo = ft.Text("Conciliación Bancaria", size=20, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE)
    subtitulo = ft.Text(
        "Agrega uno o varios estados de cuenta (elige el banco de cada uno) y compáralos "
        "contra los movimientos del sistema.",
        size=12,
        color=ft.Colors.ON_SURFACE_VARIANT,
    )
    estado_text = ft.Text("", size=12, color=ft.Colors.RED_600)
    progress = ft.ProgressRing(width=16, height=16, visible=False, stroke_width=2)

    def _texto_rango(inicio: date, fin: date) -> str:
        return f"{_fmt_fecha_larga(inicio)} – {_fmt_fecha_larga(fin)}"

    def on_cambiar_rango(e) -> None:
        picker = e.control
        if not picker.start_value or not picker.end_value:
            return
        rango_sel[0] = (picker.start_value.date(), picker.end_value.date())
        boton_rango.content.controls[1].value = _texto_rango(*rango_sel[0])
        page.update()

    date_range_picker = ft.DateRangePicker(
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2035, 12, 31),
        start_value=datetime.combine(ayer, datetime.min.time()),
        end_value=datetime.combine(ayer, datetime.min.time()),
        # Calendario siempre visible (sin modo escritura): más intuitivo para el
        # usuario — se eligen las fechas tocando los días directamente.
        entry_mode=ft.DatePickerEntryMode.CALENDAR_ONLY,
        on_change=on_cambiar_rango,
    )
    boton_rango = ft.OutlinedButton(
        content=ft.Row(
            [ft.Icon(ft.Icons.DATE_RANGE, size=16), ft.Text(_texto_rango(ayer, ayer), size=13)],
            spacing=8,
            tight=True,
        ),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        on_click=lambda _e: page.show_dialog(date_range_picker),
    )

    def _opciones_banco() -> list:
        # "Auto-detectar" + bancos habilitados. Forzar un banco es útil cuando dos
        # formatos comparten encabezados o la autodetección no basta.
        return [ft.dropdown.Option(key="", text="Auto-detectar")] + [
            ft.dropdown.Option(key=n, text=n.title()) for n in nombres_bancos()
        ]

    def _actualizar_boton() -> None:
        boton_conciliar.disabled = not archivos_banco

    def _agregar_archivo(ruta: str, nombre: str) -> None:
        dd = ft.Dropdown(value="", width=190, label="Banco", options=_opciones_banco())
        entrada: dict = {"path": ruta, "nombre": nombre, "dropdown": dd}

        def _quitar(_e) -> None:
            archivos_banco.remove(entrada)
            lista_archivos.controls.remove(entrada["fila"])
            _actualizar_boton()
            page.update()

        fila = ft.Row(
            [
                ft.Icon(ft.Icons.INSERT_DRIVE_FILE_OUTLINED, size=16, color=ft.Colors.ON_SURFACE_VARIANT),
                ft.Container(content=ft.Text(nombre, size=12, tooltip=nombre), width=260),
                dd,
                ft.IconButton(ft.Icons.DELETE_OUTLINE, icon_size=18, tooltip="Quitar archivo", on_click=_quitar),
            ],
            spacing=10,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        )
        entrada["fila"] = fila
        archivos_banco.append(entrada)
        lista_archivos.controls.append(fila)

    async def on_cargar_archivo(_e) -> None:
        # Se usa filtro ANY (no CUSTOM+allowed_extensions): en macOS el diálogo nativo
        # filtra por TIPO DE CONTENIDO (UTI), no por la extensión literal, y muchos
        # estados de cuenta .xls del portal son en realidad SpreadsheetML/HTML (o .csv
        # con tipo raro) → macOS los atenuaba aunque .xls/.csv estén permitidos. La
        # app valida el formato tras seleccionar (detección por bytes en lectura.py),
        # así que mostrar todos los archivos es seguro; los no válidos se avisan luego.
        archivos = await file_picker.pick_files(
            dialog_title="Selecciona los estados de cuenta (xlsx, xlsm, xls, xml, csv)",
            file_type=ft.FilePickerFileType.ANY,
            allow_multiple=True,
            with_data=True,
        )
        if not archivos:
            return
        for archivo in archivos:
            # Ignorar extensiones que el sistema de parsers no maneja (evita que el
            # usuario agregue un PDF/PNG por error al no haber filtro en el diálogo).
            ext = os.path.splitext(archivo.name)[1].lower().lstrip(".")
            if ext and ext not in EXTENSIONES:
                _avisar(f"«{archivo.name}»: formato .{ext} no soportado. Usa xlsx, xlsm, xls, xml o csv.")
                continue
            # En modo web: volcar bytes a un temporal CONSERVANDO la extensión (la
            # detección de algunos bancos —p. ej. BBVA .xls SpreadsheetML— depende de ella).
            if archivo.path and os.path.exists(archivo.path):
                ruta = archivo.path
            else:
                sufijo = os.path.splitext(archivo.name)[1] or ".xlsx"
                with tempfile.NamedTemporaryFile(suffix=sufijo, delete=False) as tmp:
                    tmp.write(archivo.bytes or b"")
                    ruta = tmp.name
            _agregar_archivo(ruta, archivo.name)
        _actualizar_boton()
        page.update()

    def _limpiar_archivos(_e=None) -> None:
        archivos_banco.clear()
        lista_archivos.controls.clear()
        _actualizar_boton()
        page.update()

    boton_cargar = ft.OutlinedButton(
        content=ft.Row([ft.Icon(ft.Icons.UPLOAD_FILE, size=16), ft.Text("Agregar archivos bancarios", size=13)], spacing=8, tight=True),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        on_click=on_cargar_archivo,
    )
    boton_limpiar = ft.TextButton(
        content=ft.Row([ft.Icon(ft.Icons.CLEAR_ALL, size=16), ft.Text("Limpiar todos", size=13)], spacing=6, tight=True),
        on_click=_limpiar_archivos,
    )
    boton_conciliar = ft.FilledButton(
        content=ft.Row([ft.Icon(ft.Icons.COMPARE_ARROWS, size=16), ft.Text("Conciliar", size=13)], spacing=8, tight=True),
        disabled=True,
        on_click=lambda e: page.run_task(on_conciliar, e),
    )

    def _limpiar_resultados(_e=None) -> None:
        """Vacía los paneles de resultados y el texto de estado. NO toca los archivos
        cargados (para eso está 'Limpiar todos'). `secciones` se define más abajo,
        pero solo se usa al invocar, así que la referencia ya existe entonces."""
        secciones.controls.clear()
        ultimo_resultado[0] = None
        boton_exportar_excel.disabled = True
        estado_text.value = ""
        page.update()

    boton_limpiar_resultados = ft.TextButton(
        content=ft.Row([ft.Icon(ft.Icons.DELETE_SWEEP, size=16), ft.Text("Limpiar resultados", size=13)], spacing=6, tight=True),
        on_click=_limpiar_resultados,
    )

    async def exportar_excel(_e=None) -> None:
        """Exporta la última conciliación a un .xlsx (Resumen + una hoja por sección
        con sus movimientos). El armado y el guardado van en hilos para no bloquear."""
        res = ultimo_resultado[0]
        if res is None:
            _avisar("Primero concilia para poder exportar.")
            return
        secs = _secciones_datos(res)
        generado = datetime.now().strftime("%d/%m/%Y %H:%M")
        wb = await asyncio.to_thread(_construir_workbook, res, secs, generado)

        nombre_def = f"conciliacion_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        destino = await file_picker.save_file(
            dialog_title="Guardar conciliación en Excel",
            file_name=nombre_def,
            allowed_extensions=["xlsx"],
        )
        if not destino:
            return
        if not destino.lower().endswith(".xlsx"):
            destino += ".xlsx"
        try:
            await asyncio.to_thread(wb.save, destino)
        except OSError as ex:
            _avisar(f"No se pudo guardar el Excel: {ex}")
            return
        _avisar(f"Conciliación exportada a {os.path.basename(destino)}.", error=False)

    boton_exportar_excel = ft.OutlinedButton(
        content=ft.Row([ft.Icon(ft.Icons.TABLE_VIEW, size=16), ft.Text("Exportar a Excel", size=13)], spacing=8, tight=True),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        disabled=True,
        on_click=lambda e: page.run_task(exportar_excel, e),
    )

    # --- Origen de los datos del "sistema" para comparar ------------------------
    async def on_cargar_sistema(_e) -> None:
        archivos = await file_picker.pick_files(
            dialog_title="Selecciona el reporte de Ingresos Diversos",
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx"],
            allow_multiple=False,
            with_data=True,
        )
        if not archivos:
            return
        archivo = archivos[0]
        if archivo.path and os.path.exists(archivo.path):
            archivo_sistema[0] = archivo.path
        else:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(archivo.bytes or b"")
                archivo_sistema[0] = tmp.name
        nombre_sistema.value = archivo.name
        page.update()

    boton_cargar_sistema = ft.OutlinedButton(
        content=ft.Row([ft.Icon(ft.Icons.UPLOAD_FILE, size=16), ft.Text("Cargar Ingresos Diversos", size=13)], spacing=8, tight=True),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        on_click=on_cargar_sistema,
    )

    def on_cambiar_origen(_e=None) -> None:
        es_excel = origen_group.value == "excel"
        # En modo Excel se cargan los movimientos del reporte; en modo nube se
        # consultan de BigQuery por rango de fechas.
        boton_cargar_sistema.visible = es_excel
        nombre_sistema.visible = es_excel
        boton_rango.visible = not es_excel
        page.update()

    origen_group = ft.RadioGroup(
        value="excel",  # por defecto el Excel, mientras la tabla en la nube no esté lista
        content=ft.Row(
            [
                ft.Radio(value="excel", label="Excel de Ingresos Diversos"),
                ft.Radio(value="nube", label="Datos en la nube"),
            ],
            spacing=8,
        ),
        on_change=on_cambiar_origen,
    )

    # Fila 1: archivo del banco. Fila 2: contra qué comparar. Fila 3: acción.
    # (El espaciador expand=True va en una fila SIN wrap; dentro de una fila con
    # wrap=True se renderiza como un recuadro gris enorme y descoloca los botones.)
    barra = ft.Column(
        [
            ft.Row([boton_cargar, boton_limpiar], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            lista_archivos,
            ft.Row(
                [
                    ft.Text("Comparar contra:", size=13, color=ft.Colors.ON_SURFACE_VARIANT),
                    origen_group,
                    boton_rango,
                    boton_cargar_sistema,
                    nombre_sistema,
                ],
                spacing=12,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                wrap=True,
            ),
            ft.Row(
                [progress, estado_text, ft.Container(expand=True), boton_exportar_excel, boton_limpiar_resultados, boton_conciliar],
                spacing=12,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
        ],
        spacing=10,
    )
    on_cambiar_origen()  # estado inicial de visibilidad (modo Excel)

    # --- Secciones (total + tabla unificados en un panel desplegable) -----------
    secciones = ft.Column(spacing=12)
    cuerpo = ft.Column([secciones], spacing=16)

    def _tabla(columnas: list[tuple[str, DataColumnSize | float | None]]) -> DataTable2:
        cols = []
        for etiqueta, tam in columnas:
            if isinstance(tam, (int, float)):
                cols.append(DataColumn2(ft.Text(etiqueta, weight=ft.FontWeight.BOLD), fixed_width=tam))
            else:
                cols.append(DataColumn2(ft.Text(etiqueta, weight=ft.FontWeight.BOLD), size=tam or DataColumnSize.M))
        return DataTable2(columns=cols, rows=[], min_width=700, fixed_top_rows=1, column_spacing=16, expand=True)

    def seccion_resultado(
        titulo_s: str,
        color: str,
        icono: str,
        columnas: list[tuple[str, object]],
        filas: list[list[str]],
        total_monto: float,
    ) -> ft.Control:
        """Factory reutilizable: un panel (total + tabla) por grupo. El encabezado
        muestra el conteo y el monto total; al hacer clic se despliega la tabla.

        Las 4 secciones (Conciliados, Solo Banco, Solo Sistema, Cheques) usan esta
        misma función; solo cambian columnas, filas y color."""
        # Cuerpo desplegable: la tabla, o un aviso si el grupo está vacío.
        if filas:
            tabla = _tabla(columnas)
            tabla.rows = [
                ft.DataRow(cells=[ft.DataCell(ft.Text(c, color=ft.Colors.ON_SURFACE, selectable=True)) for c in fila])
                for fila in filas
            ]
            interior = ft.Container(content=tabla, height=280, padding=ft.Padding(left=12, right=12, top=0, bottom=12))
        else:
            interior = ft.Container(
                content=ft.Text("Sin movimientos en este grupo.", italic=True, color=ft.Colors.ON_SURFACE_VARIANT),
                padding=ft.Padding(left=16, right=16, top=0, bottom=14),
            )

        # Chip con el conteo (el "total" en el que se hace clic).
        badge = ft.Container(
            content=ft.Text(str(len(filas)), size=12, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
            bgcolor=color, border_radius=12, padding=ft.Padding(left=9, right=9, top=1, bottom=1),
        )
        tile = ft.ExpansionTile(
            leading=ft.Icon(icono, color=color),
            title=ft.Row([ft.Text(titulo_s, size=14, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE), badge], spacing=10),
            subtitle=ft.Text(f"{len(filas)} movimiento(s) · {_fmt_importe(total_monto)}", size=12, color=ft.Colors.ON_SURFACE_VARIANT),
            controls=[interior],
            expanded=False,
            collapsed_bgcolor=ft.Colors.with_opacity(0.04, color),
            bgcolor=ft.Colors.with_opacity(0.04, color),
        )
        # Envoltura con borde redondeado para que cada panel se vea como tarjeta.
        return ft.Container(
            content=tile,
            border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
            border_radius=10,
            clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
        )

    def _banco(m: MovimientoConciliacion) -> str:
        # origen "BANCO:BBVA" -> "BBVA"
        return m.origen.split(":", 1)[1] if ":" in m.origen else m.origen

    def _origen_fuera(m: MovimientoConciliacion) -> str:
        """Etiqueta de origen para la sección 'Fuera de rango': lado banco -> nombre
        del banco; lado sistema -> el banco que trae el reporte (columna 'Banco'), o
        'Ingresos Diversos' si no viene."""
        if m.origen.startswith("BANCO:"):
            return f"Banco: {_banco(m)}"
        return f"Sistema: {m.raw.get('BANCO') or 'Ingresos Diversos'}"

    def _secciones_datos(res: ResultadoConciliacion) -> list[dict]:
        """Arma columnas + filas + totales por grupo. Fuente única que consumen tanto
        el render de la UI como la exportación a Excel (para no duplicar la lógica)."""
        L = DataColumnSize.L
        # Se antepone "Banco" en las tablas del lado banco (útil al combinar varios).
        cols_mov = [("Banco", 100.0), ("Fecha", 90.0), ("Descripción", L),
                    ("Referencia", 150.0), ("Importe", 120.0)]
        cols_conc = [("Banco", 100.0), ("Fecha", 90.0), ("Descripción (banco)", L),
                     ("Referencia", 150.0), ("Importe banco", 120.0), ("Importe sistema", 120.0)]
        # Repetidos: "Conciliación" (folio) + "Banco" (columna del reporte de Ingresos
        # Diversos, guardada en raw) para que el usuario sepa de qué banco es cada
        # posible duplicado y no lo confunda con el banco que subió.
        cols_repetidos = [("Conciliación", 110.0), ("Banco", 130.0), ("Fecha", 90.0),
                          ("Descripción", L), ("Referencia", 150.0), ("Importe", 120.0)]
        cols_fuera = [("Origen", 170.0), ("Fecha", 90.0), ("Descripción", L),
                      ("Referencia", 150.0), ("Importe", 120.0)]

        filas_conc = [
            [_banco(b), _fmt_fecha(b.fecha), b.descripcion, b.referencia, _fmt_importe(b.importe), _fmt_importe(s.importe)]
            for b, s in res.conciliados
        ]
        filas_banco = [[_banco(m), _fmt_fecha(m.fecha), m.descripcion, m.referencia, _fmt_importe(m.importe)] for m in res.solo_banco]
        filas_cheques = [[_banco(m), _fmt_fecha(m.fecha), m.descripcion, m.referencia, _fmt_importe(m.importe)] for m in res.devoluciones_cheque]
        filas_repetidos = [
            [str(m.raw.get("CONCILIACION") or ""), str(m.raw.get("BANCO") or ""),
             _fmt_fecha(m.fecha), m.descripcion, m.referencia, _fmt_importe(m.importe)]
            for m in res.posibles_repetidos_sistema
        ]
        filas_fuera = [[_origen_fuera(m), _fmt_fecha(m.fecha), m.descripcion, m.referencia, _fmt_importe(m.importe)] for m in res.fuera_de_rango]

        # Nota: "En sistema, no en banco" se calcula pero ya no se muestra; en su
        # lugar va "Posibles repetidos en sistema" (mismos ref+descripción+importe).
        secs = [
            {"titulo": "Movimientos conciliados", "hoja": "Conciliados", "color": _COLOR_CONCILIADOS,
             "icono": ft.Icons.CHECK_CIRCLE_OUTLINE, "columnas": cols_conc, "filas": filas_conc,
             "total": sum(b.importe for b, _ in res.conciliados)},
            {"titulo": "En banco, no en sistema", "hoja": "En banco no en sistema", "color": _COLOR_SOLO_BANCO,
             "icono": ft.Icons.ACCOUNT_BALANCE_OUTLINED, "columnas": cols_mov, "filas": filas_banco,
             "total": sum(m.importe for m in res.solo_banco)},
            {"titulo": "Posibles repetidos en sistema", "hoja": "Posibles duplicados", "color": _COLOR_REPETIDOS,
             "icono": ft.Icons.CONTENT_COPY, "columnas": cols_repetidos, "filas": filas_repetidos,
             "total": sum(m.importe for m in res.posibles_repetidos_sistema)},
            {"titulo": "Devoluciones de cheque", "hoja": "Devoluciones cheque", "color": _COLOR_CHEQUES,
             "icono": ft.Icons.MONEY_OFF, "columnas": cols_mov, "filas": filas_cheques,
             "total": sum(m.importe for m in res.devoluciones_cheque)},
        ]
        # La sección "Fuera de rango de fechas" solo aparece si hubo movimientos
        # excluidos por caer fuera de la ventana común de fechas de ambos archivos.
        if res.fuera_de_rango:
            secs.append(
                {"titulo": "Fuera de rango de fechas", "hoja": "Fuera de rango", "color": _COLOR_FUERA_RANGO,
                 "icono": ft.Icons.EVENT_BUSY, "columnas": cols_fuera, "filas": filas_fuera,
                 "total": sum(m.importe for m in res.fuera_de_rango)}
            )
        return secs

    def _render(res: ResultadoConciliacion) -> None:
        ultimo_resultado[0] = res
        secciones.controls = [
            seccion_resultado(s["titulo"], s["color"], s["icono"], s["columnas"], s["filas"], s["total"])
            for s in _secciones_datos(res)
        ]
        boton_exportar_excel.disabled = False
        page.update()

    def _avisar(mensaje: str, error: bool = True) -> None:
        snackbar.content = ft.Text(mensaje)
        snackbar.bgcolor = ft.Colors.RED_700 if error else None
        snackbar.open = True
        page.update()

    def _mostrar_error(mensaje: str) -> None:
        """Muestra el error en un diálogo FIJO y copiable, y además imprime el
        traceback completo en consola (para depurar)."""
        detalle = traceback.format_exc()
        print("\n[Conciliación] " + mensaje + "\n" + detalle)  # a consola
        cuerpo = ft.Column(
            [
                ft.Text(mensaje, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700),
                ft.SelectionArea(  # texto seleccionable/copiable
                    content=ft.Text(detalle, size=12, font_family="monospace", color=ft.Colors.ON_SURFACE),
                ),
            ],
            scroll=ft.ScrollMode.AUTO,
            tight=True,
            spacing=10,
        )
        dialogo = ft.AlertDialog(
            title=ft.Text("Error de conciliación"),
            content=ft.Container(content=cuerpo, width=620, height=360),
            actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
        )
        page.show_dialog(dialogo)

    def _mostrar_lista(titulo: str, lineas: list[str]) -> None:
        """Diálogo simple con una lista de avisos (p. ej. archivos no procesados)."""
        cuerpo = ft.Column(
            [ft.Text(l, size=13, color=ft.Colors.ON_SURFACE) for l in lineas],
            scroll=ft.ScrollMode.AUTO, tight=True, spacing=6,
        )
        dialogo = ft.AlertDialog(
            title=ft.Text(titulo),
            content=ft.Container(content=cuerpo, width=560),
            actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
        )
        page.show_dialog(dialogo)

    async def on_conciliar(_e=None) -> None:
        if not archivos_banco:
            _avisar("Agrega al menos un archivo de banco.")
            return
        # Limpiar los resultados previos ANTES de conciliar (evita mezclar tablas de
        # una corrida anterior si esta falla o cambia de archivos/origen).
        secciones.controls.clear()
        progress.visible = True
        estado_text.value = ""
        boton_conciliar.disabled = True
        page.update()
        try:
            # 1. Normalizar cada archivo con su banco (elegido o autodetectado) y
            #    combinar todos los movimientos. Los archivos con problema se listan.
            mov_banco: list[MovimientoConciliacion] = []
            problemas: list[str] = []
            resumen: list[str] = []
            for entrada in archivos_banco:
                forzado = entrada["dropdown"].value or None
                nombre, movs, estado = await asyncio.to_thread(
                    normalizar_banco, entrada["path"], forzado
                )
                if estado == "no_reconocido":
                    problemas.append(
                        f"• {entrada['nombre']}: no se reconoció el formato. Si el banco no está "
                        "en la lista, comunícate con sistemas para validarlo (o elígelo manualmente)."
                    )
                elif estado == "no_habilitado":
                    problemas.append(
                        f"• {entrada['nombre']}: parece de {nombre}, pero ese banco aún no está "
                        "habilitado. Comunícate con sistemas para validar el formato."
                    )
                else:
                    mov_banco.extend(movs)
                    resumen.append(f"{nombre} ({len(movs)})")

            if not mov_banco:
                _mostrar_lista("Ningún archivo se pudo procesar", problemas or ["Sin movimientos que conciliar."])
                return

            # 2. Traer los movimientos del sistema según el origen elegido.
            if origen_group.value == "excel":
                if not archivo_sistema[0]:
                    _avisar("Carga el Excel de Ingresos Diversos para comparar.")
                    return
                mov_sistema = await asyncio.to_thread(cargar_ingresos_diversos, archivo_sistema[0])
                origen_txt = "Ingresos Diversos (Excel)"
            else:
                fi, ff = rango_sel[0]
                crudos = await asyncio.to_thread(_repo().movimientos_crudos, fi, ff)
                mov_sistema = [MovimientoConciliacion.desde_sistema(c) for c in crudos]
                origen_txt = "nube"

            # 3. Conciliar (todos los bancos juntos) y renderizar.
            resultado = conciliar(mov_banco, mov_sistema)
            _render(resultado)
            estado_text.value = f"Bancos: {', '.join(resumen)} · Sistema ({origen_txt}): {len(mov_sistema)} mov."
            if problemas:
                _mostrar_lista("Algunos archivos no se procesaron", problemas)
        except FileNotFoundError:
            _avisar("No se encontró alguno de los archivos cargados. Vuelve a cargarlo.")
        except Exception as ex:  # noqa: BLE001 — se reporta al usuario, no debe tumbar la UI
            _mostrar_error(f"Error al conciliar: {ex}")
        finally:
            progress.visible = False
            boton_conciliar.disabled = False
            page.update()

    contenido = ft.Container(
        content=ft.Column(
            [ft.Column([titulo, subtitulo], spacing=2), barra, ft.Container(content=cuerpo, expand=True)],
            spacing=16,
            scroll=ft.ScrollMode.AUTO,
            expand=True,
        ),
        padding=20,
        expand=True,
    )

    tab = ft.Tab(label="Conciliaciones Bancarias", icon=ft.Icons.COMPARE_ARROWS)
    return tab, contenido
