import asyncio
import os
import platform
import re
import shutil
import subprocess
import tempfile
from datetime import date, datetime
from typing import Optional

import flet as ft
from flet_datatable2 import DataColumn2, DataColumnSize, DataTable2

from .catalogo import cargar_catalogo, guardar_catalogo_completo, guardar_nuevas_cuentas
from .clientes import cargar_clientes, preparar_clientes_normalizados
from .credenciales import borrar_credenciales, cargar_credenciales, guardar_credenciales
from .dashboard_cobranza import construir_tab_dashboard
from .estado_cuenta import EstadoCuenta, cargar_estado_cuenta, sugerir_sucursal_detalle
from .historial import (
    cargar_historial,
    claves_subidas,
    clave_movimiento,
    clave_movimiento_dict,
    construir_registro,
    guardar_historial,
    movimiento_desde_dict,
)
from rpa.automation import es_modo_test
from .updater import aplicar_actualizacion, reiniciar_app, revisar_actualizaciones
from .sucursales import cargar_sucursales
from .textutils import normalizar
from .empresas import EMPRESAS, EMPRESA_DEFAULT, EMPRESA_POR_CLAVE
from .matcher import extraer_cuenta, match_movimientos, match_movimientos_por_nombre
from .models import ClienteCuenta, Movimiento, OPCIONES_TIPO_MOVIMIENTO
from .ingresos_diversos import (
    aplicar_factoraje_en_sipp,
    cargar_ingresos_diversos_en_sipp,
    cargar_pagos_contado_en_sipp,
)
from .factoraje import extraer_factoraje
from .mailbox_o365 import CorreoResumen, descargar_adjuntos, listar_correos, obtener_cuenta, obtener_cuerpo
from .pagos_contado import PagoContadoExtraido, completar_con_adjunto, extraer_pago_contado
from .parsers import bbva, detectar_banco, parsear_archivo
from .rpa_folios import buscar_y_aplicar_folios, extraer_folio, extraer_folios_pendientes

EXTENSIONES_IMAGEN = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}

# Branding Grupo Petroil
NAVY = "#003C74"
GOLD = "#FBB812"
ORANGE = "#F59D00"
BORDER = "#ACABAB"
FOCUS_BORDER = "#F7F6F4"

class CustomDropdown(ft.Dropdown):
    """Dropdown solamente con borde inferior visible."""

    def __init__(self, **kwargs):
        # Estilos visuales base
        custom_styles={
            "border": ft.InputBorder.UNDERLINE,
            "border_color": BORDER,
            "focused_border_color": FOCUS_BORDER,
            "label_style": ft.TextStyle(color=ft.Colors.ON_SURFACE_VARIANT),}

        config={**custom_styles, **kwargs}
        super().__init__(**config)
        


def _parsear_monto_pago(texto: str) -> Optional[float]:
    limpio = (texto or "").replace("$", "").replace(",", "").strip()
    if not limpio:
        return None
    try:
        return float(limpio)
    except ValueError:
        return None


def es_traspaso(m: Movimiento) -> bool:
    """True si el movimiento parece un 'Traspaso a Filiales' (no es cobranza de un
    cliente). Se detecta por la palabra 'TRASPASO' en la descripción/concepto."""
    texto = f"{m.descripcion or ''} {m.concepto or ''} {m.texto_busqueda or ''}".upper()
    return "TRASPASO" in texto


def es_movimiento_portal_cliente(m: Movimiento) -> bool:
    """BBVA: los movimientos cuyo concepto empieza con 'CI' o 'CE' provienen del
    portal de clientes y YA están capturados en SIPP; subirlos por el RPA los
    duplicaría, así que se excluyen (como los traspasos)."""
    if (m.banco or "").upper() != "BBVA":
        return False
    # El archivo de movimientos EXTERNOS son cobros SPEI de otros bancos (no del
    # portal BBVA): no se excluyen aunque su leyenda traiga códigos "CI...".
    if getattr(m, "bbva_externo", False):
        return False
    texto = (m.concepto or m.descripcion or "").strip().upper()
    return bool(re.match(r"^(CI|CE)\d", texto))


# ── Exportación de movimientos a Excel ────────────────────────────────────
_EXPORT_ENCABEZADOS = [
    "Fecha", "Banco", "Descripción", "Referencia", "Abono",
    "Cliente identificado", "Cuenta", "Sucursal",
]
_EXPORT_ANCHOS = [12, 12, 48, 20, 15, 32, 16, 22]


def sucursal_export(m: Movimiento) -> str:
    """Sucursal para exportar (WYSIWYG de un snapshot): declarada > por folio >
    sugerida congelada."""
    return (
        getattr(m, "sucursal_declarada", None)
        or getattr(m, "sucursal_por_folio", None)
        or getattr(m, "sucursal_sugerida", None)
        or ""
    )


def escribir_hoja_movimientos(ws, movimientos_lista, sucursal_fn) -> None:
    """Escribe en `ws` la tabla de movimientos (encabezado + filas) con las
    columnas hasta Sucursal. `sucursal_fn(m)` resuelve la sucursal por fila."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.append(_EXPORT_ENCABEZADOS)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1B3A5B")  # NAVY
    for m in movimientos_lista:
        ws.append([
            m.fecha.strftime("%d/%m/%Y") if m.fecha else "",
            m.banco,
            m.descripcion,
            m.referencia,
            round(m.abono, 2),
            m.cliente_match or "",
            m.cuenta_match or "",
            sucursal_fn(m) or "",
        ])
    for i, w in enumerate(_EXPORT_ANCHOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for fila_celdas in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for celda in fila_celdas:
            celda.number_format = "#,##0.00"
    ws.freeze_panes = "A2"


def nombre_hoja_excel(base: str, usados: set) -> str:
    """Nombre de hoja válido (≤31 chars, sin \\ / * ? : [ ]) y único."""
    limpio = re.sub(r"[\\/*?:\[\]]", "", base or "").strip()[:28] or "Hoja"
    nombre, i = limpio, 2
    while nombre in usados:
        nombre = f"{limpio[:25]}_{i}"
        i += 1
    usados.add(nombre)
    return nombre


def abrir_archivo_con_app_predeterminada(ruta: str) -> None:
    sistema = platform.system()
    if sistema == "Darwin":
        subprocess.run(["open", ruta], check=False)
    elif sistema == "Windows":
        os.startfile(ruta)  # type: ignore[attr-defined]
    else:
        subprocess.run(["xdg-open", ruta], check=False)


def revelar_en_explorador(ruta: str) -> None:
    """Abre el explorador de archivos resaltando `ruta`."""
    sistema = platform.system()
    try:
        if sistema == "Darwin":
            subprocess.run(["open", "-R", ruta], check=False)
        elif sistema == "Windows":
            subprocess.run(["explorer", "/select,", ruta], check=False)
        else:
            subprocess.run(["xdg-open", os.path.dirname(ruta)], check=False)
    except Exception:
        pass

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HISTORIAL_PATH = os.path.join(BASE_DIR, "historial_extracciones.json")
CATALOGO_PATH = os.path.join(BASE_DIR, "Catalogos", "Cuentas_Clientes", "Catalogo_Cuentas_Clientes.csv")
CLIENTES_PATH = os.path.join(BASE_DIR, "Catalogos", "Cuentas_Clientes", "Clientes.csv")

FILTRO_TODOS = "Todos"
FILTRO_IDENTIFICADOS = "Identificados"
FILTRO_NO_IDENTIFICADOS = "No identificados"
FILTRO_EXCLUIDOS = "Excluidos"
FILTRO_SUCURSAL_TODAS = "Todas las sucursales"
FILTRO_SUCURSAL_SIN = "No identificado"  # movimientos sin sucursal aún



def main(page: ft.Page) -> None:
    page.title = "MultiHerramienta de Cobranza · Grupo Petroil"
    page.window.width = 1280
    page.window.height = 860
    page.window.icon = "Logo_Petroil.ico"
    page.padding = 0
    page.theme_mode = ft.ThemeMode.LIGHT
    page.theme = ft.Theme(color_scheme_seed=NAVY, use_material3=True)
    page.dark_theme = ft.Theme(color_scheme_seed=NAVY, use_material3=True)

    def mostrar_dialogo(dialogo: ft.AlertDialog) -> None:
        """Muestra el diálogo de forma idempotente y a prueba de estados colgados.

        pop_dialog() marca open=False pero NO retira el diálogo de la pila
        interna de Flet hasta que Flutter reporta el evento de dismiss. Tras un
        RPA pesado (Playwright) ese evento puede no llegar limpio y el diálogo
        queda colgado: en la pila pero con open=False. Entonces el guard
        `if not dialogo.open` pasa, show_dialog() lo encuentra todavía en la
        pila y truena con 'Dialog is already opened' (excepción tragada en el
        on_click) → no aparece ningún modal. Lo limpiamos a mano antes de
        reabrir."""
        pila = page._dialogs.controls
        if dialogo in pila:
            if dialogo.open:
                return  # ya está realmente abierto
            pila.remove(dialogo)  # colgado: retirarlo para poder reabrir
        page.show_dialog(dialogo)

    # Los RPA corren en modo silencioso (headless) por defecto: sin ventana de
    # navegador. El usuario puede activar "Ver navegador" para depurar.
    ver_navegador_ref = [False]

    def crear_panel_log_rpa(altura: int = 190):
        """Crea un panel de log flotante para un flujo RPA: un ListView que va
        acumulando el progreso en vivo. Devuelve (panel, log_fn, reset)."""
        lista = ft.ListView(spacing=1, auto_scroll=True, expand=True)
        panel = ft.Container(
            content=lista,
            height=altura,
            border_radius=8,
            padding=8,
            bgcolor=ft.Colors.with_opacity(0.05, ft.Colors.ON_SURFACE),
            visible=False,
        )
        colores = {
            "error": ft.Colors.RED_400,
            "warn": ft.Colors.ORANGE_700,
            "ok": ft.Colors.GREEN_700,
        }

        def log_fn(mensaje: str, nivel: str = "info") -> None:
            panel.visible = True
            lista.controls.append(
                ft.Text(mensaje, size=12, selectable=True,
                        color=colores.get(nivel, ft.Colors.ON_SURFACE))
            )
            if len(lista.controls) > 400:  # tope para no crecer sin límite
                del lista.controls[0]
            page.update()

        def reset() -> None:
            lista.controls.clear()
            panel.visible = False

        return panel, log_fn, reset

    def crear_check_ver_navegador() -> ft.Checkbox:
        return ft.Checkbox(
            label="Ver navegador (depuración)",
            value=ver_navegador_ref[0],
            on_change=lambda e: ver_navegador_ref.__setitem__(0, bool(e.control.value)),
        )

    # Botón flotante para dejar un RPA en segundo plano: al cerrar el modal de
    # avance, el FAB permite reabrirlo para ver el progreso/resumen.
    dialogo_rpa_ref: list = [None]
    rpa_corriendo_ref = [False]

    def _fab_rpa_click(_e) -> None:
        if dialogo_rpa_ref[0] is not None:
            mostrar_dialogo(dialogo_rpa_ref[0])
        if not rpa_corriendo_ref[0]:
            # Ya terminó: al abrir el resumen, se retira el botón flotante.
            fab_rpa.visible = False
            page.update()

    fab_rpa = ft.FloatingActionButton(
        icon=ft.Icons.SYNC, visible=False, bgcolor=ORANGE, on_click=_fab_rpa_click,
    )

    def rpa_inicio(dialogo) -> None:
        """Marca que un RPA está corriendo en segundo plano (muestra el FAB)."""
        dialogo_rpa_ref[0] = dialogo
        rpa_corriendo_ref[0] = True
        fab_rpa.icon = ft.Icons.SYNC
        fab_rpa.bgcolor = ORANGE
        fab_rpa.tooltip = "RPA en curso — clic para ver el avance"
        fab_rpa.visible = True
        page.update()

    def rpa_fin(ok: bool = True) -> None:
        """Marca que el RPA terminó (FAB verde/rojo, clic para ver el resumen)."""
        rpa_corriendo_ref[0] = False
        fab_rpa.icon = ft.Icons.CHECK_CIRCLE if ok else ft.Icons.ERROR
        fab_rpa.bgcolor = ft.Colors.GREEN_600 if ok else ft.Colors.RED_600
        fab_rpa.tooltip = (
            "RPA terminado — clic para ver el resumen" if ok
            else "RPA con error — clic para ver el detalle"
        )
        fab_rpa.visible = True
        page.update()

    catalogo = cargar_catalogo(CATALOGO_PATH)
    clientes_normalizados = preparar_clientes_normalizados(cargar_clientes(CLIENTES_PATH))
    sucursales_catalogo = cargar_sucursales()
    nombres_clientes = sorted({original for original, _ in clientes_normalizados})
    sugerencias_clientes = [ft.AutoCompleteSuggestion(key=nombre, value=nombre) for nombre in nombres_clientes]
    movimientos: list[Movimiento] = []

    # --- Controles de estado / resumen ---
    archivo_nombre_text = ft.Text("Ningún archivo cargado", italic=True, color=ft.Colors.ON_SURFACE_VARIANT)
    banco_detectado_text = ft.Text("")
    catalogo_info_text = ft.Text(
        f"Catálogo de clientes cargado: {len(catalogo)} cuentas",
        color=ft.Colors.WHITE_70,
        size=12,
    )

    card_total = ft.Text("0", size=28, weight=ft.FontWeight.BOLD, color=NAVY)
    card_identificados = ft.Text("0", size=28, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700)
    card_no_identificados = ft.Text("0", size=28, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700)
    card_porcentaje = ft.Text("0%", size=28, weight=ft.FontWeight.BOLD, color=ORANGE)
    # Cargos fuera de alcance por ahora (solo se procesan abonos). Se deja
    # comentado por si en el futuro se requiere mostrar también los cargos.
    # card_total_cargos = ft.Text("$0.00", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700)
    card_total_abonos = ft.Text("$0.00", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700, no_wrap=True)

    def resumen_card(titulo: str, valor_control: ft.Text) -> ft.Container:
        return ft.Container(
            content=ft.Column(
                [ft.Text(titulo, size=12, color=ft.Colors.ON_SURFACE_VARIANT), valor_control],
                spacing=4,
            ),
            padding=16,
            height=88,  # altura fija: todas las tarjetas iguales sin importar la fuente
            bgcolor=ft.Colors.SURFACE_CONTAINER_LOW,
            border=ft.Border(
                top=ft.BorderSide(3, GOLD),
                left=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
                right=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
                bottom=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
            ),
            border_radius=8,
            expand=True,
        )

    resumen_row = ft.Row(
        [
            resumen_card("Movimientos totales", card_total),
            resumen_card("Identificados", card_identificados),
            resumen_card("No identificados", card_no_identificados),
            resumen_card("% Identificado", card_porcentaje),
            # resumen_card("Total cargos", card_total_cargos),
            resumen_card("Total abonos", card_total_abonos),
        ],
        spacing=12,
    )

    def on_filtro_cambio(_e) -> None:
        refrescar_tabla()

    # --- Filtros ---
    filtro_estado = CustomDropdown(
        label="Estado",
        value=FILTRO_TODOS,
        options=[ft.dropdown.Option(o) for o in (FILTRO_TODOS, FILTRO_IDENTIFICADOS, FILTRO_NO_IDENTIFICADOS, FILTRO_EXCLUIDOS)],
        width=200,
        color=ft.Colors.ON_SURFACE,
        on_select=on_filtro_cambio,
    )
    filtro_sucursal = CustomDropdown(
        label="Sucursal",
        value=FILTRO_SUCURSAL_TODAS,
        options=[ft.dropdown.Option(FILTRO_SUCURSAL_TODAS)],
        width=220,
        color=ft.Colors.ON_SURFACE,
        editable=True,
        enable_filter=True,
        menu_height=320,
        on_select=on_filtro_cambio,
    )
    filtro_texto = ft.TextField(
        label="Buscar cliente, cuenta o descripción",
        width=400,
        color=ft.Colors.ON_SURFACE,
        on_change=on_filtro_cambio,
        on_submit=on_filtro_cambio,
        on_blur=on_filtro_cambio,
        border=ft.InputBorder.UNDERLINE,
        border_color=BORDER
    )

    def columna(
        texto: str,
        numeric: bool = False,
        fixed_width: Optional[float] = None,
        size: Optional[DataColumnSize] = None,
    ) -> DataColumn2:
        return DataColumn2(
            ft.Text(texto, weight=ft.FontWeight.BOLD, color=ft.Colors.ON_SURFACE),
            numeric=numeric,
            fixed_width=fixed_width,
            size=size,
        )

    # Columnas angostas con ancho fijo; "Descripción" y "Cliente identificado"
    # son flexibles (size=L) y se reparten el espacio restante de la ventana.
    tabla = DataTable2(
        columns=[
            columna("Fecha", fixed_width=90),
            columna("Banco", fixed_width=90),
            columna("Descripción", size=DataColumnSize.L),
            columna("Referencia", fixed_width=150),
            columna("Abono", numeric=True, fixed_width=110),
            columna("Cliente identificado", size=DataColumnSize.L),
            columna("Cuenta", fixed_width=110),
            columna("Sucursal sugerida", fixed_width=180),
            columna("Estado", fixed_width=150),
            columna("Acciones", fixed_width=180),
        ],
        rows=[],
        min_width=1050,
        fixed_top_rows=1,
        column_spacing=16,
        data_row_height=76,  # más alto para la descripción multilínea
        expand=True,
    )

    tabla_contenedor = ft.Container(content=tabla, expand=True)

    def estado_badge(m: Movimiento) -> ft.Container:
        if m.excluido:
            if es_movimiento_portal_cliente(m):
                etiqueta = "Excluido (portal)"
                tip = "BBVA portal de clientes (CI/CE): ya está en SIPP; no se sube por el RPA."
            elif es_traspaso(m):
                etiqueta = "Excluido (traspaso)"
                tip = "Traspaso a filiales: no es cobranza; no se sube por el RPA."
            else:
                etiqueta = "Excluido"
                tip = "Movimiento excluido del RPA (no se sube a SIPP)."
            return ft.Container(
                content=ft.Text(etiqueta, color=ft.Colors.WHITE, size=11, no_wrap=True),
                bgcolor=ft.Colors.RED_400,
                padding=ft.Padding.symmetric(horizontal=8, vertical=4),
                border_radius=12,
                tooltip=tip,
            )
        if m.ya_subido:
            return ft.Container(
                content=ft.Text("Ya extraído", color=ft.Colors.WHITE, size=11, no_wrap=True),
                bgcolor=ft.Colors.BLUE_GREY_400,
                padding=ft.Padding.symmetric(horizontal=8, vertical=4),
                border_radius=12,
                tooltip="Ya venía en una extracción previa (corte anterior); no se volverá a subir a SIPP.",
            )
        if m.identificado_manual:
            return ft.Container(
                content=ft.Text("Identificado (manual)", color=ft.Colors.WHITE, size=11, no_wrap=True),
                bgcolor=ft.Colors.TEAL_600,
                padding=ft.Padding.symmetric(horizontal=8, vertical=4),
                border_radius=12,
            )
        if m.identificado_por_folio:
            return ft.Container(
                content=ft.Text("Identificado (folio SIPP)", color=ft.Colors.WHITE, size=11, no_wrap=True),
                bgcolor=ft.Colors.PURPLE_600,
                padding=ft.Padding.symmetric(horizontal=8, vertical=4),
                border_radius=12,
            )
        if m.identificado_por_nombre:
            return ft.Container(
                content=ft.Text("Identificado (nombre)", color=ft.Colors.WHITE, size=11, no_wrap=True),
                bgcolor=ft.Colors.BLUE_600,
                padding=ft.Padding.symmetric(horizontal=8, vertical=4),
                border_radius=12,
            )
        if m.identificado:
            return ft.Container(
                content=ft.Text("Identificado", color=ft.Colors.WHITE, size=11, no_wrap=True),
                bgcolor=ft.Colors.GREEN_600,
                padding=ft.Padding.symmetric(horizontal=8, vertical=4),
                border_radius=12,
            )
        return ft.Container(
            content=ft.Text("No identificado", color=ft.Colors.WHITE, size=11, no_wrap=True),
            bgcolor=ft.Colors.RED_600,
            padding=ft.Padding.symmetric(horizontal=8, vertical=4),
            border_radius=12,
        )

    # --- Identificación manual ---
    movimiento_a_identificar: list[Movimiento | None] = [None]

    manual_info_text = ft.Text("")
    manual_autocomplete = ft.AutoComplete(suggestions=sugerencias_clientes)

    def on_select_cliente_manual(e: ft.AutoCompleteSelectEvent) -> None:
        mov = movimiento_a_identificar[0]
        if mov is None:
            return
        cliente = e.selection.value

        mov.cliente_match = cliente
        mov.identificado_manual = True
        # Al confirmar manualmente, deja de ser una sugerencia por nombre.
        mov.identificado_por_nombre = False

        cuenta_extraida = extraer_cuenta(mov.texto_busqueda)
        agregadas = []
        if cuenta_extraida:
            mov.cuenta_match = cuenta_extraida
            propuesta = [ClienteCuenta(cuenta=cuenta_extraida, cliente=cliente, banco=mov.banco, plaza="")]
            agregadas = guardar_nuevas_cuentas(CATALOGO_PATH, catalogo, propuesta)
            catalogo.extend(agregadas)
            catalogo_info_text.value = f"Catálogo de clientes cargado: {len(catalogo)} cuentas"

        mensaje = f"'{cliente}' asignado manualmente al movimiento."
        if agregadas:
            mensaje += " Se agregó la cuenta al catálogo."
        estado_text.value = mensaje

        page.pop_dialog()
        refrescar_resumen()
        refrescar_tabla()
        historial_guardar_snapshot()

    manual_autocomplete.on_select = on_select_cliente_manual

    def on_cancelar_manual(_e) -> None:
        page.pop_dialog()

    dialogo_manual = ft.AlertDialog(
        modal=True,
        title=ft.Text("Identificar cliente manualmente"),
        content=ft.Column(
            [manual_info_text, manual_autocomplete],
            tight=True,
            spacing=12,
            width=400,
        ),
        actions=[ft.TextButton("Cancelar", on_click=on_cancelar_manual)],
    )

    def abrir_dialogo_manual(mov: Movimiento) -> None:
        movimiento_a_identificar[0] = mov
        manual_info_text.value = f"{mov.descripcion}\nReferencia: {mov.referencia} · Abono: ${mov.abono:,.2f}"
        manual_autocomplete.value = ""
        mostrar_dialogo(dialogo_manual)

    def boton_identificar_manual(m: Movimiento) -> ft.Control:
        if not m.identificado:
            return ft.IconButton(
                icon=ft.Icons.PERSON_ADD,
                tooltip="Identificar cliente manualmente",
                icon_color=NAVY,
                on_click=lambda _e, mov=m: abrir_dialogo_manual(mov),
            )
        # Identificado por NOMBRE: permitir corregir un posible falso positivo.
        if m.identificado_por_nombre and not m.identificado_manual:
            return ft.IconButton(
                icon=ft.Icons.EDIT,
                tooltip="Cambiar cliente (identificado por nombre)",
                icon_color=NAVY,
                on_click=lambda _e, mov=m: abrir_dialogo_manual(mov),
            )
        # Identificado por cuenta/folio o ya confirmado manualmente: no editable aquí.
        return ft.Text("-", color=ft.Colors.ON_SURFACE_VARIANT)

    # --- Declaración manual de folio/texto a buscar en SIPP ---
    movimiento_a_declarar_folio: list[Movimiento | None] = [None]

    folio_manual_info_text = ft.Text("")
    folio_manual_field = ft.TextField(label="Folio o texto a buscar en SIPP")

    def on_confirmar_folio_manual(_e) -> None:
        mov = movimiento_a_declarar_folio[0]
        if mov is None:
            return
        mov.folio_manual = (folio_manual_field.value or "").strip() or None
        page.pop_dialog()
        refrescar_tabla()
        historial_guardar_snapshot()

    def on_cancelar_folio_manual(_e) -> None:
        page.pop_dialog()

    dialogo_folio_manual = ft.AlertDialog(
        modal=True,
        title=ft.Text("Declarar folio/texto para SIPP"),
        content=ft.Column(
            [folio_manual_info_text, folio_manual_field],
            tight=True,
            spacing=12,
            width=400,
        ),
        actions=[
            ft.TextButton("Cancelar", on_click=on_cancelar_folio_manual),
            ft.Button("Guardar", on_click=on_confirmar_folio_manual, bgcolor=NAVY, color=ft.Colors.WHITE),
        ],
    )

    def abrir_dialogo_folio_manual(mov: Movimiento) -> None:
        movimiento_a_declarar_folio[0] = mov
        folio_manual_info_text.value = f"{mov.descripcion}\nReferencia: {mov.referencia} · Abono: ${mov.abono:,.2f}"
        folio_manual_field.value = mov.folio_manual or ""
        mostrar_dialogo(dialogo_folio_manual)

    def boton_declarar_folio(m: Movimiento) -> ft.Control:
        if m.identificado:
            return ft.Container(width=0)
        return ft.IconButton(
            icon=ft.Icons.TRAVEL_EXPLORE,
            tooltip=f"Folio declarado: {m.folio_manual}" if m.folio_manual else "Declarar folio/texto a buscar en SIPP",
            icon_color=ORANGE if m.folio_manual else ft.Colors.ON_SURFACE_VARIANT,
            on_click=lambda _e, mov=m: abrir_dialogo_folio_manual(mov),
        )

    def toggle_excluir(m: Movimiento) -> None:
        """Alterna si el movimiento se excluye del RPA que sube a SIPP."""
        m.excluido = not m.excluido
        estado_text.value = (
            f"Movimiento excluido del RPA (no se subirá a SIPP): {m.referencia}."
            if m.excluido
            else f"Movimiento reincorporado al RPA: {m.referencia}."
        )
        refrescar_resumen()
        refrescar_tabla()
        historial_guardar_snapshot()

    def boton_excluir(m: Movimiento) -> ft.Control:
        """Botón para excluir/reincluir un movimiento del RPA. Siempre visible; se
        resalta cuando el movimiento es un traspaso (detectado) o ya está excluido."""
        if m.excluido:
            return ft.IconButton(
                icon=ft.Icons.UNDO,
                tooltip="Reincorporar al RPA (volver a subir a SIPP)",
                icon_color=ft.Colors.RED_600,
                on_click=lambda _e, mov=m: toggle_excluir(mov),
            )
        return ft.IconButton(
            icon=ft.Icons.BLOCK,
            tooltip=(
                "Excluir del RPA: es un traspaso, no se subirá a SIPP"
                if es_traspaso(m)
                else "Excluir este movimiento del RPA (no subir a SIPP)"
            ),
            icon_color=ft.Colors.RED_400 if es_traspaso(m) else ft.Colors.ON_SURFACE_VARIANT,
            on_click=lambda _e, mov=m: toggle_excluir(mov),
        )

    # --- Tipo de movimiento (checkboxes "¿Es ...?" del modal de SIPP) ---
    movimiento_a_tipificar: list[Movimiento | None] = [None]
    tipo_checkboxes: dict[str, ft.Checkbox] = {
        etiqueta: ft.Checkbox(label=f"¿Es {etiqueta}?") for etiqueta in OPCIONES_TIPO_MOVIMIENTO
    }
    tipo_info_text = ft.Text("")

    def on_guardar_tipos(_e) -> None:
        mov = movimiento_a_tipificar[0]
        if mov is None:
            return
        mov.tipos_movimiento = [
            etiqueta for etiqueta, chk in tipo_checkboxes.items() if chk.value
        ]
        page.pop_dialog()
        refrescar_tabla()
        historial_guardar_snapshot()

    dialogo_tipos = ft.AlertDialog(
        modal=True,
        title=ft.Text("Tipo de movimiento"),
        content=ft.Column(
            [
                tipo_info_text,
                ft.Text(
                    "Marca lo que aplique. Se usará al capturar en SIPP por el '+' "
                    "(bancos sin 'Subir Excel', ej. BanBajío). Vacío = Ingreso Diverso.",
                    size=11, color=ft.Colors.ON_SURFACE_VARIANT,
                ),
                *tipo_checkboxes.values(),
            ],
            tight=True, spacing=8, width=380, scroll=ft.ScrollMode.AUTO,
        ),
        actions=[
            ft.TextButton("Cancelar", on_click=lambda _e: page.pop_dialog()),
            ft.Button("Guardar", on_click=on_guardar_tipos, bgcolor=NAVY, color=ft.Colors.WHITE),
        ],
    )

    def abrir_dialogo_tipos(mov: Movimiento) -> None:
        movimiento_a_tipificar[0] = mov
        seleccion = set(getattr(mov, "tipos_movimiento", []) or [])
        for etiqueta, chk in tipo_checkboxes.items():
            chk.value = etiqueta in seleccion
        tipo_info_text.value = f"{mov.descripcion[:70]}\nAbono: ${mov.abono:,.2f}"
        mostrar_dialogo(dialogo_tipos)

    def boton_tipo(m: Movimiento) -> ft.Control:
        tipos = getattr(m, "tipos_movimiento", []) or []
        return ft.IconButton(
            icon=ft.Icons.LABEL if tipos else ft.Icons.LABEL_OUTLINE,
            tooltip=("Tipo: " + ", ".join(tipos)) if tipos else "Marcar tipo de movimiento (¿Es Contado?, ...)",
            icon_color=ORANGE if tipos else ft.Colors.ON_SURFACE_VARIANT,
            on_click=lambda _e, mov=m: abrir_dialogo_tipos(mov),
        )

    def aplicar_filtros() -> list[Movimiento]:
        resultado = movimientos
        if filtro_estado.value == FILTRO_IDENTIFICADOS:
            resultado = [m for m in resultado if m.identificado and not m.excluido]
        elif filtro_estado.value == FILTRO_NO_IDENTIFICADOS:
            resultado = [m for m in resultado if not m.identificado and not m.excluido]
        elif filtro_estado.value == FILTRO_EXCLUIDOS:
            resultado = [m for m in resultado if m.excluido]

        suc = filtro_sucursal.value
        if suc == FILTRO_SUCURSAL_SIN:
            resultado = [m for m in resultado if not sucursal_efectiva(m)]
        elif suc and suc != FILTRO_SUCURSAL_TODAS:
            resultado = [m for m in resultado if sucursal_efectiva(m) == suc]

        texto = (filtro_texto.value or "").strip().lower()
        if texto:
            resultado = [
                m
                for m in resultado
                if texto in (m.cliente_match or "").lower()
                or texto in (m.cuenta_match or "").lower()
                or texto in m.descripcion.lower()
                or texto in m.referencia.lower()
            ]
        return resultado

    def celda(texto: str) -> ft.DataCell:
        return ft.DataCell(ft.Text(texto, color=ft.Colors.ON_SURFACE))

    # Diálogo para ver la descripción completa (seleccionable/copiable).
    descripcion_dialog_text = ft.Text("", selectable=True, color=ft.Colors.ON_SURFACE)
    dialogo_descripcion = ft.AlertDialog(
        modal=True,
        title=ft.Text("Descripción del movimiento"),
        content=ft.Container(
            content=ft.Column([descripcion_dialog_text], scroll=ft.ScrollMode.AUTO, tight=True),
            width=580,
            height=260,
        ),
        actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
    )

    def abrir_dialogo_descripcion(texto: str) -> None:
        descripcion_dialog_text.value = texto or "-"
        mostrar_dialogo(dialogo_descripcion)

    def celda_desc(texto: str) -> ft.DataCell:
        """Celda de descripción: muestra hasta 3 líneas; clic abre el texto
        completo (las descripciones bancarias son largas)."""
        return ft.DataCell(
            ft.Text(
                texto,
                color=ft.Colors.ON_SURFACE,
                max_lines=3,
                overflow=ft.TextOverflow.ELLIPSIS,
                tooltip="Clic para ver la descripción completa",
            ),
            on_tap=lambda _e, d=texto: abrir_dialogo_descripcion(d),
        )

    # Cache de la sugerencia por (movimiento, cliente, abono): refrescar_tabla se
    # llama en cada tecla de los filtros y no queremos recalcular el subset-sum.
    sucursal_cache: dict = {}

    def sucursal_efectiva(m: Movimiento) -> Optional[str]:
        """Sucursal resuelta del movimiento para mostrar/filtrar: declarada por el
        usuario, leída por folio o sugerida por el estado de cuenta (en vivo o
        congelada). None si aún no tiene sucursal."""
        if not m.identificado:
            return None
        if m.sucursal_declarada:
            return m.sucursal_declarada
        if m.sucursal_por_folio:
            return m.sucursal_por_folio
        estado = estado_cuenta_ref[0]
        if estado is None:
            return m.sucursal_sugerida  # congelada del historial (o None)
        empresa = empresa_ref[0]
        clave = (id(m), m.cliente_match, m.abono, empresa.clave)
        if clave in sucursal_cache:
            detalle = sucursal_cache[clave]
        else:
            detalle = sugerir_sucursal_detalle(estado, m.cliente_match, m.abono, empresa.nombre_reporte)
            sucursal_cache[clave] = detalle
        return detalle[0] if detalle else None

    def _boton_editar_sucursal(m: Movimiento) -> ft.Control:
        return ft.IconButton(
            icon=ft.Icons.EDIT,
            tooltip="Declarar/cambiar sucursal",
            icon_color=NAVY,
            icon_size=16,
            on_click=lambda _e, mov=m: abrir_dialogo_sucursal(mov),
        )

    def celda_sucursal_sugerida(m: Movimiento) -> ft.DataCell:
        """Celda de sucursal: declarada por el usuario (override), leída de la
        factura (identificación por folio) o sugerida por el estado de cuenta.
        Vacía si el movimiento no está identificado."""
        estado = estado_cuenta_ref[0]
        if not m.identificado:
            return ft.DataCell(ft.Text("", color=ft.Colors.ON_SURFACE_VARIANT))

        # 1) Override declarado por el usuario tiene prioridad.
        if m.sucursal_declarada:
            texto = m.sucursal_declarada
            return ft.DataCell(
                ft.Row(
                    [
                        ft.Text(texto, color=ft.Colors.BLUE_700, weight=ft.FontWeight.BOLD,
                                tooltip="Declarada por el usuario"),
                        _boton_editar_sucursal(m),
                    ],
                    spacing=0,
                    tight=True,
                )
            )

        # 2) Leída de la propia factura durante la búsqueda por folio (confiable).
        if m.sucursal_por_folio:
            return ft.DataCell(
                ft.Row(
                    [
                        ft.Text(m.sucursal_por_folio, color=ft.Colors.GREEN_700,
                                weight=ft.FontWeight.BOLD, tooltip="Tomada de la factura (por folio)"),
                        _boton_editar_sucursal(m),
                    ],
                    spacing=0,
                    tight=True,
                )
            )

        # 3) Sugerida por el estado de cuenta (cacheada), filtrada por empresa.
        if estado is None:
            # Sin .xlsx cargado (ej. extracción restaurada del historial): se
            # muestra la sugerencia CONGELADA al guardar el snapshot, si existe.
            if m.sucursal_sugerida:
                motivo = m.sucursal_sugerida_motivo
                texto = m.sucursal_sugerida if motivo in (None, "única") else f"{m.sucursal_sugerida} ({motivo})"
                return ft.DataCell(
                    ft.Text(texto, color=ft.Colors.ON_SURFACE,
                            tooltip="Sugerida (guardada en el historial)")
                )
            return ft.DataCell(ft.Text("", color=ft.Colors.ON_SURFACE_VARIANT))
        empresa = empresa_ref[0]
        clave = (id(m), m.cliente_match, m.abono, empresa.clave)
        if clave in sucursal_cache:
            detalle = sucursal_cache[clave]
        else:
            detalle = sugerir_sucursal_detalle(estado, m.cliente_match, m.abono, empresa.nombre_reporte)
            sucursal_cache[clave] = detalle
        if detalle is None:
            # Cliente sin facturas de esta empresa en el estado de cuenta.
            return ft.DataCell(
                ft.Row(
                    [
                        ft.Text("— (no en edo. cuenta)", italic=True,
                                color=ft.Colors.ON_SURFACE_VARIANT),
                        _boton_editar_sucursal(m),
                    ],
                    spacing=0,
                    tight=True,
                )
            )
        sucursal, motivo, todas = detalle
        if sucursal is None:
            texto = "?"
        elif motivo == "única":
            texto = sucursal
        else:
            texto = f"{sucursal} ({motivo})"
        tooltip = None
        color = ft.Colors.ON_SURFACE
        if len(todas) > 1:
            tooltip = "Sucursales del cliente: " + ", ".join(todas)
            if motivo in ("aproximado", None):
                color = ft.Colors.ORANGE_700  # match débil: revisar
        return ft.DataCell(
            ft.Row(
                [ft.Text(texto, color=color, tooltip=tooltip), _boton_editar_sucursal(m)],
                spacing=0,
                tight=True,
            )
        )

    def tiene_folio_pendiente(m: Movimiento) -> bool:
        """El movimiento NO está identificado pero SÍ trae un folio (detectado o
        declarado por el usuario), así que puede identificarse ejecutando la
        búsqueda de folios en SIPP."""
        if m.identificado:
            return False
        return bool(m.folio_manual) or bool(extraer_folio(m.texto_busqueda))

    def color_fila(m: Movimiento):
        # Movimientos excluidos del RPA (traspasos a filiales, etc.): rojo tenue.
        # Máxima prioridad visual: el usuario decidió que no se suban.
        if m.excluido:
            return ft.Colors.with_opacity(0.14, ft.Colors.RED)
        # Movimientos ya subidos a SIPP en una extracción previa: en gris tenue
        # (tiene prioridad; ya no requieren acción).
        if m.ya_subido:
            return ft.Colors.with_opacity(0.06, ft.Colors.ON_SURFACE)
        # Resalta las filas no identificadas que tienen folio pendiente de buscar
        # en SIPP, para señalar el siguiente paso al usuario.
        if tiene_folio_pendiente(m):
            return ft.Colors.with_opacity(0.14, ft.Colors.AMBER)
        return None

    _sucursales_filtro_actual: list[frozenset] = [frozenset()]

    def actualizar_opciones_sucursal() -> None:
        """Reconstruye las opciones del filtro de sucursal a partir de las
        sucursales presentes en los movimientos. Preserva la selección vigente."""
        presentes = {s for m in movimientos if (s := sucursal_efectiva(m))}
        if frozenset(presentes) == _sucursales_filtro_actual[0]:
            return  # sin cambios: evitar reconstrucciones innecesarias
        _sucursales_filtro_actual[0] = frozenset(presentes)
        opciones = [FILTRO_SUCURSAL_TODAS] + sorted(presentes) + [FILTRO_SUCURSAL_SIN]
        filtro_sucursal.options = [ft.dropdown.Option(o) for o in opciones]
        if filtro_sucursal.value not in opciones:
            filtro_sucursal.value = FILTRO_SUCURSAL_TODAS

    def refrescar_tabla() -> None:
        actualizar_opciones_sucursal()
        filas = aplicar_filtros()
        tabla.rows = [
            ft.DataRow(
                color=color_fila(m),
                cells=[
                    celda(m.fecha.strftime("%d/%m/%Y") if m.fecha else "-"),
                    celda(m.banco),
                    celda_desc(m.descripcion),
                    celda(m.referencia[:30]),
                    # celda(f"${m.cargo:,.2f}" if m.cargo else ""),
                    celda(f"${m.abono:,.2f}" if m.abono else ""),
                    celda(m.cliente_match or "-"),
                    celda(m.cuenta_match or "-"),
                    celda_sucursal_sugerida(m),
                    ft.DataCell(estado_badge(m)),
                    ft.DataCell(ft.Row([boton_identificar_manual(m), boton_declarar_folio(m), boton_excluir(m), boton_tipo(m)], spacing=0)),
                ],
            )
            for m in filas
        ]
        page.update()

    def refrescar_resumen() -> None:
        total = len(movimientos)
        identificados = sum(1 for m in movimientos if m.identificado)
        no_identificados = total - identificados
        porcentaje = (identificados / total * 100) if total else 0
        # total_cargos = sum(m.cargo for m in movimientos)
        total_abonos = sum(m.abono for m in movimientos)

        card_total.value = str(total)
        card_identificados.value = str(identificados)
        card_no_identificados.value = str(no_identificados)
        card_porcentaje.value = f"{porcentaje:.1f}%"
        # card_total_cargos.value = f"${total_cargos:,.2f}"
        card_total_abonos.value = f"${total_abonos:,.2f}"
        page.update()


    estado_text = ft.Text("")
    ultima_ruta_csv: list[Optional[str]] = [None]

    # --- Historial local de extracciones del día ---
    historial_registros: list[dict] = cargar_historial(HISTORIAL_PATH)
    historial_id_actual: list[Optional[str]] = [None]
    historial_archivo_actual: list[str] = [""]
    historial_seleccionados: set = set()  # ids marcados para exportar a Excel

    def historial_guardar_snapshot(nuevo: bool = False) -> None:
        """Crea (nuevo=True, al cargar un CSV) o actualiza el registro actual con
        el estado vigente de los movimientos y lo persiste. Al actualizar, guarda
        los últimos cambios de cliente/sucursal/folio."""
        if not movimientos:
            return
        ahora = datetime.now()
        banco = movimientos[0].banco if movimientos else ""
        empresa_clave = empresa_ref[0].clave

        # Congelar la sucursal sugerida por estado de cuenta (si está cargado),
        # para que al restaurar la extracción se vea tal cual la dejó el usuario
        # sin necesidad de recargar el .xlsx.
        estado_ec = estado_cuenta_ref[0]
        if estado_ec is not None:
            empresa_ec = empresa_ref[0]
            for m in movimientos:
                if not m.identificado or m.sucursal_declarada or m.sucursal_por_folio:
                    continue
                detalle = sugerir_sucursal_detalle(
                    estado_ec, m.cliente_match, m.abono, empresa_ec.nombre_reporte
                )
                if detalle:
                    m.sucursal_sugerida, m.sucursal_sugerida_motivo = detalle[0], detalle[1]

        if nuevo or historial_id_actual[0] is None:
            rid = ahora.strftime("%Y%m%d_%H%M%S_") + f"{ahora.microsecond // 1000:03d}"
            historial_id_actual[0] = rid
            registro = construir_registro(
                rid, ahora.strftime("%Y-%m-%d"), ahora.strftime("%H:%M:%S"),
                banco, historial_archivo_actual[0], ultima_ruta_csv[0] or "",
                empresa_clave, movimientos,
            )
            historial_registros.insert(0, registro)
        else:
            existente = next((r for r in historial_registros if r.get("id") == historial_id_actual[0]), None)
            base = existente or {}
            registro = construir_registro(
                historial_id_actual[0], base.get("fecha", ahora.strftime("%Y-%m-%d")),
                base.get("hora", ahora.strftime("%H:%M:%S")), banco,
                base.get("archivo", historial_archivo_actual[0]),
                ultima_ruta_csv[0] or base.get("ruta_csv", ""),
                empresa_clave, movimientos,
                subido_sipp=base.get("subido_sipp", False),  # preservar la marca
            )
            if existente is not None:
                historial_registros[historial_registros.index(existente)] = registro
            else:
                historial_registros.insert(0, registro)

        guardar_historial(HISTORIAL_PATH, historial_registros)

    def marcar_movimientos_ya_subidos() -> int:
        """Marca (ya_subido=True) los movimientos que YA venían en una extracción
        PREVIA del mismo banco (cortes acumulativos del día), para mostrarlos en
        gris y excluirlos de la carga a SIPP. Devuelve cuántos quedaron marcados."""
        if not movimientos:
            return 0
        banco = movimientos[0].banco
        subidas = claves_subidas(historial_registros, banco, excluir_id=historial_id_actual[0])
        n = 0
        for m in movimientos:
            m.ya_subido = clave_movimiento(m) in subidas
            if m.ya_subido:
                n += 1
        return n

    def restaurar_registro(registro: dict) -> None:
        """Vuelve a poner en pantalla una extracción guardada, con todos sus
        últimos cambios de cliente/cuenta/sucursal/folio."""
        nonlocal movimientos
        movimientos = [movimiento_desde_dict(d) for d in registro.get("movimientos", [])]
        historial_id_actual[0] = registro.get("id")
        historial_archivo_actual[0] = registro.get("archivo", "")
        ruta = registro.get("ruta_csv") or None
        ultima_ruta_csv[0] = ruta if ruta and os.path.exists(ruta) else None
        banco = registro.get("banco", "")
        banco_detectado_text.value = f"Banco detectado: {banco}" if banco else ""
        archivo_nombre_text.value = registro.get("archivo") or "Extracción restaurada"
        archivo_nombre_text.italic = False
        archivo_nombre_text.color = ft.Colors.ON_SURFACE

        clave = registro.get("empresa_clave")
        if clave and clave in EMPRESA_POR_CLAVE:
            empresa_ref[0] = EMPRESA_POR_CLAVE[clave]
            empresa_dropdown.value = clave
            cuenta_bancaria_dropdown.options = _opciones_cuenta()
            cuenta_bancaria_dropdown.value = None

        sucursal_cache.clear()
        boton_cargar_estado_cuenta.disabled = not movimientos
        estado_text.value = (
            f"Extracción restaurada: {banco} · {registro.get('fecha','')} "
            f"{registro.get('hora','')} ({len(movimientos)} movimientos)."
        )
        page.pop_dialog()
        refrescar_resumen()
        refrescar_tabla()

    def eliminar_registro_historial(registro: dict) -> None:
        historial_registros[:] = [r for r in historial_registros if r.get("id") != registro.get("id")]
        if historial_id_actual[0] == registro.get("id"):
            historial_id_actual[0] = None
        guardar_historial(HISTORIAL_PATH, historial_registros)
        _rellenar_lista_historial()
        page.update()

    def marcar_registro_subido(registro_id: str, valor: bool) -> None:
        """Marca/desmarca un bloque como subido a SIPP (base de la deduplicación)."""
        for r in historial_registros:
            if r.get("id") == registro_id:
                r["subido_sipp"] = valor
                break
        guardar_historial(HISTORIAL_PATH, historial_registros)
        # Re-evaluar duplicados de la extracción en pantalla y refrescar.
        marcar_movimientos_ya_subidos()
        _rellenar_lista_historial()
        refrescar_resumen()
        refrescar_tabla()

    historial_lista = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO, tight=True)

    def _fila_historial(r: dict) -> ft.Control:
        total = r.get("num_movimientos", 0)
        ident = r.get("num_identificados", 0)
        pct = (ident / total * 100) if total else 0
        es_actual = r.get("id") == historial_id_actual[0]
        subido = bool(r.get("subido_sipp"))
        boton_subido = ft.IconButton(
            icon=ft.Icons.CLOUD_DONE if subido else ft.Icons.CLOUD_UPLOAD_OUTLINED,
            icon_color=ft.Colors.GREEN_600 if subido else ft.Colors.ON_SURFACE_VARIANT,
            tooltip=("Subida a SIPP (clic para desmarcar)" if subido
                     else "Marcar como subida a SIPP"),
            on_click=lambda _e, rid=r.get("id"), v=not subido: marcar_registro_subido(rid, v),
        )
        checkbox_sel = ft.Checkbox(
            value=r.get("id") in historial_seleccionados,
            on_change=lambda e, rid=r.get("id"): _toggle_historial_sel(rid, e.control.value),
            tooltip="Seleccionar para exportar a Excel",
        )
        return ft.Container(
            padding=12,
            border_radius=8,
            bgcolor=ft.Colors.with_opacity(0.12 if es_actual else 0.04, NAVY),
            content=ft.Row(
                [
                    checkbox_sel,
                    ft.Icon(ft.Icons.RECEIPT_LONG, color=NAVY),
                    ft.Column(
                        [
                            ft.Text(f"{r.get('banco','')} · {r.get('archivo','')}",
                                    weight=ft.FontWeight.BOLD, color=ft.Colors.ON_SURFACE),
                            ft.Text(f"{r.get('fecha','')}  ·  {r.get('hora','')}",
                                    size=12, color=ft.Colors.ON_SURFACE_VARIANT),
                        ],
                        spacing=2, expand=True,
                    ),
                    ft.Column(
                        [
                            ft.Text(f"${r.get('total_abonado',0):,.2f}",
                                    weight=ft.FontWeight.BOLD, color=ft.Colors.ON_SURFACE),
                            ft.Text(f"{total} movs · {ident} ident. ({pct:.0f}%)",
                                    size=12, color=ft.Colors.ON_SURFACE_VARIANT),
                        ],
                        spacing=2, horizontal_alignment=ft.CrossAxisAlignment.END,
                    ),
                    boton_subido,
                    ft.IconButton(
                        icon=ft.Icons.DELETE_OUTLINE,
                        icon_color=ft.Colors.RED_400,
                        tooltip="Eliminar del historial",
                        on_click=lambda _e, reg=r: eliminar_registro_historial(reg),
                    ),
                ],
                spacing=8,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            on_click=lambda _e, reg=r: restaurar_registro(reg),
            tooltip="Clic para abrir esta extracción",
            ink=True,
        )

    def _rellenar_lista_historial() -> None:
        if historial_registros:
            historial_lista.controls = [_fila_historial(r) for r in historial_registros]
        else:
            historial_lista.controls = [
                ft.Text("Aún no hay extracciones registradas.", italic=True,
                        color=ft.Colors.ON_SURFACE_VARIANT)
            ]

    def _toggle_historial_sel(rid: str, valor: bool) -> None:
        if valor:
            historial_seleccionados.add(rid)
        else:
            historial_seleccionados.discard(rid)
        n = len(historial_seleccionados)
        boton_descargar_historial.disabled = n == 0
        boton_descargar_historial.text = f"Descargar Excel ({n})" if n else "Descargar Excel"
        page.update()

    async def exportar_historial_excel(_e) -> None:
        """Genera UN archivo Excel con una hoja por cada extracción seleccionada."""
        seleccion = [r for r in historial_registros if r.get("id") in historial_seleccionados]
        if not seleccion:
            return
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # se crean hojas nombradas por extracción
        usados: set = set()
        total_movs = 0
        for r in seleccion:
            # Se incluyen TODOS los movimientos del snapshot, incluidos los
            # marcados como "ya extraído"/duplicados (solo la subida a SIPP los
            # excluye, no la exportación a Excel).
            movs = [movimiento_desde_dict(d) for d in r.get("movimientos", [])]
            total_movs += len(movs)
            base = f"{r.get('banco', '')}-{(r.get('hora', '') or '').replace(':', '')}"
            ws = wb.create_sheet(title=nombre_hoja_excel(base, usados))
            escribir_hoja_movimientos(ws, movs, sucursal_export)
        if not wb.sheetnames:
            wb.create_sheet("Vacío")

        nombre_def = f"historial_extracciones_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        destino = await file_picker.save_file(
            dialog_title="Guardar historial en Excel",
            file_name=nombre_def,
            allowed_extensions=["xlsx"],
        )
        if not destino:
            return
        if not destino.lower().endswith(".xlsx"):
            destino += ".xlsx"
        try:
            wb.save(destino)
        except OSError as ex:
            estado_text.value = f"No se pudo guardar el Excel: {ex}"
            page.update()
            return
        estado_text.value = (
            f"Exportadas {len(seleccion)} extracción(es) · {len(seleccion)} hoja(s) · "
            f"{total_movs} movimiento(s) (incluye duplicados/ya extraídos) → {os.path.basename(destino)}."
        )
        page.pop_dialog()
        page.update()

    boton_descargar_historial = ft.Button(
        "Descargar Excel",
        icon=ft.Icons.GRID_ON,
        disabled=True,
        bgcolor=ft.Colors.GREEN_700,
        color=ft.Colors.WHITE,
        on_click=lambda e: page.run_task(exportar_historial_excel, e),
    )

    dialogo_historial = ft.AlertDialog(
        modal=True,
        title=ft.Text("Historial de extracciones"),
        content=ft.Container(content=historial_lista, width=640, height=440),
        actions=[
            boton_descargar_historial,
            ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog()),
        ],
    )

    def abrir_historial(_e) -> None:
        # Limpia la selección previa al reabrir.
        historial_seleccionados.clear()
        boton_descargar_historial.disabled = True
        boton_descargar_historial.text = "Descargar Excel"
        _rellenar_lista_historial()
        mostrar_dialogo(dialogo_historial)

    def procesar_csv_path(
        ruta_temporal: str, nombre_archivo: Optional[str] = None, agregar: bool = False
    ) -> None:
        """Detecta el banco, parsea y matchea el CSV en ruta_temporal, actualizando
        movimientos/catálogo/estado. Reutilizado por la carga manual de archivo y
        por la descarga de adjuntos del buzón O365. Si el archivo es válido, se
        conserva su ruta (ultima_ruta_csv) para poder volver a subirlo en SIPP
        (Ingresos Diversos) sin que el usuario tenga que volver a elegirlo.

        Con `agregar=True` NO reemplaza la extracción actual: parsea el archivo y
        anexa sus movimientos a los ya cargados (BBVA descarga los movimientos en
        dos archivos —internos/externos— que se unifican en un solo grid), volviendo
        a correr el match y la deduplicación sobre la lista combinada."""
        nonlocal movimientos
        estado_text.value = "Procesando..."
        page.update()
        try:
            banco = detectar_banco(ruta_temporal)
            if banco is None:
                if agregar:
                    estado_text.value = (
                        "El archivo adicional no se reconoció; se conserva la extracción actual."
                    )
                    page.update()
                    return
                estado_text.value = "No se reconoció el formato del archivo (se esperaba Santander o BanRegio)."
                banco_detectado_text.value = ""
                movimientos = []
                refrescar_resumen()
                refrescar_tabla()
                page.update()
                return

            movs_nuevos = parsear_archivo(ruta_temporal, banco)
            if agregar:
                # BBVA: unificar internos + externos en un solo grid.
                movimientos = movimientos + movs_nuevos
            else:
                banco_detectado_text.value = f"Banco detectado: {banco}"
                movimientos = movs_nuevos
            match_movimientos(movimientos, catalogo)

            propuestas = match_movimientos_por_nombre(movimientos, clientes_normalizados)
            agregadas = guardar_nuevas_cuentas(CATALOGO_PATH, catalogo, propuestas)
            catalogo.extend(agregadas)
            catalogo_info_text.value = f"Catálogo de clientes cargado: {len(catalogo)} cuentas"

            # Nueva extracción: deduplicar contra extracciones previas del mismo
            # banco (cortes acumulativos) para no re-procesar/duplicar. Al AGREGAR
            # el archivo complementario se conserva el mismo registro de historial.
            if not agregar:
                historial_id_actual[0] = None
            ya_subidos = marcar_movimientos_ya_subidos()

            # Movimientos que NO deben subirse por el RPA (se excluyen y van en rojo):
            #  - Traspasos a filiales (no son cobranza).
            #  - BBVA: movimientos de portal de clientes (concepto CI/CE), ya en SIPP.
            traspasos = 0
            portal = 0
            for m in movimientos:
                if es_traspaso(m):
                    m.excluido = True
                    traspasos += 1
                elif es_movimiento_portal_cliente(m):
                    m.excluido = True
                    portal += 1

            identificados_por_nombre = sum(1 for m in movimientos if m.identificado_por_nombre)
            nuevos = len(movimientos) - ya_subidos
            if agregar:
                mensaje = (
                    f"Archivo adicional unificado: {len(movs_nuevos)} movimiento(s) agregado(s). "
                    f"Total en el grid: {len(movimientos)}."
                )
            else:
                mensaje = f"Archivo procesado correctamente. {len(movimientos)} movimientos leídos."
            if ya_subidos:
                mensaje += f" {ya_subidos} ya venían en un corte anterior (en gris, no se re-suben); {nuevos} nuevos."
            if traspasos:
                mensaje += f" {traspasos} traspaso(s) a filiales (en rojo, excluidos del RPA)."
            if portal:
                mensaje += f" {portal} movimiento(s) de portal BBVA (CI/CE) en rojo, excluidos (ya en SIPP)."
            if identificados_por_nombre:
                mensaje += f" {identificados_por_nombre} se identificaron por nombre."
            if agregadas:
                mensaje += f" Se agregaron {len(agregadas)} cuenta(s) nueva(s) al catálogo."
            estado_text.value = mensaje
            if agregar:
                # Conservar la ruta del primer archivo; sumar el nombre del segundo.
                base_nombre = historial_archivo_actual[0] or ""
                nuevo_nombre = nombre_archivo or os.path.basename(ruta_temporal)
                historial_archivo_actual[0] = (
                    f"{base_nombre} + {nuevo_nombre}" if base_nombre else nuevo_nombre
                )
                historial_guardar_snapshot(nuevo=False)
            else:
                ultima_ruta_csv[0] = ruta_temporal
                historial_archivo_actual[0] = nombre_archivo or os.path.basename(ruta_temporal)
                historial_guardar_snapshot(nuevo=True)
        except Exception as ex:
            estado_text.value = f"Error al procesar el archivo: {ex}"
            if not agregar:
                movimientos = []

        # El estado de cuenta solo tiene sentido con un CSV ya cargado.
        boton_cargar_estado_cuenta.disabled = not movimientos
        refrescar_resumen()
        refrescar_tabla()

    file_picker = ft.FilePicker()

    # Layout de BBVA del archivo cargado (interno/externo), para ofrecer subir el
    # complementario una sola vez y validar que el segundo sea el opuesto.
    bbva_layout_actual: list[Optional[str]] = [None]

    def _volcar_a_temporal(archivo) -> str:
        """Vuelca los bytes del archivo elegido a un temporal, conservando la
        extensión original (.csv, .xlsx de BanBajío, .xls de BBVA) para que la
        detección de banco y la subida a SIPP reciban el formato correcto. En modo
        web la API solo entrega los bytes (path es None), por eso el volcado."""
        nombre_lower = archivo.name.lower()
        if nombre_lower.endswith(".xlsx"):
            sufijo = ".xlsx"
        elif nombre_lower.endswith(".xls"):
            sufijo = ".xls"
        else:
            sufijo = ".csv"
        with tempfile.NamedTemporaryFile(suffix=sufijo, delete=False) as tmp:
            tmp.write(archivo.bytes or b"")
            return tmp.name

    _BBVA_LAYOUT_DESC = {
        bbva.LAYOUT_INTERNO: "movimientos internos (mismo banco BBVA)",
        bbva.LAYOUT_EXTERNO: "movimientos externos (otros bancos)",
    }

    async def on_agregar_complemento_bbva(_e) -> None:
        page.pop_dialog()
        archivos = await file_picker.pick_files(
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xls", "xml"],
            allow_multiple=False,
            with_data=True,
        )
        if not archivos:
            return
        archivo = archivos[0]
        ruta_temporal = _volcar_a_temporal(archivo)
        layout_nuevo = bbva.layout(ruta_temporal)
        if layout_nuevo is None:
            estado_text.value = "El archivo adicional no es un archivo de BBVA; no se agregó."
            os.unlink(ruta_temporal)
            page.update()
            return
        if layout_nuevo == bbva_layout_actual[0]:
            estado_text.value = (
                f"El archivo adicional es del MISMO tipo que el ya cargado "
                f"({_BBVA_LAYOUT_DESC.get(layout_nuevo, layout_nuevo)}); elige el complementario."
            )
            os.unlink(ruta_temporal)
            page.update()
            return
        procesar_csv_path(ruta_temporal, nombre_archivo=archivo.name, agregar=True)
        # Ya se unificaron ambos; se limpia el marcador para no volver a ofrecer.
        bbva_layout_actual[0] = None

    dialogo_complemento_bbva = ft.AlertDialog(
        modal=True,
        title=ft.Text("Unificar archivos de BBVA"),
        content=ft.Text("", width=460),
        actions=[
            ft.TextButton("No, continuar", on_click=lambda _e: page.pop_dialog()),
            ft.FilledButton("Sí, agregar", on_click=on_agregar_complemento_bbva),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    def ofrecer_complemento_bbva(layout_actual: Optional[str]) -> None:
        """Si el archivo recién cargado es de BBVA, ofrece agregar el archivo
        complementario (interno↔externo) para unificar ambos en un solo grid."""
        bbva_layout_actual[0] = layout_actual
        if layout_actual not in _BBVA_LAYOUT_DESC:
            return
        otro = bbva.LAYOUT_EXTERNO if layout_actual == bbva.LAYOUT_INTERNO else bbva.LAYOUT_INTERNO
        dialogo_complemento_bbva.content.value = (
            f"Se detectó el archivo de {_BBVA_LAYOUT_DESC[layout_actual]} de BBVA.\n\n"
            f"Para unificar la conciliación, ¿desea agregar el archivo de "
            f"{_BBVA_LAYOUT_DESC[otro]}?"
        )
        mostrar_dialogo(dialogo_complemento_bbva)

    async def on_click_cargar(_e) -> None:
        archivos = await file_picker.pick_files(
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["csv", "xlsx", "xls"],
            allow_multiple=False,
            with_data=True,
        )
        if not archivos:
            return
        archivo = archivos[0]
        archivo_nombre_text.value = archivo.name
        archivo_nombre_text.italic = False
        archivo_nombre_text.color = ft.Colors.ON_SURFACE
        page.update()

        # Se conserva el temporal al terminar: sirve para volver a subirlo a
        # "Ingresos Diversos" en SIPP (ver ultima_ruta_csv).
        if ultima_ruta_csv[0] and os.path.exists(ultima_ruta_csv[0]):
            os.unlink(ultima_ruta_csv[0])
        ruta_temporal = _volcar_a_temporal(archivo)
        procesar_csv_path(ruta_temporal, nombre_archivo=archivo.name)

        # BBVA parte los movimientos en dos archivos (internos/externos): si se
        # cargó uno válido, ofrecer unificar con el complementario.
        if movimientos:
            ofrecer_complemento_bbva(bbva.layout(ruta_temporal))
        else:
            bbva_layout_actual[0] = None

    boton_cargar = ft.Button(
        "Cargar archivo bancario (.csv)",
        icon=ft.Icons.UPLOAD_FILE,
        on_click=on_click_cargar,
        bgcolor=NAVY,
        color=ft.Colors.WHITE,
    )

    # --- Estado de cuenta (.xlsx): índice cliente→sucursal para sugerir sucursal ---
    estado_cuenta_ref: list[Optional[EstadoCuenta]] = [None]
    estado_cuenta_text = ft.Text(
        "Estado de cuenta: no cargado", italic=True, color=ft.Colors.ON_SURFACE_VARIANT
    )

    async def on_click_cargar_estado_cuenta(_e) -> None:
        archivos = await file_picker.pick_files(
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx"],
            allow_multiple=False,
            with_data=True,
        )
        if not archivos:
            return
        archivo = archivos[0]
        estado_cuenta_text.value = f"Cargando estado de cuenta '{archivo.name}'..."
        estado_cuenta_text.italic = True
        page.update()
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(archivo.bytes or b"")
                ruta_xlsx = tmp.name
            estado = await asyncio.to_thread(cargar_estado_cuenta, ruta_xlsx)
            os.unlink(ruta_xlsx)
            estado_cuenta_ref[0] = estado
            sucursal_cache.clear()
            identificados = sum(1 for m in movimientos if m.identificado)
            estado_cuenta_text.value = (
                f"Estado de cuenta cargado: {estado.num_clientes} cliente(s). "
                f"Sucursal sugerida para {identificados} movimiento(s) identificado(s)."
            )
            estado_cuenta_text.italic = False
            estado_cuenta_text.color = ft.Colors.ON_SURFACE
            refrescar_tabla()  # puebla la columna "Sucursal sugerida"
            historial_guardar_snapshot()  # congela las sugerencias en el historial
        except Exception as ex:
            estado_cuenta_ref[0] = None
            estado_cuenta_text.value = f"Error al cargar estado de cuenta: {ex}"
            estado_cuenta_text.color = ft.Colors.RED_600
        page.update()

    boton_cargar_estado_cuenta = ft.Button(
        "Cargar estado de cuenta (.xlsx)",
        icon=ft.Icons.TABLE_VIEW,
        on_click=lambda e: page.run_task(on_click_cargar_estado_cuenta, e),
        bgcolor=NAVY,
        color=ft.Colors.WHITE,
        disabled=True,  # se habilita al cargar un archivo bancario (.csv)
    )

    # --- Declaración manual de sucursal (override de la sugerida) ---
    movimiento_sucursal_edit: list[Movimiento | None] = [None]
    sucursal_edit_info = ft.Text("")
    sucursal_edit_dropdown = CustomDropdown(
        label="Sucursal",
        editable=True,
        enable_filter=True,
        menu_height=300,
        width=360,
    )

    def on_confirmar_sucursal(_e) -> None:
        mov = movimiento_sucursal_edit[0]
        if mov is not None:
            mov.sucursal_declarada = sucursal_edit_dropdown.value or None
        page.pop_dialog()
        refrescar_tabla()
        historial_guardar_snapshot()

    def on_limpiar_sucursal(_e) -> None:
        mov = movimiento_sucursal_edit[0]
        if mov is not None:
            mov.sucursal_declarada = None
        page.pop_dialog()
        refrescar_tabla()
        historial_guardar_snapshot()

    dialogo_sucursal = ft.AlertDialog(
        modal=True,
        title=ft.Text("Declarar sucursal"),
        content=ft.Column(
            [sucursal_edit_info, sucursal_edit_dropdown],
            tight=True,
            spacing=12,
            width=400,
        ),
        actions=[
            ft.TextButton("Usar sugerida", on_click=on_limpiar_sucursal),
            ft.TextButton("Cancelar", on_click=lambda _e: page.pop_dialog()),
            ft.Button("Aplicar", on_click=on_confirmar_sucursal, bgcolor=NAVY, color=ft.Colors.WHITE),
        ],
    )

    def abrir_dialogo_sucursal(mov: Movimiento) -> None:
        estado = estado_cuenta_ref[0]
        if estado is None:
            return
        empresa = empresa_ref[0]
        # Opciones: las sucursales del cliente en la empresa; si no está, todas
        # las de la empresa.
        opciones = estado.sucursales_de_cliente(mov.cliente_match or "", empresa.nombre_reporte)
        if not opciones:
            opciones = estado.sucursales(empresa.nombre_reporte)
        sucursal_edit_dropdown.options = [ft.dropdown.Option(s) for s in opciones]
        # Valor inicial: la declarada, o la sugerida.
        detalle = sugerir_sucursal_detalle(estado, mov.cliente_match, mov.abono, empresa.nombre_reporte)
        sugerida = detalle[0] if detalle else None
        sucursal_edit_dropdown.value = mov.sucursal_declarada or sugerida
        sucursal_edit_info.value = (
            f"{(mov.cliente_match or '')} · Abono ${mov.abono:,.2f}\n"
            f"Sugerida: {sugerida or '—'}"
        )
        movimiento_sucursal_edit[0] = mov
        mostrar_dialogo(dialogo_sucursal)

    # --- Búsqueda de folios pendientes en SIPP (RPA) ---
    sipp_usuario_field = ft.TextField(label="Usuario SIPP", autofocus=True)
    sipp_password_field = ft.TextField(label="Contraseña SIPP", password=True, can_reveal_password=True)
    sipp_recordar_check = ft.Checkbox(label="Recordar credenciales en este equipo")
    sipp_progreso_text = ft.Text("")
    sipp_log_panel, sipp_log_fn, sipp_log_reset = crear_panel_log_rpa()
    sipp_ver_navegador_check = crear_check_ver_navegador()

    async def on_cancelar_sipp(_e) -> None:
        page.pop_dialog()

    async def on_confirmar_sipp(_e) -> None:
        usuario = (sipp_usuario_field.value or "").strip()
        password = sipp_password_field.value or ""
        if not usuario or not password:
            sipp_progreso_text.value = "Usuario y contraseña son obligatorios."
            page.update()
            return

        if sipp_recordar_check.value:
            guardar_credenciales(usuario, password)
        else:
            borrar_credenciales()

        boton_sipp_buscar.disabled = True
        # "Cerrar" queda habilitado: permite cerrar el modal y dejar el RPA en
        # segundo plano (el FAB lo reabre).
        sipp_log_reset()
        sipp_progreso_text.value = "Ejecutando RPA en segundo plano..."
        rpa_inicio(dialogo_sipp)
        page.update()

        def log_fn(mensaje: str, nivel: str = "info") -> None:
            sipp_progreso_text.value = mensaje
            sipp_log_fn(mensaje, nivel)

        ok_rpa = True
        try:
            candidatos = extraer_folios_pendientes(movimientos)
            propuestas = await buscar_y_aplicar_folios(
                candidatos, usuario, password, empresa=empresa_ref[0],
                headless=not ver_navegador_ref[0],
                log_fn=log_fn, todos_movimientos=movimientos,
            )
            agregadas = guardar_nuevas_cuentas(CATALOGO_PATH, catalogo, propuestas)
            catalogo.extend(agregadas)
            catalogo_info_text.value = f"Catálogo de clientes cargado: {len(catalogo)} cuentas"

            identificados_por_folio = sum(1 for m in movimientos if m.identificado_por_folio)
            mensaje = f"Búsqueda en SIPP terminada. {identificados_por_folio} folio(s) identificado(s)."
            if agregadas:
                mensaje += f" Se agregaron {len(agregadas)} cuenta(s) nueva(s) al catálogo."
            estado_text.value = mensaje
            sipp_progreso_text.value = mensaje
            log_fn(mensaje, "ok")
        except Exception as ex:
            ok_rpa = False
            estado_text.value = f"Error durante la búsqueda en SIPP: {ex}"
            log_fn(f"Error: {ex}", "error")
        finally:
            boton_sipp_buscar.disabled = False
            boton_sipp_cancelar.disabled = False
            sipp_password_field.value = ""
            refrescar_resumen()
            refrescar_tabla()
            historial_guardar_snapshot()
            rpa_fin(ok_rpa)
            page.update()

    boton_sipp_cancelar = ft.TextButton("Cerrar", on_click=on_cancelar_sipp)
    boton_sipp_buscar = ft.Button("Buscar", on_click=on_confirmar_sipp, bgcolor=NAVY, color=ft.Colors.WHITE)

    dialogo_sipp = ft.AlertDialog(
        modal=True,
        title=ft.Text("Buscar folios pendientes en SIPP"),
        content=ft.Column(
            [
                sipp_progreso_text,
                sipp_usuario_field,
                sipp_password_field,
                sipp_recordar_check,
                sipp_ver_navegador_check,
                sipp_log_panel,
            ],
            tight=True,
            spacing=12,
            width=420,
        ),
        actions=[boton_sipp_cancelar, boton_sipp_buscar],
    )

    def on_click_buscar_sipp(_e) -> None:
        if not movimientos:
            estado_text.value = "Primero carga un archivo bancario."
            page.update()
            return

        candidatos = extraer_folios_pendientes(movimientos)
        if not candidatos:
            estado_text.value = "No hay folios candidatos pendientes de buscar en SIPP."
            page.update()
            return

        usuario_guardado, password_guardado = cargar_credenciales()
        sipp_usuario_field.value = usuario_guardado
        sipp_password_field.value = password_guardado
        sipp_recordar_check.value = bool(usuario_guardado)
        sipp_progreso_text.value = f"{len(candidatos)} folio(s) candidato(s) detectado(s) en la descripción."
        mostrar_dialogo(dialogo_sipp)

    boton_buscar_sipp = ft.Button(
        "Buscar Folios en SIPP",
        icon=ft.Icons.TRAVEL_EXPLORE,
        on_click=on_click_buscar_sipp,
        bgcolor=ORANGE,
        color=ft.Colors.WHITE,
    )

    # ──────────────────────────────────────────────────────
    # Factoraje (BAJA FERRIES): PDF de intereses + captura en SIPP
    # ──────────────────────────────────────────────────────
    # Instituciones del combo de SIPP (value → nombre), del HTML de factoraje.
    INSTITUCIONES_FACTORAJE = [
        ("0", "BBVA MEXICO, S.A."),
        ("1", "ARRENDADORA Y FACTOR BANORTE"),
        ("2", "BANCO SANTANDER MEXICO"),
        ("3", "START BANREGIO"),
        ("4", "BANCO VE POR MAS"),
        ("5", "BANCO J.P. MORGAN"),
        ("6", "HSBC MEXICO"),
    ]

    def _es_baja_ferries(mov: Movimiento) -> bool:
        # Detecta por el cliente identificado y, como respaldo, por la descripción
        # o el texto de búsqueda (por si el movimiento aún no se identificó).
        texto = normalizar(
            f"{mov.cliente_match or ''} {mov.descripcion or ''} {mov.texto_busqueda or ''}"
        )
        return "BAJA FERRIES" in texto or "BAJAFERRIES" in texto

    factoraje_filas: list = []
    factoraje_pares: list = []  # [(Movimiento, FilaFactoraje)]

    factoraje_folio_field = ft.TextField(label="Folio de conciliación (SIPP)", width=260)
    factoraje_institucion_dd = CustomDropdown(
        label="Institución de factoraje",
        width=320,
        value="0",
        options=[ft.dropdown.Option(key=v, text=n) for v, n in INSTITUCIONES_FACTORAJE],
    )
    factoraje_usuario_field = ft.TextField(label="Usuario SIPP")
    factoraje_password_field = ft.TextField(label="Contraseña SIPP", password=True, can_reveal_password=True)
    factoraje_info_text = ft.Text("")
    factoraje_tabla = ft.Column(spacing=4, scroll=ft.ScrollMode.AUTO, tight=True)
    factoraje_log_panel, factoraje_log_fn, factoraje_log_reset = crear_panel_log_rpa()
    factoraje_ver_navegador_check = crear_check_ver_navegador()
    boton_cargar_pdf_factoraje = ft.Button(
        "Cargar PDF de intereses",
        icon=ft.Icons.PICTURE_AS_PDF,
        on_click=lambda e: page.run_task(on_click_cargar_pdf_factoraje, e),
        bgcolor=NAVY,
        color=ft.Colors.WHITE,
    )

    def _cruzar_factoraje() -> None:
        """Empareja cada renglón del PDF con un movimiento de BAJA FERRIES por
        folio (del concepto) y, si no, por monto neto (= Monto a Recibir)."""
        nonlocal factoraje_pares
        factoraje_pares = []
        movs_bf = [m for m in movimientos if _es_baja_ferries(m)]
        usados = set()
        for fila in factoraje_filas:
            elegido = None
            for m in movs_bf:
                if id(m) in usados:
                    continue
                folio_mov = extraer_folio(m.texto_busqueda)
                if folio_mov and normalizar(folio_mov).replace(" ", "") == fila.folio:
                    elegido = m
                    break
            if elegido is None:
                for m in movs_bf:
                    if id(m) in usados:
                        continue
                    if abs(m.abono - fila.monto_recibir) < 0.01:
                        elegido = m
                        break
            if elegido is not None:
                usados.add(id(elegido))
                elegido.factoraje_interes = fila.monto_intereses
                elegido.factoraje_folio_pdf = fila.folio_texto
                factoraje_pares.append((elegido, fila))

    def _refrescar_tabla_factoraje() -> None:
        factoraje_tabla.controls = [
            ft.Row(
                [
                    ft.Text(f.folio_texto, width=110, weight=ft.FontWeight.BOLD),
                    ft.Text(m.cliente_match or "-", width=160, overflow=ft.TextOverflow.ELLIPSIS),
                    ft.Text(f"neto ${m.abono:,.2f}", width=140),
                    ft.Text(f"interés ${f.monto_intereses:,.2f}", width=150, color=ft.Colors.GREEN_700),
                    ft.Text(f"ref {m.referencia}", width=110, color=ft.Colors.ON_SURFACE_VARIANT),
                ],
                spacing=8,
            )
            for m, f in factoraje_pares
        ]
        sin_cruce = len(factoraje_filas) - len(factoraje_pares)
        factoraje_info_text.value = (
            f"{len(factoraje_pares)} movimiento(s) de BAJA FERRIES emparejado(s) con el PDF"
            + (f"; {sin_cruce} renglón(es) del PDF sin movimiento." if sin_cruce else ".")
        )
        boton_factoraje_aplicar.disabled = not factoraje_pares

    async def on_click_cargar_pdf_factoraje(_e) -> None:
        archivos = await file_picker.pick_files(
            file_type=ft.FilePickerFileType.CUSTOM, allowed_extensions=["pdf"],
            allow_multiple=False, with_data=True,
        )
        if not archivos:
            return
        nonlocal factoraje_filas
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(archivos[0].bytes or b"")
            ruta_pdf = tmp.name
        try:
            factoraje_filas = await asyncio.to_thread(extraer_factoraje, ruta_pdf)
            _cruzar_factoraje()
            _refrescar_tabla_factoraje()
            refrescar_tabla()  # refleja el interés en el grid si se muestra
        except Exception as ex:
            factoraje_info_text.value = f"Error al leer el PDF de factoraje: {ex}"
            page.update()
        finally:
            try:
                os.unlink(ruta_pdf)
            except OSError:
                pass

    async def on_confirmar_factoraje(_e) -> None:
        usuario = (factoraje_usuario_field.value or "").strip()
        password = factoraje_password_field.value or ""
        folio = (factoraje_folio_field.value or "").strip()
        if not usuario or not password:
            factoraje_info_text.value = "Usuario y contraseña son obligatorios."
            page.update()
            return
        if not folio:
            factoraje_info_text.value = "Indica el folio de conciliación de SIPP."
            page.update()
            return
        items = [
            {"folio": f.folio, "interes": f.monto_intereses, "abono": m.abono, "referencia": m.referencia}
            for m, f in factoraje_pares
        ]
        boton_factoraje_aplicar.disabled = True
        factoraje_log_reset()
        factoraje_info_text.value = "Ejecutando RPA de factoraje..."
        rpa_inicio(dialogo_factoraje)
        page.update()

        def log_fn(mensaje: str, nivel: str = "info") -> None:
            factoraje_info_text.value = mensaje
            factoraje_log_fn(mensaje, nivel)

        ok_rpa = True
        try:
            n = await aplicar_factoraje_en_sipp(
                folio, factoraje_institucion_dd.value or "0", items,
                usuario, password, empresa=empresa_ref[0],
                headless=not ver_navegador_ref[0], log_fn=log_fn,
            )
            mensaje_final = (
                f"✅ Flujo de factoraje completado: {n} movimiento(s) de BAJA FERRIES "
                f"con su interés aplicado y la conciliación guardada en SIPP."
            )
            estado_text.value = mensaje_final
            log_fn(mensaje_final, "ok")
        except Exception as ex:
            ok_rpa = False
            log_fn(f"Error en factoraje: {ex}", "error")
        finally:
            boton_factoraje_aplicar.disabled = False
            factoraje_password_field.value = ""
            rpa_fin(ok_rpa)
            page.update()

    boton_factoraje_aplicar = ft.Button(
        "Aplicar factoraje en SIPP", icon=ft.Icons.PLAY_ARROW,
        on_click=lambda e: page.run_task(on_confirmar_factoraje, e),
        bgcolor=NAVY, color=ft.Colors.WHITE, disabled=True,
    )
    dialogo_factoraje = ft.AlertDialog(
        modal=True,
        title=ft.Text("Factoraje BAJA FERRIES"),
        content=ft.Container(
            content=ft.Column(
                [
                    factoraje_info_text,
                    boton_cargar_pdf_factoraje,
                    ft.Container(content=factoraje_tabla, height=200),
                    ft.Row([factoraje_folio_field, factoraje_institucion_dd], spacing=12),
                    ft.Row([factoraje_usuario_field, factoraje_password_field], spacing=12),
                    factoraje_ver_navegador_check,
                    factoraje_log_panel,
                ],
                tight=True, spacing=12, scroll=ft.ScrollMode.AUTO,
            ),
            width=720,
        ),
        actions=[
            ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog()),
            boton_factoraje_aplicar,
        ],
    )

    def on_click_factoraje(_e) -> None:
        # El modal se abre SIEMPRE. Si no hay BAJA FERRIES en la extracción, se
        # avisa dentro del propio modal (antes se cortaba en silencio).
        movs_bf = [m for m in movimientos if _es_baja_ferries(m)]
        # Reinicia estado del modal para no arrastrar datos de una corrida previa.
        nonlocal factoraje_filas
        factoraje_filas = []
        factoraje_pares.clear()
        _refrescar_tabla_factoraje()
        if not movs_bf:
            factoraje_info_text.value = (
                "No se detectaron movimientos de BAJA FERRIES en esta extracción. "
                "Verifica que estén cargados/identificados antes de aplicar factoraje."
            )
        else:
            factoraje_info_text.value = (
                f"{len(movs_bf)} movimiento(s) de BAJA FERRIES en la extracción. "
                "Carga el PDF de intereses para cruzarlos."
            )
        usuario_guardado, password_guardado = cargar_credenciales()
        factoraje_usuario_field.value = usuario_guardado or ""
        factoraje_password_field.value = password_guardado or ""
        mostrar_dialogo(dialogo_factoraje)

    boton_factoraje = ft.Button(
        "Factoraje",
        icon=ft.Icons.REQUEST_QUOTE,
        on_click=on_click_factoraje,
        bgcolor=ORANGE,
        color=ft.Colors.WHITE,
    )

    # ──────────────────────────────────────────────────────
    # Carga a "Ingresos Diversos - Agregar" en SIPP
    # ──────────────────────────────────────────────────────
    fecha_operacion_field = ft.TextField(
        label="Fecha de Operación",
        value=date.today().strftime("%d/%m/%Y"),
        width=160,
        color=ft.Colors.ON_SURFACE,
        border=ft.InputBorder.UNDERLINE,
        border_color=BORDER,
    )
    # Empresa seleccionada: define el login de SIPP, el catálogo de cuentas y
    # las sucursales (estado de cuenta). Es global a toda la app.
    empresa_ref = [EMPRESA_DEFAULT]

    def _opciones_cuenta() -> list:
        return [ft.dropdown.Option(key=c.id_sipp, text=c.nombre) for c in empresa_ref[0].cuentas]

    def nombre_cuenta_bancaria(id_sipp: str) -> str:
        return {c.id_sipp: c.nombre for c in empresa_ref[0].cuentas}.get(id_sipp or "", "")

    cuenta_bancaria_dropdown = CustomDropdown(
        label="Cuenta Bancaria (SIPP)",
        width=420,
        color=ft.Colors.ON_SURFACE,
        editable=True,        # campo escribible
        enable_filter=True,   # filtra la lista mientras escribes
        menu_height=300,      # tope de altura: el resto hace scroll
        options=_opciones_cuenta(),
        border_color=BORDER,
    )

    def on_cambiar_empresa(e) -> None:
        empresa_ref[0] = EMPRESA_POR_CLAVE.get(e.control.value, EMPRESA_DEFAULT)
        # Las cuentas y sucursales del flujo CSV cambian con la empresa.
        cuenta_bancaria_dropdown.options = _opciones_cuenta()
        cuenta_bancaria_dropdown.value = None
        sucursal_cache.clear()
        refrescar_tabla()
        page.update()

    empresa_dropdown = CustomDropdown(
        label="Empresa (SIPP)",
        width=220,
        color=ft.Colors.ON_SURFACE,
        value=EMPRESA_DEFAULT.clave,
        options=[ft.dropdown.Option(key=e.clave, text=e.nombre) for e in EMPRESAS],
        on_select=on_cambiar_empresa,
        border_color=BORDER,
    )

    # Empresa y cuenta para el flujo de Contado, INDEPENDIENTES del CSV (puedes
    # trabajar el CSV con una empresa y Contado con otra sin regresar a cambiar).
    empresa_contado_ref = [EMPRESA_DEFAULT]

    def _opciones_cuenta_contado() -> list:
        return [ft.dropdown.Option(key=c.id_sipp, text=c.nombre) for c in empresa_contado_ref[0].cuentas]

    cuenta_contado_dropdown = CustomDropdown(
        label="Cuenta Bancaria por default (SIPP)",
        width=420,
        color=ft.Colors.ON_SURFACE,
        editable=True,
        enable_filter=True,
        menu_height=300,
        options=_opciones_cuenta_contado(),
        border_color=BORDER,
    )

    def on_cambiar_empresa_contado(e) -> None:
        empresa_contado_ref[0] = EMPRESA_POR_CLAVE.get(e.control.value, EMPRESA_DEFAULT)
        cuenta_contado_dropdown.options = _opciones_cuenta_contado()
        cuenta_contado_dropdown.value = None
        # Los ids de cuenta cambian por empresa: limpiamos la asignación previa.
        for p in pagos_contado:
            p.cuenta_bancaria = ""
        refrescar_tabla_pagos_contado()
        page.update()

    empresa_contado_dropdown = CustomDropdown(
        label="Empresa (SIPP)",
        width=220,
        color=ft.Colors.ON_SURFACE,
        value=EMPRESA_DEFAULT.clave,
        options=[ft.dropdown.Option(key=e.clave, text=e.nombre) for e in EMPRESAS],
        on_select=on_cambiar_empresa_contado,
        
    )

    ingresos_div_usuario_field = ft.TextField(label="Usuario SIPP", autofocus=True)
    ingresos_div_password_field = ft.TextField(label="Contraseña SIPP", password=True, can_reveal_password=True)
    ingresos_div_recordar_check = ft.Checkbox(label="Recordar credenciales en este equipo")
    ingresos_div_progreso_text = ft.Text("")
    ingresos_div_log_panel, ingresos_div_log_fn, ingresos_div_log_reset = crear_panel_log_rpa()

    def on_cancelar_ingresos_div(_e) -> None:
        page.pop_dialog()

    async def on_confirmar_ingresos_div(_e) -> None:
        usuario = (ingresos_div_usuario_field.value or "").strip()
        password = ingresos_div_password_field.value or ""
        if not usuario or not password:
            ingresos_div_progreso_text.value = "Usuario y contraseña son obligatorios."
            page.update()
            return

        if ingresos_div_recordar_check.value:
            guardar_credenciales(usuario, password)
        else:
            borrar_credenciales()

        boton_ingresos_div_confirmar.disabled = True
        # "Cerrar" habilitado: cerrar el modal deja el RPA en segundo plano (FAB).
        ingresos_div_log_reset()
        ingresos_div_progreso_text.value = "Ejecutando RPA... (se abrirá el navegador para revisar y guardar)"
        rpa_inicio(dialogo_ingresos_div)
        page.update()

        def log_fn(mensaje: str, nivel: str = "info") -> None:
            ingresos_div_progreso_text.value = mensaje
            ingresos_div_log_fn(mensaje, nivel)

        cuenta_nombre = nombre_cuenta_bancaria(cuenta_bancaria_dropdown.value or "")
        ok_rpa = True
        try:
            enviados = await cargar_ingresos_diversos_en_sipp(
                movimientos,
                cuenta_nombre,
                fecha_operacion_field.value or "",
                ultima_ruta_csv[0],
                usuario,
                password,
                estado_cuenta=estado_cuenta_ref[0],
                empresa=empresa_ref[0],
                headless=False,  # deja el navegador abierto para revisar/guardar
                log_fn=log_fn,
                sucursal_resolver=sucursal_efectiva,  # WYSIWYG: la del grid
            )
            mensaje = (
                f"Carga a SIPP (Ingresos Diversos) lista: {enviados} movimiento(s) identificado(s) "
                "enviados a emparejar. Revisa la previsualización y guarda manualmente en SIPP."
            )
            estado_text.value = mensaje
            log_fn(mensaje, "ok")
            # Marca la extracción actual como subida a SIPP (base de la
            # deduplicación de los próximos cortes acumulativos del banco).
            if historial_id_actual[0]:
                marcar_registro_subido(historial_id_actual[0], True)
        except Exception as ex:
            ok_rpa = False
            estado_text.value = f"Error al cargar Ingresos Diversos en SIPP: {ex}"
            log_fn(f"Error: {ex}", "error")
        finally:
            boton_ingresos_div_confirmar.disabled = False
            boton_ingresos_div_cancelar.disabled = False
            ingresos_div_password_field.value = ""
            rpa_fin(ok_rpa)
            page.update()

    boton_ingresos_div_cancelar = ft.TextButton("Cerrar", on_click=on_cancelar_ingresos_div)
    boton_ingresos_div_confirmar = ft.Button(
        "Cargar", on_click=on_confirmar_ingresos_div, bgcolor=NAVY, color=ft.Colors.WHITE
    )

    dialogo_ingresos_div = ft.AlertDialog(
        modal=True,
        title=ft.Text("Cargar a SIPP: Ingresos Diversos"),
        content=ft.Column(
            [
                ingresos_div_progreso_text,
                ingresos_div_usuario_field,
                ingresos_div_password_field,
                ingresos_div_recordar_check,
                ingresos_div_log_panel,
            ],
            tight=True,
            spacing=12,
            width=420,
        ),
        actions=[boton_ingresos_div_cancelar, boton_ingresos_div_confirmar],
    )

    def on_click_cargar_ingresos_div(_e) -> None:
        if not movimientos:
            estado_text.value = "Primero carga un archivo bancario."
            page.update()
            return
        if not ultima_ruta_csv[0] or not os.path.exists(ultima_ruta_csv[0]):
            estado_text.value = "No hay un archivo bancario disponible para volver a subir; carga uno de nuevo."
            page.update()
            return
        if not cuenta_bancaria_dropdown.value:
            estado_text.value = "Selecciona la Cuenta Bancaria (SIPP) antes de continuar."
            page.update()
            return
        if not (fecha_operacion_field.value or "").strip():
            estado_text.value = "Indica la Fecha de Operación antes de continuar."
            page.update()
            return
        identificados = sum(1 for m in movimientos if m.identificado)
        if identificados == 0:
            estado_text.value = "No hay movimientos identificados (con cliente asignado) para enviar."
            page.update()
            return

        usuario_guardado, password_guardado = cargar_credenciales()
        ingresos_div_usuario_field.value = usuario_guardado
        ingresos_div_password_field.value = password_guardado
        ingresos_div_recordar_check.value = bool(usuario_guardado)
        ingresos_div_progreso_text.value = f"{identificados} movimiento(s) identificado(s) se intentarán emparejar."
        mostrar_dialogo(dialogo_ingresos_div)

    boton_cargar_ingresos_div = ft.Button(
        "Cargar a SIPP (Ingresos Diversos)",
        icon=ft.Icons.ACCOUNT_BALANCE,
        on_click=on_click_cargar_ingresos_div,
        bgcolor=NAVY,
        color=ft.Colors.WHITE,
    )

    # ──────────────────────────────────────────────────────
    # Pestaña: Buzón O365 (Microsoft Graph)
    # ──────────────────────────────────────────────────────
    correos_o365: list[CorreoResumen] = []
    estado_o365_text = ft.Text("")

    def columna_o365(
        texto: str,
        fixed_width: Optional[float] = None,
        size: Optional[DataColumnSize] = None,
    ) -> DataColumn2:
        return DataColumn2(
            ft.Text(texto, weight=ft.FontWeight.BOLD, color=ft.Colors.ON_SURFACE),
            fixed_width=fixed_width,
            size=size,
        )

    # Correos marcados para extraer (por id). Al cargar la bandeja se marcan
    # todos por default; el usuario desmarca los que no quiera.
    correos_seleccionados: set[str] = set()

    def on_toggle_todos_o365(e) -> None:
        if e.control.value:
            correos_seleccionados.update(c.id_correo for c in correos_o365)
        else:
            correos_seleccionados.clear()
        refrescar_tabla_o365()

    check_todos_o365 = ft.Checkbox(value=True, on_change=on_toggle_todos_o365, tooltip="Seleccionar todos")

    def on_toggle_correo(correo: CorreoResumen, valor: bool) -> None:
        if valor:
            correos_seleccionados.add(correo.id_correo)
        else:
            correos_seleccionados.discard(correo.id_correo)
        # Mantener el "seleccionar todo" en sincronía.
        check_todos_o365.value = bool(correos_o365) and all(
            c.id_correo in correos_seleccionados for c in correos_o365
        )
        page.update()

    tabla_o365 = DataTable2(
        columns=[
            DataColumn2(check_todos_o365, fixed_width=55),
            columna_o365("Fecha", fixed_width=150),
            columna_o365("Remitente", fixed_width=260),
            columna_o365("Asunto", size=DataColumnSize.L),
            columna_o365("Adjuntos", fixed_width=90),
            columna_o365("Acciones", fixed_width=110),
        ],
        rows=[],
        min_width=1150,
        fixed_top_rows=1,
        column_spacing=16,
        expand=True,
    )

    async def on_descargar_adjunto(_e, correo: CorreoResumen) -> None:
        estado_o365_text.value = f"Descargando adjuntos de '{correo.asunto}'..."
        page.update()
        destino_dir = tempfile.mkdtemp(prefix="mh_cobranza_o365_")
        try:
            rutas = await asyncio.to_thread(descargar_adjuntos, correo.mensaje, destino_dir)
            if not rutas:
                estado_o365_text.value = f"'{correo.asunto}' no tiene adjuntos."
                page.update()
                return

            # Si hay un CSV, es el archivo bancario: se carga en Conciliación.
            csv_rutas = [r for r in rutas if r.lower().endswith(".csv")]
            if csv_rutas:
                archivo_nombre_text.value = os.path.basename(csv_rutas[0])
                archivo_nombre_text.italic = False
                archivo_nombre_text.color = ft.Colors.ON_SURFACE
                procesar_csv_path(csv_rutas[0], nombre_archivo=os.path.basename(csv_rutas[0]))
                estado_o365_text.value = f"Adjunto de '{correo.asunto}' cargado en Conciliación Bancaria."
                tabs.selected_index = 0
                page.update()
                return

            # Cualquier otro adjunto (PDF, imagen, etc.): se copia a Descargas y
            # se revela en el explorador para que el usuario lo guarde/abra.
            descargas_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            os.makedirs(descargas_dir, exist_ok=True)
            guardados = []
            for ruta in rutas:
                destino = os.path.join(descargas_dir, os.path.basename(ruta))
                await asyncio.to_thread(shutil.copy2, ruta, destino)
                guardados.append(destino)
            if guardados:
                revelar_en_explorador(guardados[0])
            nombres = ", ".join(os.path.basename(g) for g in guardados)
            estado_o365_text.value = f"{len(guardados)} adjunto(s) descargado(s) a Descargas: {nombres}"
            page.update()
        except Exception as ex:
            estado_o365_text.value = f"Error al descargar adjuntos: {ex}"
            page.update()

    def celda_o365(texto: str) -> ft.DataCell:
        return ft.DataCell(ft.Text(texto, color=ft.Colors.ON_SURFACE))

    def boton_descargar_o365(correo: CorreoResumen) -> ft.Control:
        if not correo.tiene_adjuntos:
            return ft.Text("-", color=ft.Colors.ON_SURFACE_VARIANT)
        return ft.IconButton(
            icon=ft.Icons.DOWNLOAD,
            tooltip="Descargar adjunto(s) (CSV se carga en Conciliación; otros se guardan en Descargas)",
            icon_color=NAVY,
            on_click=lambda e, c=correo: page.run_task(on_descargar_adjunto, e, c),
        )

    # --- Detalle de correo (cuerpo completo + visor de adjuntos) ---
    detalle_titulo_text = ft.Text("", weight=ft.FontWeight.BOLD, size=16)
    detalle_meta_text = ft.Text("", color=ft.Colors.ON_SURFACE_VARIANT, size=12)
    detalle_cuerpo_text = ft.Text("", selectable=True)
    detalle_adjuntos_column = ft.Column([], spacing=8)

    def on_cerrar_detalle(_e) -> None:
        page.pop_dialog()

    dialogo_detalle = ft.AlertDialog(
        modal=True,
        title=detalle_titulo_text,
        content=ft.Column(
            [
                detalle_meta_text,
                ft.Divider(),
                detalle_cuerpo_text,
                ft.Divider(),
                ft.Text("Adjuntos:", weight=ft.FontWeight.BOLD),
                detalle_adjuntos_column,
            ],
            tight=True,
            spacing=10,
            width=560,
            height=520,
            scroll=ft.ScrollMode.ALWAYS,
        ),
        actions=[ft.TextButton("Cerrar", on_click=on_cerrar_detalle)],
    )

    def control_adjunto(ruta: str) -> ft.Control:
        nombre = os.path.basename(ruta)
        extension = os.path.splitext(nombre)[1].lower()
        if extension in EXTENSIONES_IMAGEN:
            with open(ruta, "rb") as f:
                datos = f.read()
            return ft.Column(
                [
                    ft.Text(nombre, size=12),
                    ft.Image(src=datos, width=480, fit=ft.BoxFit.CONTAIN),
                ],
                spacing=4,
            )
        return ft.Row(
            [
                ft.Text(nombre, expand=True),
                ft.IconButton(
                    icon=ft.Icons.OPEN_IN_NEW,
                    tooltip="Abrir con la aplicación predeterminada del sistema",
                    icon_color=NAVY,
                    on_click=lambda _e, r=ruta: abrir_archivo_con_app_predeterminada(r),
                ),
            ],
        )

    async def on_ver_detalle(_e, correo: CorreoResumen) -> None:
        detalle_titulo_text.value = correo.asunto
        fecha_texto = correo.fecha.strftime("%d/%m/%Y %H:%M") if correo.fecha else "-"
        detalle_meta_text.value = f"De: {correo.remitente} · {fecha_texto}"
        detalle_cuerpo_text.value = "Cargando..."
        detalle_adjuntos_column.controls = []
        mostrar_dialogo(dialogo_detalle)
        page.update()

        try:
            cuerpo = await asyncio.to_thread(obtener_cuerpo, correo.mensaje)
            detalle_cuerpo_text.value = cuerpo or "(sin contenido)"

            if correo.tiene_adjuntos:
                destino_dir = tempfile.mkdtemp(prefix="mh_cobranza_o365_detalle_")
                rutas = await asyncio.to_thread(descargar_adjuntos, correo.mensaje, destino_dir)
                detalle_adjuntos_column.controls = [control_adjunto(ruta) for ruta in rutas]
            else:
                detalle_adjuntos_column.controls = [ft.Text("(sin adjuntos)", color=ft.Colors.ON_SURFACE_VARIANT)]
        except Exception as ex:
            detalle_cuerpo_text.value = f"Error al cargar el correo: {ex}"
        page.update()

    def boton_ver_detalle(correo: CorreoResumen) -> ft.Control:
        return ft.IconButton(
            icon=ft.Icons.VISIBILITY,
            tooltip="Ver mensaje completo",
            icon_color=NAVY,
            on_click=lambda e, c=correo: page.run_task(on_ver_detalle, e, c),
        )

    filtro_o365_concepto = ft.TextField(
        label="Filtrar por concepto/asunto",
        prefix_icon=ft.Icons.SEARCH,
        dense=True,
        width=320,
        on_change=lambda _e: refrescar_tabla_o365(),
        border=ft.InputBorder.UNDERLINE,
        border_color=BORDER
    )
    # Filtro de fecha vía DatePicker (en vez de teclear la fecha completa).
    fecha_filtro_sel: list[Optional[date]] = [None]
    texto_fecha_filtro = ft.Text("", color=ft.Colors.ON_SURFACE)

    def on_cambiar_fecha_filtro(e) -> None:
        val = e.control.value  # datetime seleccionado (o None)
        fecha_filtro_sel[0] = val.date() if val else None
        texto_fecha_filtro.value = val.strftime("%d/%m/%Y") if val else ""
        refrescar_tabla_o365()

    date_picker_o365 = ft.DatePicker(
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2035, 12, 31),
        on_change=on_cambiar_fecha_filtro,
    )

    boton_filtro_fecha = ft.OutlinedButton(
        "Filtrar por fecha",
        icon=ft.Icons.CALENDAR_MONTH,
        on_click=lambda _e: page.show_dialog(date_picker_o365),
    )

    def limpiar_filtro_fecha(_e) -> None:
        fecha_filtro_sel[0] = None
        date_picker_o365.value = None
        texto_fecha_filtro.value = ""
        refrescar_tabla_o365()

    boton_limpiar_fecha = ft.IconButton(
        icon=ft.Icons.CLOSE,
        tooltip="Quitar filtro de fecha",
        on_click=limpiar_filtro_fecha,
    )

    def refrescar_tabla_o365() -> None:
        f_concepto = (filtro_o365_concepto.value or "").strip().lower()
        f_fecha = fecha_filtro_sel[0]

        def pasa(correo: CorreoResumen) -> bool:
            if f_concepto and f_concepto not in (correo.asunto or "").lower():
                return False
            if f_fecha is not None:
                if not correo.fecha or correo.fecha.date() != f_fecha:
                    return False
            return True

        visibles = [c for c in correos_o365 if pasa(c)]
        tabla_o365.rows = [
            ft.DataRow(
                cells=[
                    ft.DataCell(
                        ft.Checkbox(
                            value=correo.id_correo in correos_seleccionados,
                            on_change=lambda e, c=correo: on_toggle_correo(c, e.control.value),
                        )
                    ),
                    celda_o365(correo.fecha.strftime("%d/%m/%Y %H:%M") if correo.fecha else "-"),
                    celda_o365(correo.remitente),
                    celda_o365(correo.asunto[:80]),
                    celda_o365("Sí" if correo.tiene_adjuntos else "No"),
                    ft.DataCell(
                        ft.Row([boton_ver_detalle(correo), boton_descargar_o365(correo)], spacing=0)
                    ),
                ],
            )
            for correo in visibles
        ]
        page.update()

    correos_vistos_ids: set[str] = set()
    primera_carga_o365 = [True]

    def notificar(mensaje: str) -> None:
        notificacion_snackbar.content = ft.Text(mensaje)
        notificacion_snackbar.open = True
        page.update()

    async def actualizar_bandeja_o365(_e=None) -> None:
        nonlocal correos_o365
        estado_o365_text.value = "Conectando con el buzón..."
        page.update()
        try:
            cuenta = await asyncio.to_thread(obtener_cuenta)
            if cuenta is None:
                estado_o365_text.value = (
                    "No autenticado. Corre `python auth_o365.py` en una terminal, en la "
                    "raíz del proyecto, para generar el token y vuelve a intentar."
                )
                correos_o365 = []
                refrescar_tabla_o365()
                page.update()
                return

            correos_o365 = await asyncio.to_thread(listar_correos, cuenta, 150, "Contado")
            # Por default se marcan todos para extraer; el usuario desmarca.
            correos_seleccionados.clear()
            correos_seleccionados.update(c.id_correo for c in correos_o365)
            check_todos_o365.value = bool(correos_o365)
            estado_o365_text.value = (
                f"{len(correos_o365)} correo(s) con 'Contado' en el asunto cargado(s) de la bandeja de entrada."
            )
            refrescar_tabla_o365()

            if primera_carga_o365[0]:
                # Primera carga: solo establece la línea base, sin disparar
                # extracción/notificación para correos que ya estaban ahí.
                correos_vistos_ids.update(c.id_correo for c in correos_o365)
                primera_carga_o365[0] = False
            else:
                nuevos = [c for c in correos_o365 if c.id_correo not in correos_vistos_ids]
                for correo in nuevos:
                    correos_vistos_ids.add(correo.id_correo)
                    pago = await extraer_un_pago_contado(correo)
                    pagos_contado.append(pago)
                    notificar(f"Pago de contado detectado y extraído: {correo.asunto[:80]}")
                if nuevos:
                    refrescar_tabla_pagos_contado()
        except Exception as ex:
            estado_o365_text.value = f"Error al consultar el buzón: {ex}"
        page.update()

    boton_actualizar_o365 = ft.Button(
        "Actualizar bandeja",
        icon=ft.Icons.REFRESH,
        on_click=actualizar_bandeja_o365,
        bgcolor=NAVY,
        color=ft.Colors.WHITE,
    )

    # ──────────────────────────────────────────────────────
    # Extracción de "Pagos de Contado" desde los correos filtrados
    # (Fecha de Operación y Cuenta Bancaria se declaran una sola vez,
    # arriba, en la pestaña Conciliación Bancaria — igual que en SIPP)
    # ──────────────────────────────────────────────────────
    pagos_contado: list[PagoContadoExtraido] = []
    pagos_progreso_text = ft.Text("")

    pago_a_identificar: list[Optional[PagoContadoExtraido]] = [None]
    pago_cliente_info_text = ft.Text("")
    pago_cliente_autocomplete = ft.AutoComplete(suggestions=sugerencias_clientes)

    def on_select_cliente_pago(e: ft.AutoCompleteSelectEvent) -> None:
        pago = pago_a_identificar[0]
        if pago is None:
            return
        pago.cliente_match = e.selection.value
        page.pop_dialog()
        refrescar_tabla_pagos_contado()

    pago_cliente_autocomplete.on_select = on_select_cliente_pago

    def on_cancelar_pago_cliente(_e) -> None:
        page.pop_dialog()

    dialogo_pago_cliente = ft.AlertDialog(
        modal=True,
        title=ft.Text("Seleccionar cliente"),
        content=ft.Column(
            [pago_cliente_info_text, pago_cliente_autocomplete],
            tight=True,
            spacing=12,
            width=400,
        ),
        actions=[ft.TextButton("Cancelar", on_click=on_cancelar_pago_cliente)],
    )

    def abrir_dialogo_pago_cliente(pago: PagoContadoExtraido) -> None:
        pago_a_identificar[0] = pago
        pago_cliente_info_text.value = f"{pago.concepto[:90]}\nSugerencia detectada: {pago.cliente_texto or '(ninguna)'}"
        pago_cliente_autocomplete.value = pago.cliente_texto or ""
        mostrar_dialogo(dialogo_pago_cliente)

    texto_adjunto_pago_text = ft.Text("", selectable=True)
    dialogo_texto_adjunto = ft.AlertDialog(
        modal=True,
        title=ft.Text("Texto extraído del comprobante"),
        content=ft.Column(
            [texto_adjunto_pago_text],
            tight=True,
            spacing=12,
            width=560,
            height=420,
            scroll=ft.ScrollMode.ALWAYS,
        ),
        actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
    )

    def abrir_dialogo_texto_adjunto(pago: PagoContadoExtraido) -> None:
        if pago.error and not pago.texto_adjunto:
            texto_adjunto_pago_text.value = pago.error
        else:
            texto_adjunto_pago_text.value = pago.texto_adjunto or "(sin texto extraído)"
        mostrar_dialogo(dialogo_texto_adjunto)

    def quitar_pago_contado(pago: PagoContadoExtraido) -> None:
        if pago in pagos_contado:
            pagos_contado.remove(pago)
        refrescar_tabla_pagos_contado()

    def celda_campo_pago(valor: str, on_change) -> ft.DataCell:
        return ft.DataCell(
            ft.TextField(value=valor, dense=True, content_padding=6, on_change=on_change)
        )

    tabla_pagos_contado = DataTable2(
        columns=[
            columna_o365("Fecha"),
            columna_o365("Concepto"),
            columna_o365("Tipo"),
            columna_o365("Banco"),
            columna_o365("Plaza"),
            columna_o365("Monto"),
            columna_o365("Cliente"),
            columna_o365("Referencia"),
            columna_o365("Cuenta Bancaria"),
            columna_o365("Acciones"),
        ],
        rows=[],
        min_width=1500,
        fixed_top_rows=1,
        column_spacing=12,
        expand=True,
    )

    def _norm_ref(s: str) -> str:
        import re as _re
        return _re.sub(r"[^A-Z0-9]", "", (s or "").upper())

    def cruzar_contado_con_bancos() -> None:
        """Marca los pagos de contado que YA vienen en algún movimiento de una
        extracción bancaria (match por monto y referencia). Esos no se suben por
        el RPA de contado: se identifican en el bloque bancario."""
        mov_banco = [
            (r, d)
            for r in historial_registros
            for d in r.get("movimientos", [])
        ]
        for pago in pagos_contado:
            pago.en_bloque_bancario = False
            pago.bloque_ref = pago.bloque_id = pago.bloque_clave = ""
            if pago.monto is None:
                continue
            por_monto = [(r, d) for (r, d) in mov_banco if abs((d.get("abono") or 0) - pago.monto) <= 0.01]
            if not por_monto:
                continue
            ref = _norm_ref(pago.referencia)
            elegido = None
            if ref:
                for (r, d) in por_monto:
                    desc = _norm_ref((d.get("descripcion", "") or "") + (d.get("referencia", "") or ""))
                    if ref in desc:
                        elegido = (r, d)
                        break
            if elegido is None and len(por_monto) == 1:
                elegido = por_monto[0]  # monto único: match razonable
            if elegido:
                r, d = elegido
                pago.en_bloque_bancario = True
                pago.bloque_id = r.get("id", "")
                pago.bloque_clave = clave_movimiento_dict(d)
                pago.bloque_ref = f"{r.get('banco','')} {r.get('hora','')} · ref {d.get('referencia','')}"

    def identificar_pago_en_bloque(pago: PagoContadoExtraido) -> None:
        """Aplica el cliente del pago de contado al movimiento bancario del bloque
        coincidente (en el historial y, si está cargado, en pantalla)."""
        if not (pago.bloque_id and pago.cliente_match and pago.bloque_clave):
            estado_o365_text.value = "Falta cliente en el pago para identificar en el bloque."
            page.update()
            return
        for r in historial_registros:
            if r.get("id") != pago.bloque_id:
                continue
            for d in r.get("movimientos", []):
                if clave_movimiento_dict(d) == pago.bloque_clave:
                    d["cliente_match"] = pago.cliente_match
                    d["identificado_manual"] = True
                    break
            r["num_identificados"] = sum(1 for d in r.get("movimientos", []) if d.get("cliente_match"))
            break
        guardar_historial(HISTORIAL_PATH, historial_registros)

        # Si esa extracción está cargada en pantalla, reflejarlo al instante.
        if historial_id_actual[0] == pago.bloque_id:
            for m in movimientos:
                if clave_movimiento(m) == pago.bloque_clave:
                    m.cliente_match = pago.cliente_match
                    m.identificado_manual = True
                    break
            refrescar_resumen()
            refrescar_tabla()
        estado_o365_text.value = (
            f"'{pago.cliente_match}' identificado en el bloque bancario ({pago.bloque_ref}). "
            "Este pago no se subirá por Contado."
        )
        refrescar_tabla_pagos_contado()

    def refrescar_tabla_pagos_contado() -> None:
        cruzar_contado_con_bancos()
        filas = []
        for pago in pagos_contado:
            fecha_texto = pago.correo.fecha.strftime("%d/%m/%Y") if pago.correo.fecha else "-"

            def on_change_concepto(e, p=pago):
                p.concepto = e.control.value

            def on_change_plaza(e, p=pago):
                p.plaza = e.control.value

            def on_change_monto(e, p=pago):
                p.monto = _parsear_monto_pago(e.control.value)

            def on_change_referencia(e, p=pago):
                p.referencia = e.control.value

            def on_change_tipo(e, p=pago):
                p.tipo_movimiento = e.control.value

            def on_change_cuenta(e, p=pago):
                p.cuenta_bancaria = e.control.value or ""

            # Hereda la cuenta del encabezado (Conciliación Bancaria) si la fila
            # aún no tiene una; el usuario puede sobrescribirla por fila.
            if not pago.cuenta_bancaria and cuenta_contado_dropdown.value:
                pago.cuenta_bancaria = cuenta_contado_dropdown.value

            cuenta_dropdown = ft.DataCell(
                ft.Dropdown(
                    value=pago.cuenta_bancaria or None,
                    dense=True,
                    editable=True,        # campo escribible
                    enable_filter=True,   # filtra la lista mientras escribes
                    menu_height=300,      # tope de altura: el resto hace scroll
                    options=[
                        ft.dropdown.Option(key=c.id_sipp, text=c.nombre)
                        for c in empresa_contado_ref[0].cuentas
                    ],
                    on_select=on_change_cuenta,
                )
            )

            tipo_dropdown = ft.DataCell(
                ft.Dropdown(
                    value=pago.tipo_movimiento or None,
                    dense=True,
                    options=[
                        ft.dropdown.Option(""),
                        ft.dropdown.Option("Anticipo"),
                        ft.dropdown.Option("Contado"),
                    ],
                    on_select=on_change_tipo,
                )
            )

            acciones = ft.DataCell(
                ft.Row(
                    [
                        *(
                            [ft.IconButton(
                                icon=ft.Icons.LINK,
                                icon_color=ft.Colors.ORANGE_700,
                                tooltip=(f"Ya viene en un bloque bancario ({pago.bloque_ref}). "
                                         "No se subirá por Contado. Clic para identificar el cliente en el bloque."),
                                on_click=lambda _e, p=pago: identificar_pago_en_bloque(p),
                            )]
                            if pago.en_bloque_bancario else []
                        ),
                        ft.IconButton(
                            icon=ft.Icons.PERSON_SEARCH,
                            tooltip="Seleccionar cliente",
                            icon_color=NAVY,
                            on_click=lambda _e, p=pago: abrir_dialogo_pago_cliente(p),
                        ),
                        ft.IconButton(
                            icon=ft.Icons.ARTICLE,
                            tooltip="Ver texto extraído del comprobante",
                            icon_color=NAVY,
                            on_click=lambda _e, p=pago: abrir_dialogo_texto_adjunto(p),
                        ),
                        ft.IconButton(
                            icon=ft.Icons.DELETE_OUTLINE,
                            tooltip="Quitar de la lista",
                            icon_color=ft.Colors.RED_600,
                            on_click=lambda _e, p=pago: quitar_pago_contado(p),
                        ),
                    ],
                    spacing=0,
                )
            )

            filas.append(
                ft.DataRow(
                    color=(ft.Colors.with_opacity(0.10, ft.Colors.ORANGE) if pago.en_bloque_bancario else None),
                    cells=[
                        celda_o365(fecha_texto),
                        celda_campo_pago(pago.concepto, on_change_concepto),
                        tipo_dropdown,
                        celda_o365(pago.banco_detectado or "-"),
                        celda_campo_pago(pago.plaza, on_change_plaza),
                        celda_campo_pago(f"{pago.monto:.2f}" if pago.monto is not None else "", on_change_monto),
                        celda_o365(pago.cliente_match or f"(sugerido: {pago.cliente_texto})" if pago.cliente_texto else (pago.cliente_match or "-")),
                        celda_campo_pago(pago.referencia, on_change_referencia),
                        cuenta_dropdown,
                        acciones,
                    ],
                )
            )
        tabla_pagos_contado.rows = filas
        page.update()

    async def extraer_un_pago_contado(correo: CorreoResumen) -> PagoContadoExtraido:
        try:
            cuerpo = await asyncio.to_thread(obtener_cuerpo, correo.mensaje)
            pago = extraer_pago_contado(correo, cuerpo, clientes_normalizados, sucursales_catalogo)
            destino_dir = tempfile.mkdtemp(prefix="mh_cobranza_pago_contado_")
            await asyncio.to_thread(completar_con_adjunto, pago, destino_dir)
        except Exception as ex:
            pago = PagoContadoExtraido(correo=correo, concepto=correo.asunto, error=f"Error al extraer: {ex}")
        return pago

    async def on_click_extraer_pagos_contado(_e) -> None:
        ids_extraidos = {p.correo.id_correo for p in pagos_contado}
        nuevos = [
            c for c in correos_o365
            if c.id_correo in correos_seleccionados and c.id_correo not in ids_extraidos
        ]
        if not nuevos:
            pagos_progreso_text.value = (
                "No hay correos marcados por extraer (marca al menos uno; los ya extraídos se omiten)."
            )
            page.update()
            return

        for i, correo in enumerate(nuevos):
            pagos_progreso_text.value = f"Extrayendo {i + 1}/{len(nuevos)}: {correo.asunto[:60]}..."
            page.update()
            pagos_contado.append(await extraer_un_pago_contado(correo))
            refrescar_tabla_pagos_contado()

        pagos_progreso_text.value = f"Extracción completa: {len(pagos_contado)} pago(s) de contado en revisión."
        page.update()

    boton_extraer_pagos_contado = ft.Button(
        "Extraer pagos de contado",
        icon=ft.Icons.FIND_IN_PAGE,
        on_click=lambda e: page.run_task(on_click_extraer_pagos_contado, e),
        bgcolor=ORANGE,
        color=ft.Colors.WHITE,
    )

    # --- Carga de Pagos de Contado a SIPP (reutiliza Fecha/Cuenta Bancaria
    # declarados arriba, en la pestaña Conciliación Bancaria) ---
    enviar_automaticamente_switch = ft.Switch(
        label="Enviar automáticamente (Guardar y Enviar) sin pausa para revisión",
        value=False,
    )
    pagos_sipp_usuario_field = ft.TextField(label="Usuario SIPP", autofocus=True)
    pagos_sipp_password_field = ft.TextField(label="Contraseña SIPP", password=True, can_reveal_password=True)
    pagos_sipp_recordar_check = ft.Checkbox(label="Recordar credenciales en este equipo")
    pagos_sipp_progreso_text = ft.Text("")
    pagos_sipp_log_panel, pagos_sipp_log_fn, pagos_sipp_log_reset = crear_panel_log_rpa()
    pagos_sipp_ver_navegador_check = crear_check_ver_navegador()

    def on_cancelar_pagos_sipp(_e) -> None:
        page.pop_dialog()

    async def on_confirmar_pagos_sipp(_e) -> None:
        usuario = (pagos_sipp_usuario_field.value or "").strip()
        password = pagos_sipp_password_field.value or ""
        if not usuario or not password:
            pagos_sipp_progreso_text.value = "Usuario y contraseña son obligatorios."
            page.update()
            return

        if pagos_sipp_recordar_check.value:
            guardar_credenciales(usuario, password)
        else:
            borrar_credenciales()

        boton_pagos_sipp_confirmar.disabled = True
        # "Cerrar" habilitado: cerrar el modal deja el RPA en segundo plano (FAB).
        pagos_sipp_log_reset()
        auto = enviar_automaticamente_switch.value
        # En modo envío automático no hay revisión manual → puede correr silencioso.
        # En modo revisión se necesita el navegador visible para revisar las pestañas.
        headless = bool(auto) and not ver_navegador_ref[0]
        pagos_sipp_progreso_text.value = (
            "Ejecutando RPA en segundo plano..." if headless
            else "Ejecutando RPA... (se abrirá el navegador para revisar)"
        )
        rpa_inicio(dialogo_pagos_sipp)
        page.update()

        def log_fn(mensaje: str, nivel: str = "info") -> None:
            pagos_sipp_progreso_text.value = mensaje
            pagos_sipp_log_fn(mensaje, nivel)

        # La plaza es OPCIONAL: un pago sin plaza igual se envía para que el RPA
        # verifique si ya es duplicado (y lo omita). Si no es duplicado y no trae
        # plaza, el RPA lo salta con aviso (contado necesita sucursal para crearlo).
        confirmados = [p for p in pagos_contado if p.cliente_match and p.monto is not None and not p.en_bloque_bancario]
        ok_rpa = True
        try:
            enviados, duplicados = await cargar_pagos_contado_en_sipp(
                confirmados,
                fecha_operacion_field.value or "",
                usuario,
                password,
                empresa=empresa_contado_ref[0],
                headless=headless,
                log_fn=log_fn,
                enviar_automaticamente=auto,
            )
            nuevos = enviados - len(duplicados)
            mensaje = f"Carga a SIPP (Pagos de Contado) lista: {nuevos} nuevo(s)"
            if duplicados:
                mensaje += f", {len(duplicados)} omitido(s) por duplicado"
            mensaje += "."
            estado_o365_text.value = mensaje
            log_fn(mensaje, "ok")
            if duplicados:
                mostrar_dialogo_duplicados_contado(duplicados)
        except Exception as ex:
            ok_rpa = False
            estado_o365_text.value = f"Error al cargar Pagos de Contado en SIPP: {ex}"
            log_fn(f"Error: {ex}", "error")
        finally:
            boton_pagos_sipp_confirmar.disabled = False
            boton_pagos_sipp_cancelar.disabled = False
            pagos_sipp_password_field.value = ""
            rpa_fin(ok_rpa)
            page.update()

    def mostrar_dialogo_duplicados_contado(duplicados: list) -> None:
        """Resumen visual de los pagos que el RPA omitió por ya estar subidos,
        indicando en qué conciliación (folio) se encontró cada uno."""
        filas = []
        for d in duplicados:
            suc = d.get("sucursal") or "—"
            filas.append(
                ft.Container(
                    padding=10,
                    border_radius=8,
                    bgcolor=ft.Colors.with_opacity(0.08, ft.Colors.ORANGE),
                    content=ft.Row(
                        [
                            ft.Icon(ft.Icons.WARNING_AMBER_ROUNDED, color=ORANGE),
                            ft.Column(
                                [
                                    ft.Text(d.get("cliente", ""), weight=ft.FontWeight.BOLD,
                                            color=ft.Colors.ON_SURFACE),
                                    ft.Text(f"${d.get('monto', 0):,.2f}  ·  sucursal {suc}",
                                            size=12, color=ft.Colors.ON_SURFACE_VARIANT),
                                ],
                                spacing=2, expand=True,
                            ),
                            ft.Container(
                                content=ft.Text(f"Conciliación {d.get('folio', '?')}",
                                                color=ft.Colors.WHITE, size=12, weight=ft.FontWeight.BOLD),
                                bgcolor=NAVY,
                                padding=ft.Padding.symmetric(horizontal=10, vertical=6),
                                border_radius=12,
                            ),
                        ],
                        spacing=10,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                )
            )
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Row(
                [ft.Icon(ft.Icons.CONTENT_COPY, color=ORANGE),
                 ft.Text("Pagos omitidos por duplicado")],
                spacing=8,
            ),
            content=ft.Container(
                content=ft.Column(
                    [
                        ft.Text(
                            f"{len(duplicados)} pago(s) ya estaban subidos en SIPP y se omitieron:",
                            size=13,
                        ),
                        ft.Divider(),
                        *filas,
                    ],
                    spacing=10, scroll=ft.ScrollMode.AUTO, tight=True,
                ),
                width=560,
                height=min(140 + len(duplicados) * 72, 460),
            ),
            actions=[ft.TextButton("Entendido", on_click=lambda _e: page.pop_dialog())],
        )
        mostrar_dialogo(dlg)

    boton_pagos_sipp_cancelar = ft.TextButton("Cerrar", on_click=on_cancelar_pagos_sipp)
    boton_pagos_sipp_confirmar = ft.Button(
        "Cargar", on_click=on_confirmar_pagos_sipp, bgcolor=NAVY, color=ft.Colors.WHITE
    )

    dialogo_pagos_sipp = ft.AlertDialog(
        modal=True,
        title=ft.Text("Cargar Pagos de Contado a SIPP"),
        content=ft.Column(
            [
                pagos_sipp_progreso_text,
                pagos_sipp_usuario_field,
                pagos_sipp_password_field,
                pagos_sipp_recordar_check,
                pagos_sipp_ver_navegador_check,
                pagos_sipp_log_panel,
            ],
            tight=True,
            spacing=12,
            width=420,
        ),
        actions=[boton_pagos_sipp_cancelar, boton_pagos_sipp_confirmar],
    )

    def on_click_cargar_pagos_sipp(_e) -> None:
        if not pagos_contado:
            estado_o365_text.value = "Primero extrae los pagos de contado."
            page.update()
            return
        if not (fecha_operacion_field.value or "").strip():
            estado_o365_text.value = "Indica la Fecha de Operación, arriba en Conciliación Bancaria."
            page.update()
            return
        confirmados = [p for p in pagos_contado if p.cliente_match and p.monto is not None and not p.en_bloque_bancario]
        if not confirmados:
            estado_o365_text.value = (
                "Ningún pago tiene Cliente + Monto confirmados todavía; revísalos en la tabla."
            )
            page.update()
            return
        if any(not p.cuenta_bancaria for p in confirmados):
            estado_o365_text.value = (
                "Asigna la Cuenta Bancaria a cada pago confirmado en la columna 'Cuenta Bancaria' de la tabla."
            )
            page.update()
            return

        usuario_guardado, password_guardado = cargar_credenciales()
        pagos_sipp_usuario_field.value = usuario_guardado
        pagos_sipp_password_field.value = password_guardado
        pagos_sipp_recordar_check.value = bool(usuario_guardado)
        pagos_sipp_progreso_text.value = f"{len(confirmados)}/{len(pagos_contado)} pago(s) listo(s) para enviar."
        mostrar_dialogo(dialogo_pagos_sipp)

    boton_cargar_pagos_sipp = ft.Button(
        "Cargar Pagos de Contado a SIPP",
        icon=ft.Icons.ACCOUNT_BALANCE,
        on_click=on_click_cargar_pagos_sipp,
        bgcolor=NAVY,
        color=ft.Colors.WHITE,
    )

    contenido_o365 = ft.Container(
        content=ft.Column(
            [
                ft.Row([boton_actualizar_o365], spacing=16),
                estado_o365_text,
                ft.Row(
                    [
                        boton_filtro_fecha,
                        texto_fecha_filtro,
                        boton_limpiar_fecha,
                        filtro_o365_concepto,
                    ],
                    spacing=12,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                ft.Divider(),
                # Altura fija: el buzón hace scroll INTERNO (no empuja la sección de
                # extracción cuando hay muchos correos).
                ft.Container(content=tabla_o365, height=380),
                ft.Divider(),
                ft.Text("Pagos de Contado (revisar antes de cargar a SIPP)", weight=ft.FontWeight.BOLD, size=16),
                ft.Row(
                    [empresa_contado_dropdown, cuenta_contado_dropdown],
                    spacing=16,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                ft.Row([boton_extraer_pagos_contado], spacing=16),
                pagos_progreso_text,
                ft.Container(content=tabla_pagos_contado, expand=True),
                ft.Row([enviar_automaticamente_switch, boton_cargar_pagos_sipp], spacing=16),
            ],
            spacing=12,
            scroll=ft.ScrollMode.ALWAYS,
        ),
        padding=20,
        expand=True,
    )

    # ──────────────────────────────────────────────────────
    # Pestaña: Catálogos
    # ──────────────────────────────────────────────────────
    # --- Catálogo Cuentas-Clientes (CRUD completo) ---
    filtro_catalogo_token = [0]

    async def _filtrar_catalogo_con_debounce(token: int) -> None:
        await asyncio.sleep(0.3)
        if token == filtro_catalogo_token[0]:
            refrescar_tabla_catalogo()

    def on_cambio_filtro_catalogo(_e) -> None:
        filtro_catalogo_token[0] += 1
        page.run_task(_filtrar_catalogo_con_debounce, filtro_catalogo_token[0])

    filtro_catalogo_texto = ft.TextField(
        label="Buscar cuenta, cliente, banco o plaza",
        width=400,
        color=ft.Colors.ON_SURFACE,
        on_change=on_cambio_filtro_catalogo,
        on_submit=lambda _e: refrescar_tabla_catalogo(),
        border=ft.InputBorder.UNDERLINE,
        border_color=BORDER,
    )
    catalogo_estado_text = ft.Text("")
    LIMITE_RESULTADOS_CATALOGO = 200

    def columna_catalogos(texto: str) -> DataColumn2:
        return DataColumn2(ft.Text(texto, weight=ft.FontWeight.BOLD, color=ft.Colors.ON_SURFACE))

    tabla_catalogo = DataTable2(
        columns=[
            columna_catalogos("Cuenta"),
            columna_catalogos("Cliente"),
            columna_catalogos("Banco"),
            columna_catalogos("Plaza"),
            columna_catalogos("Acciones"),
        ],
        rows=[],
        min_width=800,
        fixed_top_rows=1,
        column_spacing=16,
        expand=True,
    )

    item_catalogo_editando: list[Optional[ClienteCuenta]] = [None]
    campo_catalogo_cuenta = ft.TextField(label="Cuenta")
    campo_catalogo_cliente = ft.TextField(label="Cliente")
    campo_catalogo_banco = ft.TextField(label="Banco")
    campo_catalogo_plaza = ft.TextField(label="Plaza")

    def aplicar_filtro_catalogo() -> list[ClienteCuenta]:
        texto = (filtro_catalogo_texto.value or "").strip().lower()
        if not texto:
            return catalogo
        return [
            c for c in catalogo
            if texto in c.cuenta.lower()
            or texto in c.cliente.lower()
            or texto in c.banco.lower()
            or texto in c.plaza.lower()
        ]

    def on_cerrar_dialogo_catalogo(_e) -> None:
        page.pop_dialog()

    def on_guardar_item_catalogo(_e) -> None:
        cuenta = (campo_catalogo_cuenta.value or "").strip()
        cliente = (campo_catalogo_cliente.value or "").strip()
        banco = (campo_catalogo_banco.value or "").strip()
        plaza = (campo_catalogo_plaza.value or "").strip()
        if not cuenta or not cliente:
            catalogo_estado_text.value = "Cuenta y Cliente son obligatorios."
            page.update()
            return

        item = item_catalogo_editando[0]
        if item is None:
            if any(c.cuenta == cuenta for c in catalogo):
                catalogo_estado_text.value = f"Ya existe una cuenta '{cuenta}' en el catálogo."
                page.update()
                return
            catalogo.append(ClienteCuenta(cuenta=cuenta, cliente=cliente, banco=banco, plaza=plaza))
        else:
            item.cuenta, item.cliente, item.banco, item.plaza = cuenta, cliente, banco, plaza

        guardar_catalogo_completo(CATALOGO_PATH, catalogo)
        catalogo_info_text.value = f"Catálogo de clientes cargado: {len(catalogo)} cuentas"
        catalogo_estado_text.value = "Catálogo actualizado y guardado."
        page.pop_dialog()
        refrescar_tabla_catalogo()

    dialogo_catalogo = ft.AlertDialog(
        modal=True,
        title=ft.Text("Cuenta del catálogo"),
        content=ft.Column(
            [campo_catalogo_cuenta, campo_catalogo_cliente, campo_catalogo_banco, campo_catalogo_plaza],
            tight=True,
            spacing=12,
            width=350,
        ),
        actions=[
            ft.TextButton("Cancelar", on_click=on_cerrar_dialogo_catalogo),
            ft.Button("Guardar", on_click=on_guardar_item_catalogo, bgcolor=NAVY, color=ft.Colors.WHITE),
        ],
    )

    def abrir_dialogo_catalogo(item: Optional[ClienteCuenta]) -> None:
        item_catalogo_editando[0] = item
        campo_catalogo_cuenta.value = item.cuenta if item else ""
        campo_catalogo_cliente.value = item.cliente if item else ""
        campo_catalogo_banco.value = item.banco if item else ""
        campo_catalogo_plaza.value = item.plaza if item else ""
        dialogo_catalogo.title = ft.Text("Editar cuenta" if item else "Agregar cuenta")
        mostrar_dialogo(dialogo_catalogo)

    def eliminar_item_catalogo(item: ClienteCuenta) -> None:
        catalogo.remove(item)
        guardar_catalogo_completo(CATALOGO_PATH, catalogo)
        catalogo_info_text.value = f"Catálogo de clientes cargado: {len(catalogo)} cuentas"
        catalogo_estado_text.value = f"Cuenta '{item.cuenta}' eliminada."
        refrescar_tabla_catalogo()

    def refrescar_tabla_catalogo() -> None:
        todas = aplicar_filtro_catalogo()
        truncado = len(todas) > LIMITE_RESULTADOS_CATALOGO
        filas = todas[:LIMITE_RESULTADOS_CATALOGO]
        catalogo_estado_text.value = (
            f"Mostrando los primeros {LIMITE_RESULTADOS_CATALOGO} de {len(todas)} "
            "resultados; refina la búsqueda para ver otros." if truncado else f"{len(filas)} resultado(s)."
        )
        tabla_catalogo.rows = [
            ft.DataRow(
                cells=[
                    celda_o365(item.cuenta),
                    celda_o365(item.cliente),
                    celda_o365(item.banco),
                    celda_o365(item.plaza),
                    ft.DataCell(
                        ft.Row(
                            [
                                ft.IconButton(
                                    icon=ft.Icons.EDIT,
                                    tooltip="Editar",
                                    icon_color=NAVY,
                                    on_click=lambda _e, i=item: abrir_dialogo_catalogo(i),
                                ),
                                ft.IconButton(
                                    icon=ft.Icons.DELETE_OUTLINE,
                                    tooltip="Eliminar",
                                    icon_color=ft.Colors.RED_600,
                                    on_click=lambda _e, i=item: eliminar_item_catalogo(i),
                                ),
                            ],
                            spacing=0,
                        )
                    ),
                ],
            )
            for item in filas
        ]
        page.update()

    boton_agregar_catalogo = ft.Button(
        "Agregar cuenta",
        icon=ft.Icons.ADD,
        on_click=lambda _e: abrir_dialogo_catalogo(None),
        bgcolor=NAVY,
        color=ft.Colors.WHITE,
    )

    refrescar_tabla_catalogo()

    contenido_catalogos = ft.Container(
        content=ft.Column(
            [
                ft.Text("Catálogo Cuentas-Clientes", weight=ft.FontWeight.BOLD, size=16),
                ft.Row([filtro_catalogo_texto, boton_agregar_catalogo], spacing=16),
                catalogo_estado_text,
                ft.Container(content=tabla_catalogo, expand=True),
            ],
            spacing=12,
        ),
        padding=20,
        expand=True,
    )

    def on_click_tema(_e) -> None:
        page.theme_mode = (
            ft.ThemeMode.DARK if page.theme_mode == ft.ThemeMode.LIGHT else ft.ThemeMode.LIGHT
        )
        boton_tema.icon = (
            ft.Icons.LIGHT_MODE if page.theme_mode == ft.ThemeMode.DARK else ft.Icons.DARK_MODE
        )
        page.update()

    boton_tema = ft.IconButton(
        icon=ft.Icons.DARK_MODE,
        icon_color=ft.Colors.WHITE,
        tooltip="Cambiar tema claro/oscuro",
        on_click=on_click_tema,
    )

    # --- Actualizaciones desde GitHub ---
    update_estado_text = ft.Text("")
    update_progress = ft.ProgressRing(width=18, height=18, visible=False)
    update_resumen_text = ft.Text("", size=12, color=ft.Colors.ON_SURFACE_VARIANT, selectable=True)

    async def _aplicar_update(_e=None) -> None:
        boton_aplicar_update.disabled = True
        update_progress.visible = True
        update_estado_text.value = "Descargando actualización..."
        page.update()
        ok, salida = await asyncio.to_thread(aplicar_actualizacion, BASE_DIR)
        if ok:
            update_estado_text.value = "Actualización aplicada. Reiniciando la aplicación..."
            page.update()
            await asyncio.sleep(1.2)
            reiniciar_app(BASE_DIR)  # macOS: reemplaza el proceso (no regresa)
            # Windows: reiniciar_app lanzó una instancia nueva y regresó aquí →
            # hay que CERRAR esta ventana y matar este proceso para que no queden
            # dos. Un hilo aparte fuerza la salida aunque el loop se esté cerrando;
            # window.destroy() cierra la ventana/cliente Flet mientras tanto.
            import threading
            import time as _time

            def _forzar_salida() -> None:
                _time.sleep(1.5)
                os._exit(0)

            threading.Thread(target=_forzar_salida, daemon=True).start()
            try:
                # destroy() es ASÍNCRONO: hay que await, si no el coroutine nunca
                # corre y la ventana vieja se queda abierta.
                await page.window.destroy()
            except Exception:
                pass
        else:
            update_progress.visible = False
            boton_aplicar_update.disabled = False
            update_estado_text.value = "No se pudo actualizar automáticamente."
            update_resumen_text.value = salida
            update_estado_text.color = ft.Colors.RED_600
            page.update()

    boton_aplicar_update = ft.Button(
        "Actualizar y reiniciar",
        icon=ft.Icons.SYSTEM_UPDATE_ALT,
        bgcolor=NAVY,
        color=ft.Colors.WHITE,
        visible=False,
        on_click=lambda e: page.run_task(_aplicar_update, e),
    )

    dialogo_update = ft.AlertDialog(
        modal=True,
        title=ft.Text("Actualizaciones"),
        content=ft.Container(
            content=ft.Column(
                [ft.Row([update_progress, update_estado_text], spacing=8), update_resumen_text],
                tight=True, spacing=10,
            ),
            width=460,
        ),
        actions=[
            ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog()),
            boton_aplicar_update,
        ],
    )

    async def comprobar_actualizaciones(mostrar: bool = True, auto: bool = False) -> None:
        """Consulta GitHub. Si `mostrar`, abre el diálogo con el resultado. Si
        `auto` y hay actualización, la aplica y reinicia sin intervención."""
        if mostrar:
            update_estado_text.value = "Comprobando actualizaciones..."
            update_estado_text.color = ft.Colors.ON_SURFACE
            update_resumen_text.value = ""
            update_progress.visible = True
            boton_aplicar_update.visible = False
            mostrar_dialogo(dialogo_update)
            page.update()

        res = await asyncio.to_thread(revisar_actualizaciones, BASE_DIR)
        update_progress.visible = False

        if res.get("error"):
            if mostrar:
                update_estado_text.value = f"No se pudo comprobar: {res['error']}"
                update_estado_text.color = ft.Colors.ON_SURFACE
                page.update()
            return

        if not res.get("disponible"):
            if mostrar:
                update_estado_text.value = "La aplicación está al día."
                update_estado_text.color = ft.Colors.GREEN_700
                page.update()
            return

        # Hay actualización disponible.
        detras = res.get("detras", 0)
        update_estado_text.value = f"Hay una actualización disponible ({detras} cambio(s))."
        update_estado_text.color = ft.Colors.ON_SURFACE
        update_resumen_text.value = res.get("resumen", "")
        if auto:
            if not mostrar:
                mostrar_dialogo(dialogo_update)
            page.update()
            await _aplicar_update()
        else:
            boton_aplicar_update.visible = True
            page.update()

    boton_update = ft.IconButton(
        icon=ft.Icons.SYNC,
        icon_color=ft.Colors.WHITE,
        tooltip="Comprobar actualizaciones",
        on_click=lambda e: page.run_task(comprobar_actualizaciones, True, False),
    )

    encabezado = ft.Container(
        content=ft.Row(
            [
                ft.Container(
                    content=ft.Image(src="grupopetroil.png", height=36, fit=ft.BoxFit.CONTAIN),
                    bgcolor=ft.Colors.WHITE,
                    padding=ft.Padding.symmetric(horizontal=12, vertical=8),
                    border_radius=8,
                ),
                ft.Column(
                    [
                        ft.Text("MultiHerramienta de Cobranza", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
                        catalogo_info_text,
                    ],
                    spacing=0,
                ),
                ft.Container(expand=True),
                *(
                    [
                        ft.Container(
                            content=ft.Text(
                                "● SIPP PRUEBAS",
                                size=12,
                                weight=ft.FontWeight.BOLD,
                                color=ft.Colors.WHITE,
                            ),
                            bgcolor=ft.Colors.RED_700,
                            padding=ft.Padding.symmetric(horizontal=12, vertical=6),
                            border_radius=20,
                            tooltip="El RPA apunta al entorno de pruebas (stage.sipp.petroil.dev)",
                        )
                    ]
                    if es_modo_test()
                    else []
                ),
                boton_update,
                boton_tema,
            ],
            spacing=16,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        bgcolor=NAVY,
        padding=ft.Padding.symmetric(horizontal=24, vertical=14),
        border=ft.Border(bottom=ft.BorderSide(4, GOLD)),
    )

    def paso_badge(n: int) -> ft.Control:
        return ft.Container(
            content=ft.Text(str(n), color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD, size=14),
            width=26,
            height=26,
            bgcolor=NAVY,
            border_radius=13,
            alignment=ft.Alignment.CENTER,
        )

    _pasos_ayuda = [
        (1, "Carga el archivo bancario (.csv) del banco (Santander / BanRegio)."),
        (2, "Opcional: si quedan movimientos sin cliente, usa 'Buscar folios en SIPP' para identificarlos."),
        (3, "Carga el Estado de Cuenta (.xlsx): sugiere la sucursal de cada movimiento."),
        (4, "Elige Empresa, Cuenta Bancaria (SIPP) y Fecha de Operación."),
        (5, "Presiona 'Cargar a SIPP (Ingresos Diversos)' para enviarlo al RPA."),
    ]
    dialogo_ayuda_csv = ft.AlertDialog(
        modal=True,
        title=ft.Text("¿Cómo funciona la Conciliación Bancaria?"),
        content=ft.Column(
            [
                ft.Row(
                    [paso_badge(n), ft.Text(txt, expand=True)],
                    spacing=12,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                )
                for n, txt in _pasos_ayuda
            ],
            tight=True,
            spacing=14,
            width=540,
        ),
        actions=[ft.TextButton("Entendido", on_click=lambda _e: page.pop_dialog())],
    )
    boton_ayuda_csv = ft.OutlinedButton(
        "¿Cómo funciona?",
        icon=ft.Icons.HELP_OUTLINE,
        on_click=lambda _e: mostrar_dialogo(dialogo_ayuda_csv),
    )

    def fila_paso(n: int, fila: ft.Control) -> ft.Control:
        return ft.Row(
            [paso_badge(n), fila],
            spacing=12,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        )

    # Bloque de controles superiores (pasos + resumen) que se puede ocultar para
    # dar toda la pantalla a la tabla de movimientos.
    bloque_superior_csv = ft.Column(
        [
            ft.Row(
                [
                    paso_badge(1),
                    boton_cargar,
                    boton_buscar_sipp,
                    boton_factoraje,
                    archivo_nombre_text,
                    banco_detectado_text,
                    ft.Container(expand=True),
                    boton_ayuda_csv,
                ],
                spacing=16,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            fila_paso(
                2,
                ft.Row(
                    [boton_cargar_estado_cuenta, estado_cuenta_text],
                    spacing=16,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
            ),
            ft.Divider(),
            fila_paso(
                3,
                ft.Row(
                    [fecha_operacion_field, empresa_dropdown, cuenta_bancaria_dropdown, boton_cargar_ingresos_div],
                    spacing=16,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
            ),
            estado_text,
            ft.Divider(),
            resumen_row,
            ft.Divider(),
        ],
        spacing=12,
    )

    tabla_csv_expandida = [False]

    def on_toggle_expandir_csv(_e) -> None:
        expandida = not tabla_csv_expandida[0]
        tabla_csv_expandida[0] = expandida
        bloque_superior_csv.visible = not expandida
        boton_expandir_csv.icon = ft.Icons.FULLSCREEN_EXIT if expandida else ft.Icons.FULLSCREEN
        boton_expandir_csv.tooltip = "Restaurar vista" if expandida else "Expandir tabla"
        page.update()

    boton_expandir_csv = ft.IconButton(
        icon=ft.Icons.FULLSCREEN,
        tooltip="Expandir tabla",
        icon_color=NAVY,
        on_click=on_toggle_expandir_csv,
    )

    boton_historial = ft.IconButton(
        icon=ft.Icons.HISTORY,
        tooltip="Historial de extracciones",
        icon_color=NAVY,
        on_click=abrir_historial,
    )

    async def exportar_grid_excel(_e) -> None:
        """Descarga el grid (columnas hasta Sucursal) a un archivo .xlsx,
        respetando los filtros vigentes."""
        filas = aplicar_filtros()
        if not filas:
            estado_text.value = "No hay movimientos para exportar."
            page.update()
            return

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        escribir_hoja_movimientos(ws, filas, sucursal_efectiva)

        banco = filas[0].banco or "movimientos"
        nombre_def = f"movimientos_{banco}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        destino = await file_picker.save_file(
            dialog_title="Guardar movimientos en Excel",
            file_name=nombre_def,
            allowed_extensions=["xlsx"],
        )
        if not destino:
            return
        if not destino.lower().endswith(".xlsx"):
            destino += ".xlsx"
        try:
            wb.save(destino)
        except OSError as ex:
            estado_text.value = f"No se pudo guardar el Excel: {ex}"
            page.update()
            return
        estado_text.value = f"Exportados {len(filas)} movimiento(s) a {os.path.basename(destino)}."
        page.update()

    boton_exportar_excel = ft.IconButton(
        icon=ft.Icons.GRID_ON,
        tooltip="Descargar grid a Excel (.xlsx)",
        icon_color=ft.Colors.GREEN_700,
        on_click=lambda e: page.run_task(exportar_grid_excel, e),
    )

    def _leyenda_item(color, texto: str) -> ft.Control:
        """Muestra de color (fiel al tinte de la fila) + descripción, para la
        leyenda de colores del grid."""
        return ft.Row(
            [
                ft.Container(
                    width=14,
                    height=14,
                    bgcolor=color,
                    border_radius=3,
                    border=ft.Border.all(1, ft.Colors.with_opacity(0.25, ft.Colors.ON_SURFACE)),
                ),
                ft.Text(texto, size=11, color=ft.Colors.ON_SURFACE_VARIANT),
            ],
            spacing=6,
            tight=True,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        )

    # Los tintes coinciden con color_fila(): mismo hue, un poco más opaco para que
    # la muestra sea legible fuera de una fila completa.
    leyenda_colores = ft.Row(
        [
            ft.Text("Colores:", size=11, weight=ft.FontWeight.BOLD, color=ft.Colors.ON_SURFACE_VARIANT),
            _leyenda_item(
                ft.Colors.with_opacity(0.30, ft.Colors.RED),
                "Excluido del RPA (traspaso a filiales / portal BBVA)",
            ),
            _leyenda_item(
                ft.Colors.with_opacity(0.30, ft.Colors.AMBER),
                "Folio pendiente de buscar en SIPP",
            ),
            _leyenda_item(
                ft.Colors.with_opacity(0.12, ft.Colors.ON_SURFACE),
                "Ya subido a SIPP (corte anterior, no se re-sube)",
            ),
        ],
        spacing=20,
        wrap=True,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
    )

    contenido_conciliacion = ft.Container(
        content=ft.Column(
            [
                bloque_superior_csv,
                ft.Row(
                    [filtro_estado, filtro_sucursal, filtro_texto, ft.Container(expand=True), boton_exportar_excel, boton_historial, boton_expandir_csv],
                    spacing=16,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                leyenda_colores,
                ft.Container(content=tabla_contenedor, expand=True),
            ],
            spacing=12,
        ),
        padding=20,
        expand=True,
    )

    def on_tabs_change(e) -> None:
        if tabs.selected_index == 1:
            page.run_task(actualizar_bandeja_o365)

    tab_dashboard, contenido_dashboard = construir_tab_dashboard(page)

    tabs = ft.Tabs(
        length=4,
        selected_index=0,
        expand=True,
        on_change=on_tabs_change,
        content=ft.Column(
            expand=True,
            controls=[
                ft.TabBar(
                    tabs=[
                        ft.Tab(label="Identificación Bancaria", icon=ft.Icons.ACCOUNT_BALANCE),
                        ft.Tab(label="Extracción de Contados", icon=ft.Icons.MAIL_OUTLINE),
                        ft.Tab(label="Catálogos", icon=ft.Icons.FOLDER_OPEN),
                        tab_dashboard,
                    ],
                ),
                ft.TabBarView(
                    expand=True,
                    controls=[contenido_conciliacion, contenido_o365, contenido_catalogos, contenido_dashboard],
                ),
            ],
        ),
    )

    notificacion_snackbar = ft.SnackBar(content=ft.Text(""))
    page.overlay.append(notificacion_snackbar)

    async def bucle_refresco_o365() -> None:
        while True:
            await asyncio.sleep(300)
            await actualizar_bandeja_o365()

    page.floating_action_button = fab_rpa
    page.add(
        ft.Column(
            [
                encabezado,
                tabs,
            ],
            expand=True,
            spacing=0,
        )
    )

    page.run_task(bucle_refresco_o365)

    # Al iniciar: avisar y comprobar GitHub en segundo plano; auto-actualizar si
    # hay cambios (sin trabajo abierto que perder al reiniciar).
    async def _comprobar_actualizaciones_inicio() -> None:
        notificacion_snackbar.content = ft.Text("Comprobando actualizaciones...")
        notificacion_snackbar.open = True
        page.update()
        await comprobar_actualizaciones(mostrar=False, auto=True)

    page.run_task(_comprobar_actualizaciones_inicio)


if __name__ == "__main__":
    ft.run(main)
