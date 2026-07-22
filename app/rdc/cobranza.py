"""Panel 'Cobranza' (mitad derecha de la sub-pestaña Proyección): lo
efectivamente cobrado en el periodo seleccionado (por defecto, la semana
anterior a hoy), sobre `Tableros.IgresosClientes`.

Las reglas de negocio y las consultas a BigQuery viven en
`app/services/cobranza_semanal_repository.py` (capa de datos); este módulo
solo arma la UI del panel — ver ese archivo para el detalle de cada regla
(exclusiones de razón social/sucursal, reclasificación de Petroplazas,
conversión de USD, etc.).
"""

import asyncio
from datetime import date, datetime, timedelta

import flet as ft

from ..dashboard.componentes import (
    color_slot,
    encabezado_seccion,
    escribir_hoja_excel,
    guardar_workbook,
    mostrar_dialogo,
    nombre_hoja_valido,
    sombra_tarjeta,
    tile_compacta,
)
from ..services.cobranza_semanal_repository import SEGMENTOS, CobranzaSemanalRepository

# El repositorio se crea perezosamente (necesita credenciales de BigQuery); así
# no falla al importar este módulo si aún no hay credenciales configuradas.
_repo_holder: list[CobranzaSemanalRepository | None] = [None]


def _repo() -> CobranzaSemanalRepository:
    if _repo_holder[0] is None:
        _repo_holder[0] = CobranzaSemanalRepository()
    return _repo_holder[0]


def construir_panel_cobranza(page: ft.Page) -> ft.Control:
    """Contenido del panel de Cobranza (sin Tab propio — vive dentro de la
    mitad derecha de la sub-pestaña Proyección, ver app/rdc/vista.py)."""
    hoy = date.today()
    hace_una_semana = hoy - timedelta(days=7)
    rango_sel: list[tuple[date, date]] = [(hace_una_semana, hoy)]
    segmentos_significativos: list[list[str]] = [[]]  # [] = Todos (los 3 segmentos)

    def _dark() -> bool:
        return page.theme_mode == ft.ThemeMode.DARK

    def _texto_rango(inicio: date, fin: date) -> str:
        return f"{inicio.strftime('%d %b %Y')} – {fin.strftime('%d %b %Y')}"

    titulo = ft.Text("Cobranza", size=20, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE)
    subtitulo = ft.Text(
        "Distribuidora, Asociados y Petroplazas · lo efectivamente cobrado (im_Movimiento) "
        "en el periodo seleccionado.",
        size=12,
        color=ft.Colors.ON_SURFACE_VARIANT,
    )

    estado_text = ft.Text("", size=12, color=ft.Colors.RED_600)
    progress = ft.ProgressRing(width=16, height=16, visible=False, stroke_width=2)

    hero_contenedor = ft.ResponsiveRow(spacing=10, run_spacing=10)
    seccion_detalle = ft.Container()
    seccion_significativos = ft.Container()
    cuerpo = ft.Column(
        [hero_contenedor, seccion_detalle, seccion_significativos],
        spacing=16,
        opacity=1.0,
        animate_opacity=200,
    )

    # En Flet 0.85 el FilePicker es un servicio: se crea y se usa directamente
    # (NO se agrega a page.overlay; hacerlo provoca "Unknown control: FilePicker").
    file_picker = ft.FilePicker()

    # Datos de la última consulta exitosa, para exportar sin recalcular. Se
    # habilita el botón de descarga solo si las 3 consultas (concentrado,
    # significativos, por día) tuvieron éxito a la vez.
    ultimo_datos: dict = {"items": None, "significativos": None, "por_dia": None}

    def _refrescar(resultado, significativos, por_dia) -> None:
        dark = _dark()
        if isinstance(resultado, Exception):
            ultimo_datos["items"] = None
            hero_contenedor.controls = []
            seccion_detalle.content = ft.Container(
                content=ft.Text(f"No se pudo consultar: {resultado}", size=12, color=ft.Colors.RED_600),
                height=120,
                alignment=ft.Alignment.CENTER,
            )
        else:
            por_segmento = {fila["segmento"]: fila for fila in resultado}
            items = [
                (
                    segmento,
                    (por_segmento.get(segmento, {}).get("total_mxn") or 0),
                    (por_segmento.get(segmento, {}).get("total_usd") or 0),
                    (por_segmento.get(segmento, {}).get("total_usd_convertido") or 0),
                    (por_segmento.get(segmento, {}).get("total_usd_sin_tc") or 0),
                )
                for segmento in SEGMENTOS
            ]
            ultimo_datos["items"] = items
            total_mxn = sum(v for _s, v, _u, _c, _st in items)
            total_usd = sum(u for _s, _v, u, _c, _st in items)
            total_convertido = sum(c for _s, _v, _u, c, _st in items)
            total_sin_tc = sum(st for _s, _v, _u, _c, st in items)
            total_final = total_mxn + total_convertido

            hero_contenedor.controls = [
                tile_compacta("Total cobrado", total_final, color_slot(2, dark), ft.Icons.PAYMENTS_OUTLINED,
                               "MXN + USD convertido"),
                tile_compacta("Distribuidora", items[0][1] + items[0][3], color_slot(0, dark),
                               ft.Icons.LOCAL_SHIPPING_OUTLINED),
                tile_compacta("Asociados", items[1][1] + items[1][3], color_slot(1, dark),
                               ft.Icons.HANDSHAKE_OUTLINED),
                tile_compacta("Petroplazas", items[2][1] + items[2][3], color_slot(4, dark),
                               ft.Icons.LOCAL_GAS_STATION_OUTLINED),
            ]

            filas_tabla = []
            for segmento, valor_mxn, valor_usd, convertido, sin_tc in items:
                filas_tabla.append(ft.DataRow(cells=[
                    ft.DataCell(ft.Text(segmento, size=11)),
                    ft.DataCell(ft.Text(f"${valor_mxn:,.2f}", size=11)),
                    ft.DataCell(ft.Text(f"US${valor_usd:,.2f}" if valor_usd else "—", size=11)),
                    ft.DataCell(ft.Text(f"${convertido:,.2f}" if convertido else "—", size=11)),
                    ft.DataCell(ft.Text(f"${valor_mxn + convertido:,.2f}", size=11,
                                        weight=ft.FontWeight.W_600)),
                    ft.DataCell(ft.Text(f"US${sin_tc:,.2f}" if sin_tc else "—", size=11)),
                ]))
            filas_tabla.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text("Total", size=11, weight=ft.FontWeight.W_700)),
                ft.DataCell(ft.Text(f"${total_mxn:,.2f}", size=11, weight=ft.FontWeight.W_700)),
                ft.DataCell(ft.Text(f"US${total_usd:,.2f}" if total_usd else "—", size=11,
                                    weight=ft.FontWeight.W_700)),
                ft.DataCell(ft.Text(f"${total_convertido:,.2f}" if total_convertido else "—", size=11,
                                    weight=ft.FontWeight.W_700)),
                ft.DataCell(ft.Text(f"${total_final:,.2f}", size=11, weight=ft.FontWeight.W_700)),
                ft.DataCell(ft.Text(f"US${total_sin_tc:,.2f}" if total_sin_tc else "—", size=11,
                                    weight=ft.FontWeight.W_700)),
            ]))
            tabla = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("Segmento", size=11)),
                    ft.DataColumn(ft.Text("MXN", size=11), numeric=True),
                    ft.DataColumn(ft.Text("USD", size=11), numeric=True),
                    ft.DataColumn(ft.Text("MXN convertido", size=11), numeric=True),
                    ft.DataColumn(ft.Text("Total final", size=11), numeric=True),
                    ft.DataColumn(ft.Text("USD sin TC", size=11), numeric=True),
                ],
                rows=filas_tabla,
                data_row_max_height=34,
                heading_row_height=34,
                column_spacing=16,
            )
            seccion_detalle.content = ft.Container(
                content=ft.Column(
                    [
                        encabezado_seccion(ft.Icons.TABLE_CHART_OUTLINED, color_slot(3, dark),
                                           "Detalle por segmento", "MXN, USD y su conversión por segmento"),
                        ft.Row([tabla], scroll=ft.ScrollMode.AUTO),
                    ],
                    spacing=10,
                ),
                padding=16,
                bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
                border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
                border_radius=12,
                shadow=sombra_tarjeta(),
            )

        if isinstance(significativos, Exception) or isinstance(por_dia, Exception):
            ultimo_datos["significativos"] = None
            ultimo_datos["por_dia"] = None
            errores = [
                str(error)
                for error in (significativos, por_dia)
                if isinstance(error, Exception)
            ]
            seccion_significativos.content = ft.Container(
                content=ft.Text(f"No se pudo consultar Ingresos Significativos: {' · '.join(errores)}",
                                size=12, color=ft.Colors.RED_600),
                height=120,
                alignment=ft.Alignment.CENTER,
            )
        else:
            ultimo_datos["significativos"] = significativos
            ultimo_datos["por_dia"] = por_dia
            filas_top = []
            for puesto, fila in enumerate(significativos, start=1):
                sin_tc = fila.get("total_usd_sin_tc") or 0
                filas_top.append(ft.DataRow(cells=[
                    ft.DataCell(ft.Text(str(puesto), size=10)),
                    ft.DataCell(ft.Container(
                        ft.Text(fila["razon_social"], size=10, max_lines=1,
                                overflow=ft.TextOverflow.ELLIPSIS, tooltip=fila["razon_social"]),
                        width=210,
                    )),
                    ft.DataCell(ft.Text(fila.get("segmento") or "—", size=10)),
                    ft.DataCell(ft.Text(f"${(fila.get('total_mxn') or 0):,.2f}", size=10)),
                    ft.DataCell(ft.Text(
                        f"US${(fila.get('total_usd') or 0):,.2f}" if fila.get("total_usd") else "—", size=10
                    )),
                    ft.DataCell(ft.Text(
                        f"${(fila.get('total_usd_convertido') or 0):,.2f}"
                        if fila.get("total_usd_convertido") else "—", size=10
                    )),
                    ft.DataCell(ft.Text(f"${(fila.get('total_final') or 0):,.2f}", size=10,
                                        weight=ft.FontWeight.W_600)),
                    ft.DataCell(ft.Text(f"US${sin_tc:,.2f}" if sin_tc else "—", size=10,
                                        color=ft.Colors.RED_600 if sin_tc else ft.Colors.ON_SURFACE_VARIANT)),
                ]))
            tabla_top = ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text("#", size=10), numeric=True),
                    ft.DataColumn(ft.Text("Razón social", size=10)),
                    ft.DataColumn(ft.Text("Tipo de negocio", size=10)),
                    ft.DataColumn(ft.Text("MXN", size=10), numeric=True),
                    ft.DataColumn(ft.Text("USD", size=10), numeric=True),
                    ft.DataColumn(ft.Text("MXN convertido", size=10), numeric=True),
                    ft.DataColumn(ft.Text("Total final", size=10), numeric=True),
                    ft.DataColumn(ft.Text("USD sin TC", size=10), numeric=True),
                ],
                rows=filas_top,
                data_row_max_height=32,
                heading_row_height=34,
                column_spacing=14,
            )

            def _tabla_dia(filas: list[dict]) -> ft.DataTable:
                filas_dia = []
                for fila in filas:
                    sin_tc = fila.get("total_usd_sin_tc") or 0
                    filas_dia.append(ft.DataRow(cells=[
                        ft.DataCell(ft.Text(fila["fecha"].strftime("%d/%m/%Y"), size=10)),
                        ft.DataCell(ft.Text(f"${(fila.get('total_mxn') or 0):,.2f}", size=10)),
                        ft.DataCell(ft.Text(
                            f"US${(fila.get('total_usd') or 0):,.2f}" if fila.get("total_usd") else "—", size=10
                        )),
                        ft.DataCell(ft.Text(
                            f"${(fila.get('total_usd_convertido') or 0):,.2f}"
                            if fila.get("total_usd_convertido") else "—", size=10
                        )),
                        ft.DataCell(ft.Text(f"${(fila.get('total_final') or 0):,.2f}", size=10,
                                            weight=ft.FontWeight.W_600)),
                        ft.DataCell(ft.Text(f"US${sin_tc:,.2f}" if sin_tc else "—", size=10,
                                            color=ft.Colors.RED_600 if sin_tc else ft.Colors.ON_SURFACE_VARIANT)),
                    ]))
                return ft.DataTable(
                    columns=[
                        ft.DataColumn(ft.Text("Fecha", size=10)),
                        ft.DataColumn(ft.Text("MXN", size=10), numeric=True),
                        ft.DataColumn(ft.Text("USD", size=10), numeric=True),
                        ft.DataColumn(ft.Text("MXN convertido", size=10), numeric=True),
                        ft.DataColumn(ft.Text("Total final", size=10), numeric=True),
                        ft.DataColumn(ft.Text("USD sin TC", size=10), numeric=True),
                    ],
                    rows=filas_dia,
                    data_row_max_height=32,
                    heading_row_height=34,
                    column_spacing=14,
                )

            # Con más de un tipo de negocio presente en el periodo, una tabla por
            # fecha con columna "Tipo de negocio" queda difícil de leer (fechas
            # repetidas, una por segmento) — mejor una tabla POR tipo de negocio,
            # cada una ordenada por fecha. Con uno solo (o ninguno), la tabla
            # única de siempre.
            segmentos_en_dia = [s for s in SEGMENTOS if any(f.get("segmento") == s for f in por_dia)]
            if len(segmentos_en_dia) > 1:
                bloques_dia = []
                for segmento in segmentos_en_dia:
                    filas_segmento = [f for f in por_dia if f.get("segmento") == segmento]
                    bloques_dia.append(
                        ft.Column(
                            [
                                ft.Text(segmento, size=12, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE),
                                ft.Row([_tabla_dia(filas_segmento)], scroll=ft.ScrollMode.ALWAYS),
                            ],
                            spacing=6,
                        )
                    )
                contenido_dia: ft.Control = ft.Column(bloques_dia, spacing=14)
            else:
                contenido_dia = ft.Row([_tabla_dia(por_dia)], scroll=ft.ScrollMode.ALWAYS)

            seccion_significativos.content = ft.Container(
                content=ft.Column(
                    [
                        encabezado_seccion(
                            ft.Icons.STAR_OUTLINE, color_slot(4, dark),
                            "Ingresos Significativos", "Top 20 agregado por razón social · ordenado por total final en MXN",
                            [_construir_filtro_segmento()],
                        ),
                        ft.Divider(height=1),
                        ft.Row([tabla_top], scroll=ft.ScrollMode.ALWAYS),
                        ft.Divider(height=1),
                        ft.Row(
                            [
                                ft.Icon(ft.Icons.CALENDAR_VIEW_DAY_OUTLINED, size=14, color=ft.Colors.ON_SURFACE_VARIANT),
                                ft.Text("Ingresos por día", size=13, weight=ft.FontWeight.W_600,
                                        color=ft.Colors.ON_SURFACE),
                            ],
                            spacing=6,
                        ),
                        ft.Text(
                            "Desglose diario de todos los ingresos del periodo con el filtro seleccionado; "
                            "con más de un tipo de negocio, una tabla por tipo de negocio.",
                            size=10,
                            color=ft.Colors.ON_SURFACE_VARIANT,
                        ),
                        contenido_dia,
                    ],
                    spacing=10,
                ),
                padding=16,
                bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
                border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
                border_radius=12,
                shadow=sombra_tarjeta(),
            )

        boton_exportar.disabled = any(valor is None for valor in ultimo_datos.values())

    async def cargar(_e=None) -> None:
        cuerpo.opacity = 0.5
        progress.visible = True
        estado_text.value = ""
        boton_rango.disabled = True
        page.update()

        fecha_inicio, fecha_fin = rango_sel[0]
        resultado, significativos, por_dia = await asyncio.gather(
            asyncio.to_thread(_repo().cobranza_por_segmento, fecha_inicio, fecha_fin),
            asyncio.to_thread(_repo().ingresos_significativos, fecha_inicio, fecha_fin, segmentos_significativos[0]),
            asyncio.to_thread(_repo().ingresos_por_dia, fecha_inicio, fecha_fin, segmentos_significativos[0]),
            return_exceptions=True,
        )

        _refrescar(resultado, significativos, por_dia)

        progress.visible = False
        boton_rango.disabled = False
        cuerpo.opacity = 1.0
        if any(isinstance(valor, Exception) for valor in (resultado, significativos, por_dia)):
            estado_text.value = "No se pudo consultar BigQuery (ver detalle abajo)."
        page.update()

    def on_cambiar_rango(e) -> None:
        picker = e.control
        if not picker.start_value or not picker.end_value:
            return
        inicio = picker.start_value.date()
        fin = picker.end_value.date()
        rango_sel[0] = (inicio, fin)
        boton_rango.content.controls[1].value = _texto_rango(inicio, fin)
        page.update()
        page.run_task(cargar)

    date_range_picker = ft.DateRangePicker(
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2035, 12, 31),
        start_value=datetime.combine(hace_una_semana, datetime.min.time()),
        end_value=datetime.combine(hoy, datetime.min.time()),
        entry_mode=ft.DatePickerEntryMode.CALENDAR_ONLY,
        on_change=on_cambiar_rango,
    )

    boton_rango = ft.OutlinedButton(
        content=ft.Row(
            [ft.Icon(ft.Icons.DATE_RANGE, size=16), ft.Text(_texto_rango(hace_una_semana, hoy), size=13)],
            spacing=8,
            tight=True,
        ),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        on_click=lambda _e: page.show_dialog(date_range_picker),
    )

    # Dropdown anclado (PopupMenuButton), NO modal: un PopupMenuItem con un
    # Checkbox propio por tipo de negocio, igual que el mismo patrón en
    # app/dashboard/explorador.py. Se reconstruye en cada `_refrescar` (ver más
    # abajo) para reflejar la selección vigente en su resumen.
    def _resumen_segmentos(seleccionados: list[str]) -> str:
        if not seleccionados:
            return "Todos"
        if len(seleccionados) == 1:
            return seleccionados[0]
        return f"{len(seleccionados)} seleccionados"

    def _construir_filtro_segmento() -> ft.PopupMenuButton:
        seleccion_previa = set(segmentos_significativos[0])
        pendiente: dict[str, bool] = {segmento: (segmento in seleccion_previa) for segmento in SEGMENTOS}
        checks: list[ft.Checkbox] = []

        def _marcar_todos(_e) -> None:
            for chk in checks:
                chk.value = True
                pendiente[chk.label] = True
            page.update()

        def _limpiar(_e) -> None:
            for chk in checks:
                chk.value = False
                pendiente[chk.label] = False
            page.update()

        def _toggle(segmento: str):
            def _h(e) -> None:
                pendiente[segmento] = e.control.value
            return _h

        def _aplicar(_e) -> None:
            segmentos_significativos[0] = [segmento for segmento in SEGMENTOS if pendiente[segmento]]
            page.run_task(cargar)

        filas_valor = []
        for segmento in SEGMENTOS:
            chk = ft.Checkbox(value=pendiente[segmento], label=segmento, on_change=_toggle(segmento))
            checks.append(chk)
            filas_valor.append(ft.PopupMenuItem(content=ft.Row([chk])))

        items = [
            ft.PopupMenuItem(
                content=ft.Row([
                    ft.TextButton("Limpiar", on_click=_limpiar),
                    ft.TextButton("Todos", on_click=_marcar_todos),
                ]),
            ),
            *filas_valor,
            ft.PopupMenuItem(
                content=ft.Container(
                    content=ft.Text("Aplicar", weight=ft.FontWeight.BOLD, color=ft.Colors.PRIMARY),
                    alignment=ft.Alignment.CENTER,
                ),
                on_click=_aplicar,
            ),
        ]

        return ft.PopupMenuButton(
            content=ft.Container(
                content=ft.Row(
                    [ft.Icon(ft.Icons.CATEGORY_OUTLINED, size=16), ft.Text("Tipo de negocio", size=13),
                     ft.Text(_resumen_segmentos(segmentos_significativos[0]), size=12,
                             color=ft.Colors.ON_SURFACE_VARIANT)],
                    spacing=6, tight=True,
                ),
                padding=ft.Padding(left=12, right=12, top=6, bottom=6),
                border=ft.Border.all(1, ft.Colors.OUTLINE),
                border_radius=8,
            ),
            items=items,
        )

    def _abrir_info(_e) -> None:
        lineas = [
            "Se excluyen registros cuya Razón Social sea exactamente 'Abastecedora de Combustibles "
            "del Pacifico', 'ACP Combustibles' o 'Petro Smart Combustibles'.",
            "Los registros cuya Razón Social empiece con 'Petroplazas' (incluye variantes como "
            "PETROPLAZAS AEROPUERTO o PETROPLAZAS ESTACIONES) se cuentan como segmento 'Petroplazas', "
            "sin importar su tipo de negocio original.",
            "Solo se cuentan los tipos de negocio 'Distribuidora' y 'Asociados'; cualquier otro "
            "(por ejemplo 'GasPetroil', o vacío) se excluye — salvo que ya se haya reclasificado "
            "como Petroplazas por el punto anterior.",
            "Se excluyen las sucursales cuyo nombre contenga 'GAS', 'AUTOTANQUE', 'GC' o 'Corporativo'.",
            "Los registros en dólares se muestran separados (columna 'USD') y se convierten a MXN "
            "(columna 'MXN convertido') con el promedio diario de im_TipoCambio en "
            "DocumentosClientesCobranza. Si no existe tipo de cambio en la fecha exacta, se usa la "
            "fecha disponible más cercana. La columna 'USD sin TC' (en rojo cuando tiene saldo) es la "
            "parte del USD para la que no existe NINGÚN tipo de cambio en toda la historia de Cobranza "
            "— no se descarta ni se convierte con un valor inventado, y queda fuera de 'Total final' "
            "(= MXN + MXN convertido) porque no hay con qué convertirla.",
            "Ingresos Significativos muestra los 20 mayores ingresos agregados por Razón Social, "
            "ordenados por Total final = MXN + USD convertido. Puede filtrarse por uno o varios "
            "tipos de negocio (Distribuidora, Asociados, Petroplazas) a la vez; sin ninguno marcado "
            "se muestran los tres.",
            "Dentro de Ingresos Significativos, el apartado Ingresos por día agrega los movimientos de "
            "cada fecha del periodo por tipo de negocio (una fila por día y tipo de negocio, no un total "
            "combinado por día), respeta el mismo filtro de tipo de negocio y el mismo tratamiento de "
            "MXN, USD, conversión y USD sin tipo de cambio.",
            "La fecha usada para filtrar es fh_Envio; por defecto se muestra la semana anterior a hoy "
            "(espejo de la semana a futuro que muestra el panel de Proyección, a la izquierda).",
        ]
        dialogo = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cómo se calculan estos datos"),
            content=ft.Container(
                content=ft.Column(
                    [ft.Text(f"•  {l}", size=12, selectable=True) for l in lineas],
                    spacing=10, scroll=ft.ScrollMode.AUTO,
                ),
                width=480, height=340,
            ),
            actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
        )
        mostrar_dialogo(page, dialogo)

    boton_info = ft.IconButton(
        icon=ft.Icons.INFO_OUTLINE,
        icon_size=18,
        tooltip="Ver cómo se calculan estos datos",
        on_click=_abrir_info,
    )

    async def exportar_excel(_e) -> None:
        """Descarga un Excel con las 3 vistas del panel (mismos datos que ya
        están en pantalla, sin volver a consultar BigQuery): concentrado por
        segmento, Top 20 de Ingresos Significativos e Ingresos por día."""
        items = ultimo_datos["items"]
        significativos = ultimo_datos["significativos"]
        por_dia = ultimo_datos["por_dia"]
        if items is None or significativos is None or por_dia is None:
            return
        boton_exportar.disabled = True
        estado_text.value = "Generando Excel…"
        page.update()

        import openpyxl

        wb = openpyxl.Workbook()
        nombres_usados: set = set()

        ws_concentrado = wb.active
        ws_concentrado.title = nombre_hoja_valido("Concentrado", nombres_usados)
        total_mxn = sum(mxn for _s, mxn, _u, _c, _st in items)
        total_usd = sum(usd for _s, _m, usd, _c, _st in items)
        total_convertido = sum(c for _s, _m, _u, c, _st in items)
        total_sin_tc = sum(st for _s, _m, _u, _c, st in items)
        filas_concentrado = [
            [segmento, round(mxn, 2), round(usd, 2), round(convertido, 2),
             round(mxn + convertido, 2), round(sin_tc, 2)]
            for segmento, mxn, usd, convertido, sin_tc in items
        ]
        filas_concentrado.append([
            "Total", round(total_mxn, 2), round(total_usd, 2), round(total_convertido, 2),
            round(total_mxn + total_convertido, 2), round(total_sin_tc, 2),
        ])
        escribir_hoja_excel(
            ws_concentrado,
            ["Segmento", "MXN", "USD", "MXN convertido", "Total final", "USD sin TC"],
            filas_concentrado,
        )
        for fila_celdas in ws_concentrado.iter_rows(min_row=2, min_col=2, max_col=6):
            for celda in fila_celdas:
                celda.number_format = "#,##0.00"

        ws_top = wb.create_sheet(nombre_hoja_valido("Ingresos Significativos", nombres_usados))
        escribir_hoja_excel(
            ws_top,
            ["#", "Razón social", "Tipo de negocio", "MXN", "USD", "MXN convertido", "Total final", "USD sin TC"],
            [
                [
                    puesto,
                    fila["razon_social"],
                    fila.get("segmento") or "",
                    round(fila.get("total_mxn") or 0, 2),
                    round(fila.get("total_usd") or 0, 2),
                    round(fila.get("total_usd_convertido") or 0, 2),
                    round(fila.get("total_final") or 0, 2),
                    round(fila.get("total_usd_sin_tc") or 0, 2),
                ]
                for puesto, fila in enumerate(significativos, start=1)
            ],
        )
        for fila_celdas in ws_top.iter_rows(min_row=2, min_col=4, max_col=8):
            for celda in fila_celdas:
                celda.number_format = "#,##0.00"

        # Igual que en pantalla: con más de un tipo de negocio presente, una
        # tabla por tipo de negocio (apiladas en la misma hoja, con su nombre
        # como título) en vez de una sola tabla con columna "Tipo de negocio".
        def _escribir_bloque_dia(ws, fila_inicio: int, titulo: str | None, filas: list[dict]) -> int:
            """Escribe encabezado + filas de un bloque de 'Ingresos por día' (con
            un título en negritas arriba si `titulo` no es None) y devuelve la
            fila donde debe empezar el siguiente bloque."""
            from openpyxl.styles import Font

            if titulo is not None:
                ws.cell(row=fila_inicio, column=1, value=titulo).font = Font(bold=True, size=12)
                fila_encabezado = fila_inicio + 1
            else:
                fila_encabezado = fila_inicio
            escribir_hoja_excel(
                ws,
                ["Fecha", "MXN", "USD", "MXN convertido", "Total final", "USD sin TC"],
                [
                    [
                        fila["fecha"],
                        round(fila.get("total_mxn") or 0, 2),
                        round(fila.get("total_usd") or 0, 2),
                        round(fila.get("total_usd_convertido") or 0, 2),
                        round(fila.get("total_final") or 0, 2),
                        round(fila.get("total_usd_sin_tc") or 0, 2),
                    ]
                    for fila in filas
                ],
                fila_inicio=fila_encabezado,
            )
            primera_fila_datos = fila_encabezado + 1
            ultima_fila_datos = fila_encabezado + len(filas)
            for fila_celdas in ws.iter_rows(min_row=primera_fila_datos, max_row=ultima_fila_datos,
                                            min_col=1, max_col=1):
                for celda in fila_celdas:
                    celda.number_format = "dd/mm/yyyy"
            for fila_celdas in ws.iter_rows(min_row=primera_fila_datos, max_row=ultima_fila_datos,
                                            min_col=2, max_col=6):
                for celda in fila_celdas:
                    celda.number_format = "#,##0.00"
            return ultima_fila_datos + 2  # +1 fila en blanco de separación antes del siguiente bloque

        ws_dia = wb.create_sheet(nombre_hoja_valido("Ingresos por dia", nombres_usados))
        segmentos_en_dia = [s for s in SEGMENTOS if any(f.get("segmento") == s for f in por_dia)]
        if len(segmentos_en_dia) > 1:
            fila_actual = 1
            for segmento in segmentos_en_dia:
                filas_segmento = [f for f in por_dia if f.get("segmento") == segmento]
                fila_actual = _escribir_bloque_dia(ws_dia, fila_actual, segmento, filas_segmento)
        else:
            _escribir_bloque_dia(ws_dia, 1, None, por_dia)

        fecha_inicio, fecha_fin = rango_sel[0]
        nombre_def = f"rdc_cobranza_{fecha_inicio:%Y%m%d}_{fecha_fin:%Y%m%d}.xlsx"
        ok, mensaje = await guardar_workbook(page, file_picker, wb, nombre_def)
        boton_exportar.disabled = False
        estado_text.value = mensaje
        page.update()

    boton_exportar = ft.IconButton(
        icon=ft.Icons.DOWNLOAD,
        icon_size=18,
        tooltip="Descargar Excel (concentrado + ingresos significativos + por día)",
        disabled=True,
        on_click=lambda e: page.run_task(exportar_excel, e),
    )

    barra_herramientas = ft.Container(
        content=ft.Row(
            [boton_rango, progress, estado_text, ft.Container(expand=True), boton_exportar, boton_info],
            spacing=12,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        padding=ft.Padding(left=14, right=14, top=10, bottom=10),
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOW,
        border_radius=12,
    )

    contenido = ft.Column(
        [
            ft.Column([titulo, subtitulo], spacing=2),
            barra_herramientas,
            ft.Container(content=cuerpo, expand=True),
        ],
        spacing=16,
        scroll=ft.ScrollMode.AUTO,
        expand=True,
    )

    page.run_task(cargar)

    return contenido
