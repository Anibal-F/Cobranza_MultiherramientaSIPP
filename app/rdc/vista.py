"""Pestaña 'RDC': antigüedad de saldos vencidos, segmentada en Distribuidora,
Asociados y Petroplazas — réplica de las macros `CargarAntiguedadSaldos` /
`CargarAntiguedadAsociados` del Excel de Proyección, sobre la tabla
`documentosClientes_AntiguedadSaldosVencidoPorClienteDetalle` de BigQuery."""

import math
import os
from datetime import date, datetime, timedelta

import flet as ft
import flet_charts as fc

from ..dashboard.componentes import (
    color_slot,
    encabezado_seccion,
    formato_compacto,
    mostrar_dialogo,
    preparar_tema_date_picker,
    sombra_tarjeta,
    tile_compacta,
)
from .cobranza import construir_panel_cobranza
from .consultas import SEGMENTOS, consultar_antiguedad_saldos, consultar_detalle_periodo

# Columnas de fecha en el detalle crudo (SELECT * de la tabla): se formatean
# como fecha en el Excel en vez de dejarlas como el datetime completo de BigQuery.
_COLUMNAS_FECHA_DETALLE = {"fh_Documento", "fh_Venta", "fh_Vencimiento"}

_COLOR_SLOT_VIGENTE = 0
_COLOR_SLOT_VENCIDO = 5


def _redondear_max_y(valor: float) -> float:
    """Sube `valor` al siguiente escalón "limpio" para el tope del eje Y (ver
    misma lógica en el dashboard de ingresos: evita un *1.15 crudo que deja el
    tope pegado con el último tick "redondo")."""
    if valor <= 0:
        return 10
    magnitud = 10 ** math.floor(math.log10(valor))
    paso = magnitud / 2
    return math.ceil(valor * 1.05 / paso) * paso


def _leyenda_metricas(dark: bool) -> ft.Row:
    def _item(color: str, etiqueta: str) -> ft.Row:
        return ft.Row(
            [
                ft.Container(width=10, height=10, bgcolor=color, border_radius=5),
                ft.Text(etiqueta, size=11, color=ft.Colors.ON_SURFACE),
            ],
            spacing=6,
        )

    return ft.Row(
        [
            _item(color_slot(_COLOR_SLOT_VIGENTE, dark), "Saldo vigente"),
            _item(color_slot(_COLOR_SLOT_VENCIDO, dark), "Vencido a 30 días"),
        ],
        spacing=20,
    )


def _construir_barra_segmentos(items: list[tuple[str, float, float]], dark: bool) -> ft.Control:
    """Barra apilada por segmento: vigente (abajo) + vencido a 30 días
    (arriba). Con solo 3 categorías y 2 métricas, una barra apilada comunica
    mejor la composición que una dona (parte-todo de UNA métrica) o un
    leaderboard (UNA métrica, muchas categorías) — ninguno de los dos encaja."""
    color_vigente = color_slot(_COLOR_SLOT_VIGENTE, dark)
    color_vencido = color_slot(_COLOR_SLOT_VENCIDO, dark)
    max_total = max((vigente + vencido30 for _s, vigente, vencido30 in items), default=0)

    grupos = []
    for i, (_segmento, vigente, vencido30) in enumerate(items):
        total = vigente + vencido30
        stack_items = [
            fc.BarChartRodStackItem(from_y=0, to_y=vigente, color=color_vigente),
            fc.BarChartRodStackItem(from_y=vigente, to_y=total, color=color_vencido),
        ]
        grupos.append(
            fc.BarChartGroup(
                x=i,
                rods=[
                    fc.BarChartRod(
                        from_y=0,
                        to_y=total if total > 0 else 0.0001,
                        width=48,
                        color=color_vigente,
                        stack_items=stack_items,
                        border_radius=ft.BorderRadius(top_left=6, top_right=6, bottom_left=0, bottom_right=0),
                        tooltip=f"Vigente {formato_compacto(vigente)} · Vencido 30d {formato_compacto(vencido30)}",
                    )
                ],
            )
        )

    grafica = fc.BarChart(
        groups=grupos,
        max_y=_redondear_max_y(max_total),
        group_spacing=40,
        bottom_axis=fc.ChartAxis(
            labels=[
                fc.ChartAxisLabel(value=float(i), label=ft.Text(segmento, size=11, color=ft.Colors.ON_SURFACE))
                for i, (segmento, _v, _v3) in enumerate(items)
            ],
            show_labels=True,
        ),
        left_axis=fc.ChartAxis(show_labels=False),
        expand=True,
    )
    return ft.Container(grafica, height=220)


def _construir_tabla_segmentos(items: list[tuple[str, float, float]]) -> ft.Control:
    filas = []
    for segmento, vigente, vencido30 in items:
        filas.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(segmento, size=12)),
                    ft.DataCell(ft.Text(f"${vigente:,.2f}", size=12)),
                    ft.DataCell(ft.Text(f"${vencido30:,.2f}", size=12)),
                    ft.DataCell(ft.Text(f"${vigente + vencido30:,.2f}", size=12, weight=ft.FontWeight.W_600)),
                ]
            )
        )
    return ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("Segmento", size=12)),
            ft.DataColumn(ft.Text("Vigente", size=12), numeric=True),
            ft.DataColumn(ft.Text("Vencido 30 días", size=12), numeric=True),
            ft.DataColumn(ft.Text("Total", size=12), numeric=True),
        ],
        rows=filas,
        data_row_max_height=36,
        heading_row_height=36,
        column_spacing=24,
    )


def _construir_workbook_reporte(items: list[tuple[str, float, float]], detalle: list[dict]):
    """Arma el Excel de descarga con 2 hojas:

    - 'Concentrado': el mismo resumen por segmento que muestra el dashboard
      (vigente, vencido a 30 días y total), con una fila de gran total.
    - 'Detalle periodo': todos los registros de la tabla cuya fh_Vencimiento
      cae en el rango seleccionado, SIN los filtros de negocio del concentrado
      (cliente/factura vacíos, ICV, Totales, prefijo FCOR, segmentación) — solo
      el filtro de fecha, para poder auditar el concentrado fila por fila."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    def _estilizar_encabezado(ws) -> None:
        for celda in ws[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="1B3A5B")

    wb = openpyxl.Workbook()

    ws_concentrado = wb.active
    ws_concentrado.title = "Concentrado"
    ws_concentrado.append(["Segmento", "Saldo vigente", "Vencido a 30 días", "Total"])
    _estilizar_encabezado(ws_concentrado)
    for segmento, vigente, vencido30 in items:
        ws_concentrado.append([segmento, round(vigente, 2), round(vencido30, 2), round(vigente + vencido30, 2)])
    total_vigente = sum(v for _s, v, _v3 in items)
    total_vencido30 = sum(v3 for _s, _v, v3 in items)
    ws_concentrado.append(["Total", round(total_vigente, 2), round(total_vencido30, 2),
                           round(total_vigente + total_vencido30, 2)])
    for celda in ws_concentrado[ws_concentrado.max_row]:
        celda.font = Font(bold=True)
    for fila_celdas in ws_concentrado.iter_rows(min_row=2, min_col=2, max_col=4):
        for celda in fila_celdas:
            celda.number_format = "#,##0.00"
    for i, w in enumerate([18, 18, 20, 18], start=1):
        ws_concentrado.column_dimensions[get_column_letter(i)].width = w
    ws_concentrado.freeze_panes = "A2"

    ws_detalle = wb.create_sheet("Detalle periodo")
    encabezados = list(detalle[0].keys()) if detalle else []
    ws_detalle.append(encabezados)
    _estilizar_encabezado(ws_detalle)
    for fila in detalle:
        ws_detalle.append([fila.get(col) for col in encabezados])
    for i, col in enumerate(encabezados, start=1):
        letra = get_column_letter(i)
        ws_detalle.column_dimensions[letra].width = max(12, min(30, len(col) + 2))
        if col in _COLUMNAS_FECHA_DETALLE:
            for fila_celdas in ws_detalle.iter_rows(min_row=2, min_col=i, max_col=i):
                for celda in fila_celdas:
                    celda.number_format = "dd/mm/yyyy"
    ws_detalle.freeze_panes = "A2"

    return wb


def construir_tab_rdc(page: ft.Page) -> tuple[ft.Tab, ft.Control]:
    """Pestaña 'RDC'. Devuelve (tab, contenido): el `Tab` va en `TabBar.tabs` y
    `contenido` en `TabBarView.controls`, en la misma posición (así arma las
    pestañas main.py)."""
    preparar_tema_date_picker(page)

    hoy = date.today()
    una_semana_adelante = hoy + timedelta(days=7)
    rango_sel: list[tuple[date, date]] = [(hoy, una_semana_adelante)]

    def _dark() -> bool:
        return page.theme_mode == ft.ThemeMode.DARK

    def _texto_rango(inicio: date, fin: date) -> str:
        return f"{inicio.strftime('%d %b %Y')} – {fin.strftime('%d %b %Y')}"

    titulo = ft.Text("Proyección", size=20, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE)
    subtitulo = ft.Text(
        "Distribuidora, Asociados y Petroplazas · el saldo vigente se filtra por fecha de vencimiento "
        "dentro del rango; el vencido a 30 días es el acumulado total a la fecha de corte.",
        size=12,
        color=ft.Colors.ON_SURFACE_VARIANT,
    )

    estado_text = ft.Text("", size=12, color=ft.Colors.RED_600)
    progress = ft.ProgressRing(width=16, height=16, visible=False, stroke_width=2)

    hero_contenedor = ft.ResponsiveRow(spacing=16, run_spacing=16)
    seccion_grafica = ft.Container()
    cuerpo = ft.Column([hero_contenedor, seccion_grafica], spacing=20, opacity=1.0, animate_opacity=200)

    # En Flet 0.85 el FilePicker es un servicio: se crea y se usa directamente
    # (NO se agrega a page.overlay; hacerlo provoca "Unknown control: FilePicker").
    file_picker = ft.FilePicker()

    ultimo_items: list = [[]]  # concentrado de la última consulta exitosa, para exportar sin recalcular

    def _refrescar(resultado) -> None:
        dark = _dark()
        if isinstance(resultado, Exception):
            hero_contenedor.controls = []
            seccion_grafica.content = ft.Container(
                content=ft.Text(f"No se pudo consultar: {resultado}", size=12, color=ft.Colors.RED_600),
                height=160,
                alignment=ft.Alignment.CENTER,
            )
            boton_exportar.disabled = True
            return

        por_segmento = {fila["segmento"]: fila for fila in resultado}
        items = [
            (
                segmento,
                (por_segmento.get(segmento, {}).get("saldo_vigente") or 0),
                (por_segmento.get(segmento, {}).get("saldo_vencido_30") or 0),
            )
            for segmento in SEGMENTOS
        ]
        ultimo_items[0] = items
        boton_exportar.disabled = False

        total_vigente = sum(v for _s, v, _v3 in items)
        total_vencido30 = sum(v3 for _s, _v, v3 in items)

        # col explícito a 3 por fila: este panel vive a media pantalla (junto al
        # de Cobranza), así que el default de tile_compacta (2 por fila) dejaría
        # una tarjeta sola y angosta en la 2a fila — con exactamente 3 tarjetas,
        # un tercio cada una aprovecha todo el ancho disponible sin huecos.
        col_tercio = {"xs": 12, "sm": 4}
        hero_contenedor.controls = [
            tile_compacta("Total cartera", total_vigente + total_vencido30, color_slot(2, dark),
                          ft.Icons.ACCOUNT_BALANCE_OUTLINED, "Distribuidora + Asociados + Petroplazas",
                          col=col_tercio),
            tile_compacta("Saldo vigente", total_vigente, color_slot(_COLOR_SLOT_VIGENTE, dark),
                          ft.Icons.SCHEDULE_OUTLINED, "Facturas con vencimiento en el rango seleccionado",
                          col=col_tercio),
            tile_compacta("Vencido a 30 días", total_vencido30, color_slot(_COLOR_SLOT_VENCIDO, dark),
                          ft.Icons.WARNING_AMBER_OUTLINED, "Acumulado total a la fecha de corte",
                          col=col_tercio),
        ]

        seccion_grafica.content = ft.Container(
            content=ft.Column(
                [
                    encabezado_seccion(
                        ft.Icons.BAR_CHART_OUTLINED, color_slot(2, dark),
                        "Vigente vs. vencido a 30 días", "Por segmento: Distribuidora, Asociados y Petroplazas",
                        [_leyenda_metricas(dark)],
                    ),
                    ft.Divider(height=1),
                    _construir_barra_segmentos(items, dark),
                    ft.Container(
                        content=_construir_tabla_segmentos(items),
                        padding=ft.Padding(top=8, right=0, left=0, bottom=0),
                    ),
                ],
                spacing=10,
            ),
            padding=16,
            bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
            border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
            border_radius=12,
            shadow=sombra_tarjeta(),
        )

    async def cargar(_e=None) -> None:
        cuerpo.opacity = 0.5
        progress.visible = True
        estado_text.value = ""
        boton_rango.disabled = True
        page.update()

        fecha_inicio, fecha_fin = rango_sel[0]
        try:
            resultado = await consultar_antiguedad_saldos(fecha_inicio, fecha_fin)
        except Exception as error:  # noqa: BLE001 - se muestra en la sección, igual que el dashboard de ingresos
            resultado = error

        _refrescar(resultado)

        progress.visible = False
        boton_rango.disabled = False
        cuerpo.opacity = 1.0
        if isinstance(resultado, Exception):
            estado_text.value = "No se pudo consultar BigQuery (ver detalle abajo)."
        page.update()

    def on_cambiar_rango(e) -> None:
        picker = e.control
        if not picker.start_value or not picker.end_value:
            return
        inicio = picker.start_value.date()
        fin = picker.end_value.date()
        rango_sel[0] = (inicio, fin)
        boton_rango.content.controls[1].value = _texto_rango(inicio, fin)
        page.update()
        page.run_task(cargar)

    date_range_picker = ft.DateRangePicker(
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2035, 12, 31),
        start_value=datetime.combine(hoy, datetime.min.time()),
        end_value=datetime.combine(una_semana_adelante, datetime.min.time()),
        entry_mode=ft.DatePickerEntryMode.CALENDAR_ONLY,
        on_change=on_cambiar_rango,
    )

    boton_rango = ft.OutlinedButton(
        content=ft.Row(
            [ft.Icon(ft.Icons.DATE_RANGE, size=16), ft.Text(_texto_rango(hoy, una_semana_adelante), size=13)],
            spacing=8,
            tight=True,
        ),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        on_click=lambda _e: page.show_dialog(date_range_picker),
    )

    def _abrir_info(_e) -> None:
        lineas = [
            "Distribuidora y Asociados: se agrupan por nb_TipoDeNegocio, excluyendo las filas "
            "cuyo cliente sea PETROPLAZAS (esas se cuentan aparte).",
            "Petroplazas: se identifica por nombre de cliente ('PETROPLAZAS'), sin importar su "
            "nb_TipoDeNegocio — igual que en la macro de Excel, donde este cliente aparecía tanto "
            "en el reporte de Distribuidora como en el de Asociados.",
            "El segmento GasPetroil (y filas sin tipo de negocio) no entra en ninguna categoría — "
            "la macro original nunca los procesaba.",
            "Se excluyen filas sin cliente o sin factura, el cliente 'ICV' y los folios que "
            "empiezan con 'FCOR' (el filtro configurado en Config_Filtros > Antigüedad de Saldos).",
            "El saldo vigente (im_CarteraVigente) solo se suma si la fecha de vencimiento de la "
            "factura cae dentro del rango de fechas seleccionado.",
            "El vencido a 30 días (im_Vencido30Dias) se suma completo, sin filtrar por fecha — es "
            "el acumulado total a la fecha de corte, tal como hacía la macro original.",
        ]
        dialogo = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cómo se calculan estos datos"),
            content=ft.Container(
                content=ft.Column(
                    [ft.Text(f"•  {l}", size=12, selectable=True) for l in lineas],
                    spacing=10, scroll=ft.ScrollMode.AUTO,
                ),
                width=520, height=340,
            ),
            actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
        )
        mostrar_dialogo(page, dialogo)

    boton_info = ft.IconButton(
        icon=ft.Icons.INFO_OUTLINE,
        icon_size=18,
        tooltip="Ver cómo se calculan estos datos",
        on_click=_abrir_info,
    )

    async def exportar_excel(_e) -> None:
        """Descarga un Excel con 2 hojas: el concentrado que muestra el
        dashboard, y el detalle crudo del periodo (solo filtro de fecha, sin
        los filtros de negocio del concentrado) para poder auditarlo."""
        if not ultimo_items[0]:
            return
        boton_exportar.disabled = True
        estado_text.value = "Consultando detalle del periodo…"
        page.update()

        fecha_inicio, fecha_fin = rango_sel[0]
        try:
            detalle = await consultar_detalle_periodo(fecha_inicio, fecha_fin)
            wb = _construir_workbook_reporte(ultimo_items[0], detalle)
        except Exception as error:  # noqa: BLE001 - se muestra en estado_text, igual que el resto de la pestaña
            estado_text.value = f"No se pudo generar el Excel: {error}"
            boton_exportar.disabled = False
            page.update()
            return

        nombre_def = f"rdc_antiguedad_saldos_{fecha_inicio:%Y%m%d}_{fecha_fin:%Y%m%d}.xlsx"
        carpeta_descargas = os.path.expanduser("~/Downloads")
        destino = await file_picker.save_file(
            dialog_title="Guardar antigüedad de saldos en Excel",
            file_name=nombre_def,
            initial_directory=carpeta_descargas if os.path.isdir(carpeta_descargas) else None,
            allowed_extensions=["xlsx"],
        )
        boton_exportar.disabled = False
        if not destino:
            estado_text.value = ""
            page.update()
            return
        if not destino.lower().endswith(".xlsx"):
            destino += ".xlsx"
        try:
            wb.save(destino)
        except OSError as error:
            # En algunos equipos macOS el diálogo nativo regresa una ruta no
            # escribible (ej. la raíz del sistema) aunque el usuario haya
            # navegado a una carpeta válida — se reintenta en Descargas antes
            # de rendirse, para no perder el reporte ya generado.
            respaldo = os.path.join(carpeta_descargas, os.path.basename(destino))
            try:
                os.makedirs(carpeta_descargas, exist_ok=True)
                wb.save(respaldo)
            except OSError:
                estado_text.value = f"No se pudo guardar el Excel: {error}"
                page.update()
                return
            destino = respaldo
            estado_text.value = (
                f"No se pudo guardar en la ubicación elegida; se guardó en Descargas: "
                f"{os.path.basename(destino)}."
            )
            page.update()
            return
        estado_text.value = f"Exportado: {len(detalle)} registro(s) de detalle → {os.path.basename(destino)}."
        page.update()

    boton_exportar = ft.IconButton(
        icon=ft.Icons.DOWNLOAD,
        icon_size=18,
        tooltip="Descargar Excel (concentrado + detalle del periodo)",
        disabled=True,
        on_click=lambda e: page.run_task(exportar_excel, e),
    )

    barra_herramientas = ft.Container(
        content=ft.Row(
            [boton_rango, progress, estado_text, ft.Container(expand=True), boton_exportar, boton_info],
            spacing=12,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        padding=ft.Padding(left=14, right=14, top=10, bottom=10),
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOW,
        border_radius=12,
    )

    panel_izquierdo = ft.Container(
        content=ft.Column(
            [
                ft.Column([titulo, subtitulo], spacing=2),
                barra_herramientas,
                ft.Container(content=cuerpo, expand=True),
            ],
            spacing=16,
            scroll=ft.ScrollMode.AUTO,
            expand=True,
        ),
        padding=ft.Padding(left=20, right=16, top=20, bottom=20),
        border=ft.Border(right=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT)),
        expand=True,
        col={"xs": 12, "lg": 6},
    )

    # Mitad derecha: lo efectivamente cobrado la semana anterior (espejo de la
    # proyección a futuro de la izquierda) — ver app/rdc/cobranza.py.
    panel_derecho = ft.Container(
        content=construir_panel_cobranza(page),
        padding=ft.Padding(left=16, right=20, top=20, bottom=20),
        expand=True,
        col={"xs": 12, "lg": 6},
    )

    contenido = ft.ResponsiveRow(
        [panel_izquierdo, panel_derecho],
        spacing=0,
        run_spacing=0,
        vertical_alignment=ft.CrossAxisAlignment.START,
        expand=True,
    )

    page.run_task(cargar)

    tab = ft.Tab(label="Proyección", icon=ft.Icons.HISTORY_EDU)
    return tab, contenido
