"""Piezas visuales compartidas del dashboard: paleta categórica, formato de
montos, tarjetas KPI, dona, ranking, tabla simple, barras del timeline y
helpers de tema/diálogo. Sin ninguna dependencia de BigQuery."""

import dataclasses
import os
import re
from datetime import date

import flet as ft
import flet_charts as fc

# Paleta categórica validada (orden fijo, CVD-safe): slot 0 = azul, 1 = aqua,
# 2 = amarillo, ... Un juego de tonos para tema claro y otro para oscuro.
# Nunca se asignan más de ~6-8 slots a la vez.
PALETA_CATEGORICA_LIGHT = ["#2a78d6", "#1baf7a", "#eda100", "#008300", "#4a3aa7", "#e34948", "#e87ba4", "#eb6834"]
PALETA_CATEGORICA_DARK = ["#3987e5", "#199e70", "#c98500", "#008300", "#9085e9", "#e66767", "#d55181", "#d95926"]


def color_slot(indice: int, dark: bool) -> str:
    paleta = PALETA_CATEGORICA_DARK if dark else PALETA_CATEGORICA_LIGHT
    return paleta[indice % len(paleta)]


def formato_compacto(valor: float) -> str:
    """$580,497,376.78 -> $580.5M (cifra escaneable en una tarjeta KPI)."""
    signo = "-" if valor < 0 else ""
    valor = abs(valor)
    if valor >= 1_000_000_000:
        return f"{signo}${valor / 1_000_000_000:,.1f}B"
    if valor >= 1_000_000:
        return f"{signo}${valor / 1_000_000:,.1f}M"
    if valor >= 1_000:
        return f"{signo}${valor / 1_000:,.1f}K"
    return f"{signo}${valor:,.2f}"


def preparar_tema_date_picker(page: ft.Page) -> None:
    """El diálogo compacto (entry_mode=INPUT) del DateRangePicker es angosto y
    el texto grande del rango ("Jul 1 – Jul 7") se corta a 2 líneas con el
    tamaño de fuente por defecto. No hay una prop por instancia para esto en
    DateRangePicker — solo a nivel tema, así que se ajusta aquí (sin tocar
    main.py) para claro y oscuro."""
    tema = ft.DatePickerTheme(
        range_picker_header_headline_text_style=ft.TextStyle(size=18, weight=ft.FontWeight.W_600)
    )
    if page.theme is not None:
        page.theme = dataclasses.replace(page.theme, date_picker_theme=tema)
    if page.dark_theme is not None:
        page.dark_theme = dataclasses.replace(page.dark_theme, date_picker_theme=tema)


def mostrar_dialogo(page: ft.Page, dialogo: ft.AlertDialog) -> None:
    """Réplica local del helper `mostrar_dialogo` de app/main.py (no es
    importable: vive local a `main(page)`). Evita 'Dialog is already opened'
    si un diálogo anterior quedó colgado en la pila interna de Flet."""
    pila = page._dialogs.controls
    if dialogo in pila:
        if dialogo.open:
            return
        pila.remove(dialogo)
    page.show_dialog(dialogo)


# --- Exportación a Excel (Segmentado / Timeline / Detalle) --------------------
# Helpers compartidos por las 3 descargas del dashboard: escribir una hoja con
# encabezado estilizado, y el flujo completo de guardado (diálogo nativo +
# reintento si la ruta elegida resulta no escribible).

def nombre_hoja_valido(base: str, usados: set) -> str:
    """Nombre de hoja de Excel válido (≤31 chars, sin \\ / * ? : [ ]) y único
    dentro de `usados` (que se muta con el nombre devuelto)."""
    limpio = re.sub(r"[\\/*?:\[\]]", "", base or "").strip()[:28] or "Hoja"
    nombre, i = limpio, 2
    while nombre in usados:
        nombre = f"{limpio[:25]}_{i}"
        i += 1
    usados.add(nombre)
    return nombre


def escribir_hoja_excel(ws, encabezados: list[str], filas: list[list], fila_inicio: int = 1) -> None:
    """Escribe encabezado (estilizado en azul marino) + filas de datos en `ws`,
    empezando en `fila_inicio` (por defecto la primera fila de la hoja) — usar
    un valor > 1 cuando antes se escribió alguna fila de contexto/metadatos
    (ej. los filtros aplicados en la exportación de Timeline). Columnas
    anchadas por longitud del encabezado; freeze_panes justo debajo del
    encabezado."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    for col, valor in enumerate(encabezados, start=1):
        ws.cell(row=fila_inicio, column=col, value=valor)
    for celda in ws[fila_inicio]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1B3A5B")
    for offset, fila in enumerate(filas, start=1):
        for col, valor in enumerate(fila, start=1):
            ws.cell(row=fila_inicio + offset, column=col, value=valor)
    for col, encabezado in enumerate(encabezados, start=1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(30, len(str(encabezado)) + 2))
    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1).coordinate


async def guardar_workbook(page: ft.Page, file_picker: ft.FilePicker, wb, nombre_sugerido: str) -> tuple[bool, str]:
    """Abre el diálogo nativo de guardado y escribe `wb` en la ruta elegida.
    Devuelve (ok, mensaje) para que el llamador lo muestre en su propio texto
    de estado. En algunos equipos macOS el diálogo regresa una ruta no
    escribible (ej. la raíz del sistema, de solo lectura) aunque el usuario
    haya navegado a una carpeta válida — se reintenta en ~/Downloads antes de
    rendirse, para no perder el reporte ya generado."""
    destino = await file_picker.save_file(
        dialog_title="Guardar Excel",
        file_name=nombre_sugerido,
        allowed_extensions=["xlsx"],
    )
    if not destino:
        return False, ""
    if not destino.lower().endswith(".xlsx"):
        destino += ".xlsx"
    try:
        wb.save(destino)
        return True, f"Exportado: {os.path.basename(destino)}."
    except OSError as error:
        carpeta_respaldo = os.path.expanduser("~/Downloads")
        respaldo = os.path.join(carpeta_respaldo, os.path.basename(destino))
        try:
            os.makedirs(carpeta_respaldo, exist_ok=True)
            wb.save(respaldo)
            return True, f"No se pudo guardar en la ubicación elegida; se guardó en Descargas: {os.path.basename(respaldo)}."
        except OSError:
            return False, f"No se pudo guardar el Excel: {error}"


def chip_total(total: float) -> ft.Container:
    """Pastilla discreta 'Total · $X' para la cabecera de cada sección."""
    return ft.Container(
        content=ft.Row(
            [
                ft.Text("Total", size=10, color=ft.Colors.ON_SURFACE_VARIANT),
                ft.Text(formato_compacto(total), size=12, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE),
            ],
            spacing=6,
            tight=True,
        ),
        padding=ft.Padding(left=10, right=10, top=4, bottom=4),
        bgcolor=ft.Colors.SURFACE_CONTAINER_HIGH,
        border_radius=20,
    )


def chip_total_usd(total_usd: float) -> ft.Control:
    """Pastilla 'USD · $X' — el monto en dólares de una sección, mostrado
    SIEMPRE aparte de `chip_total` (nunca sumado a los pesos: son montos en
    monedas distintas). No renderiza nada si la sección no tuvo movimientos
    en dólares — la mayoría no los tiene, y una pastilla en $0.00 sería ruido."""
    if not total_usd:
        return ft.Container()
    return ft.Container(
        content=ft.Row(
            [
                ft.Text("USD", size=10, color=ft.Colors.ON_SURFACE_VARIANT),
                ft.Text(f"${total_usd:,.2f}", size=12, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE),
            ],
            spacing=6,
            tight=True,
        ),
        padding=ft.Padding(left=10, right=10, top=4, bottom=4),
        bgcolor=ft.Colors.with_opacity(0.14, "#eda100"),
        border_radius=20,
    )


def sombra_tarjeta() -> ft.BoxShadow:
    """Sombra sutil compartida por TODAS las tarjetas del dashboard (KPI,
    secciones, paneles de filtro) — antes cada tarjeta era un rectángulo
    plano de 1px (bgcolor + border), correcto pero sin profundidad ni
    jerarquía visual frente al fondo. Una sombra chica (blur 16, sin spread)
    las separa del fondo sin verse "flotadas"; se centraliza aquí para poder
    ajustar el look de toda la app en un solo lugar."""
    return ft.BoxShadow(
        spread_radius=0,
        blur_radius=16,
        color=ft.Colors.with_opacity(0.10, ft.Colors.SHADOW),
        offset=ft.Offset(0, 3),
    )


def encabezado_seccion(icono, color: str, titulo: str, subtitulo: str, chips: list[ft.Control] | None = None) -> ft.Row:
    """Cabecera estándar de tarjeta de sección: ícono en círculo de acento +
    título + subtítulo a la izquierda, chips (Total/USD) a la derecha —
    mismo lenguaje visual que `hero_tile`/`tile_compacta`, para que las
    tarjetas de sección (donas, rankings, tablas) se sientan parte del mismo
    sistema que las tarjetas KPI en vez de solo texto plano sin acento."""
    return ft.Row(
        [
            ft.Container(
                ft.Icon(icono, color=color, size=16),
                width=30, height=30, border_radius=9,
                bgcolor=ft.Colors.with_opacity(0.14, color),
                alignment=ft.Alignment.CENTER,
            ),
            ft.Column(
                [
                    ft.Text(titulo, size=14, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE),
                    ft.Text(subtitulo, size=10, color=ft.Colors.ON_SURFACE_VARIANT,
                            max_lines=2, overflow=ft.TextOverflow.ELLIPSIS),
                ],
                spacing=2,
                expand=True,
            ),
            ft.Row(chips or [], spacing=6, tight=True),
        ],
        spacing=10,
        vertical_alignment=ft.CrossAxisAlignment.START,
    )


def hero_tile(etiqueta: str, valor, color: str, icono, subtexto: str = "") -> ft.Container:
    """Tarjeta grande de la banda superior (hero): ícono en acento + valor
    grande + etiqueta. `valor` puede ser una Exception (consulta fallida)."""
    if isinstance(valor, Exception):
        valor_texto, valor_color = "—", ft.Colors.ON_SURFACE_VARIANT
    else:
        valor_texto, valor_color = formato_compacto(valor), ft.Colors.ON_SURFACE
    contenido = [
        ft.Row(
            [
                ft.Container(
                    ft.Icon(icono, color=color, size=18),
                    width=34, height=34, border_radius=10,
                    bgcolor=ft.Colors.with_opacity(0.14, color),
                    alignment=ft.Alignment.CENTER,
                ),
                ft.Text(etiqueta, size=12, color=ft.Colors.ON_SURFACE_VARIANT, expand=True,
                        max_lines=2, overflow=ft.TextOverflow.ELLIPSIS),
            ],
            spacing=10,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        ft.Text(valor_texto, size=26, weight=ft.FontWeight.W_700, color=valor_color),
    ]
    if subtexto:
        contenido.append(ft.Text(subtexto, size=10, color=ft.Colors.ON_SURFACE_VARIANT))
    return ft.Container(
        content=ft.Column(contenido, spacing=6),
        padding=16,
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
        border=ft.Border(
            top=ft.BorderSide(3, color),
            left=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
            right=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
            bottom=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
        ),
        border_radius=12,
        shadow=sombra_tarjeta(),
        height=124,  # alto fijo → las 4 tarjetas del hero quedan uniformes
        col={"xs": 12, "sm": 6, "lg": 3},  # ancho responsivo: 4 por fila en pantallas anchas
    )


def tile_compacta(etiqueta: str, valor, color: str, icono, subtexto: str = "", col=None) -> ft.Container:
    """Tarjeta KPI compacta: como `hero_tile` pero pensada para paneles a
    media pantalla (ej. Proyección/Cobranza, cada uno la mitad del ancho).
    Sin alto fijo (crece con el contenido) y con `max_lines` + elipsis tanto
    en la etiqueta como en el subtexto — `hero_tile` solo protege la
    etiqueta, así que su subtexto se desbordaba del alto fijo de 124px en
    columnas angostas (texto amontonado saliéndose de la tarjeta). `col` por
    defecto dos tarjetas por fila; pásalo explícito si el número de tarjetas
    pide otra fracción (ej. 3 tarjetas → col={"xs": 12, "sm": 4})."""
    if isinstance(valor, Exception):
        valor_texto, valor_color = "—", ft.Colors.ON_SURFACE_VARIANT
    else:
        valor_texto, valor_color = formato_compacto(valor), ft.Colors.ON_SURFACE
    contenido = [
        ft.Row(
            [
                ft.Container(
                    ft.Icon(icono, color=color, size=15),
                    width=28, height=28, border_radius=8,
                    bgcolor=ft.Colors.with_opacity(0.14, color),
                    alignment=ft.Alignment.CENTER,
                ),
                ft.Text(etiqueta, size=11, color=ft.Colors.ON_SURFACE_VARIANT, expand=True,
                        max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
            ],
            spacing=8,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        ft.Text(valor_texto, size=19, weight=ft.FontWeight.W_700, color=valor_color),
    ]
    if subtexto:
        contenido.append(ft.Text(subtexto, size=9.5, color=ft.Colors.ON_SURFACE_VARIANT,
                                  max_lines=1, overflow=ft.TextOverflow.ELLIPSIS, tooltip=subtexto))
    return ft.Container(
        content=ft.Column(contenido, spacing=5),
        padding=12,
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
        border=ft.Border(
            top=ft.BorderSide(3, color),
            left=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
            right=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
            bottom=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
        ),
        border_radius=10,
        shadow=sombra_tarjeta(),
        col=col or {"xs": 12, "sm": 6},
    )


def _leyenda_fila(color: str, etiqueta: str, valor: float, total: float) -> ft.Row:
    pct = (valor / total * 100) if total else 0
    return ft.Row(
        [
            ft.Container(width=10, height=10, bgcolor=color, border_radius=5),
            ft.Text(etiqueta, size=11, color=ft.Colors.ON_SURFACE, expand=True,
                    max_lines=1, overflow=ft.TextOverflow.ELLIPSIS, tooltip=etiqueta),
            ft.Text(formato_compacto(valor), size=11, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE),
            ft.Container(
                ft.Text(f"{pct:.0f}%", size=10, color=ft.Colors.ON_SURFACE_VARIANT),
                width=36, alignment=ft.Alignment.CENTER_RIGHT,
            ),
        ],
        spacing=8,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
    )


def construir_donut(items: list[tuple[str, float]], dark: bool) -> ft.Control:
    """Composición parte-todo (Empresa, Tipo de negocio): gráfica de DONA con el
    total en el centro y una leyenda (color · nombre · monto · %) al lado."""
    total = sum(v for _, v in items) or 1
    secciones = [
        fc.PieChartSection(
            value=float(valor) if valor > 0 else 0.0001,
            color=color_slot(i, dark),
            radius=26,
            title="",
        )
        for i, (_etiqueta, valor) in enumerate(items)
    ]
    pie = fc.PieChart(sections=secciones, center_space_radius=44, sections_space=2, expand=True)
    centro = ft.Container(
        content=ft.Column(
            [
                ft.Text("Total", size=9, color=ft.Colors.ON_SURFACE_VARIANT),
                ft.Text(formato_compacto(total), size=15, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE),
            ],
            spacing=0,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            alignment=ft.MainAxisAlignment.CENTER,
        ),
        alignment=ft.Alignment.CENTER,
    )
    grafica = ft.Container(ft.Stack([pie, centro]), width=150, height=150)
    leyenda = ft.Column(
        [_leyenda_fila(color_slot(i, dark), et, val, total) for i, (et, val) in enumerate(items)],
        spacing=10,
        expand=True,
    )
    return ft.Row(
        [grafica, ft.Container(leyenda, expand=True, padding=ft.Padding(left=8, right=0, top=6, bottom=6))],
        spacing=14,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
        height=180,
    )


_ANCHO_TRACK = 190  # px del carril del ranking, a valor máximo (versión compacta)


def construir_ranked_list(items: list[tuple[str, float]], dark: bool) -> ft.Control:
    """Ranking tipo leaderboard (muchas categorías: sucursales / empresa+sucursal):
    posición + nombre + barra sobre un carril tenue + monto. El carril de fondo da
    la escala visual sin necesidad de ejes; orden descendente (viene de la query).
    Column con scroll propio: nada se trunca."""
    color = color_slot(0, dark)
    track_bg = ft.Colors.with_opacity(0.08, ft.Colors.ON_SURFACE)
    max_total = max((v for _, v in items), default=0) or 1
    filas = []
    for i, (etiqueta, valor) in enumerate(items):
        ancho = max(3, round(_ANCHO_TRACK * (valor / max_total)))
        filas.append(
            ft.Row(
                [
                    ft.Container(
                        ft.Text(f"{i + 1}", size=10, color=ft.Colors.ON_SURFACE_VARIANT),
                        width=16, alignment=ft.Alignment.CENTER_RIGHT,
                    ),
                    ft.Container(
                        ft.Text(etiqueta, size=10, color=ft.Colors.ON_SURFACE,
                                max_lines=1, overflow=ft.TextOverflow.ELLIPSIS, tooltip=etiqueta),
                        width=104,
                    ),
                    ft.Container(
                        content=ft.Stack(
                            [
                                ft.Container(width=_ANCHO_TRACK, height=12, bgcolor=track_bg, border_radius=6),
                                ft.Container(width=ancho, height=12, bgcolor=color, border_radius=6,
                                             tooltip=formato_compacto(valor)),
                            ]
                        ),
                        width=_ANCHO_TRACK, height=12,
                    ),
                    ft.Text(formato_compacto(valor), size=10, weight=ft.FontWeight.W_500,
                            color=ft.Colors.ON_SURFACE),
                ],
                spacing=8,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            )
        )
    return ft.Column(filas, spacing=8, scroll=ft.ScrollMode.AUTO, height=200)


def estado_vacio() -> ft.Control:
    """Sin esto, un rango sin movimientos se veía como una pantalla en blanco
    (¿falló? ¿sigue cargando?) — un empty state explícito es parte del
    contrato de cualquier vista con datos, no un caso aparte."""
    return ft.Container(
        content=ft.Column(
            [
                ft.Icon(ft.Icons.INBOX_OUTLINED, size=24, color=ft.Colors.OUTLINE_VARIANT),
                ft.Text("Sin movimientos en este rango", size=12, color=ft.Colors.ON_SURFACE_VARIANT),
            ],
            spacing=6,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        alignment=ft.Alignment.CENTER,
        height=120,
    )


def placeholder_carga(mensaje: str = "Consultando BigQuery…") -> ft.Control:
    return ft.Container(
        content=ft.Column(
            [
                ft.ProgressRing(width=24, height=24, stroke_width=2),
                ft.Text(mensaje, size=12, color=ft.Colors.ON_SURFACE_VARIANT),
            ],
            spacing=10,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        alignment=ft.Alignment.CENTER,
        height=200,
    )


def construir_tabla(items: list[tuple[str, float]], dark: bool, un_solo_color: bool) -> ft.Control:
    filas = []
    for i, (etiqueta, valor) in enumerate(items):
        color = color_slot(0, dark) if un_solo_color else color_slot(i, dark)
        filas.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(
                        ft.Row(
                            [
                                ft.Container(width=8, height=8, bgcolor=color, border_radius=4),
                                ft.Text(etiqueta, size=11),
                            ],
                            spacing=6,
                        )
                    ),
                    ft.DataCell(ft.Text(f"${valor:,.2f}", size=11)),
                ]
            )
        )
    tabla = ft.DataTable(
        columns=[ft.DataColumn(ft.Text("Categoría", size=11)), ft.DataColumn(ft.Text("Total", size=11), numeric=True)],
        rows=filas,
        data_row_max_height=32,
        heading_row_height=32,
        column_spacing=16,
    )
    return ft.Column([tabla], scroll=ft.ScrollMode.AUTO, height=200)


def construir_tabla_moneda(items: list[tuple[str, float, float, float, float]], dark: bool) -> ft.Control:
    """Tabla detallada con el desglose pesos/dólares — la usan las secciones
    del dashboard en modo tabla, para separar el USD por categoría (no solo
    el agregado de la pastilla). Cada item es
    (etiqueta, mxn, usd, usd_convertido, usd_sin_tc):

    - MXN (sin USD): pesos únicamente.
    - USD: dólares crudos, sin convertir.
    - MXN convertido: USD × tipo de cambio promedio del día que le tocó.
    - Total final: MXN (sin USD) + MXN convertido — el USD sin tipo de
      cambio (columna aparte, si aplica) NO entra en este total, porque no
      hay con qué convertirlo, no porque se ignore: sigue visible en su
      propia columna."""
    hay_sin_tc = any(sin_tc for _et, _m, _u, _c, sin_tc in items)
    columnas = [
        ft.DataColumn(ft.Text("Categoría", size=11)),
        ft.DataColumn(ft.Text("MXN (sin USD)", size=11), numeric=True),
        ft.DataColumn(ft.Text("USD", size=11), numeric=True),
        ft.DataColumn(ft.Text("MXN convertido", size=11), numeric=True),
        ft.DataColumn(ft.Text("Total final", size=11), numeric=True),
    ]
    if hay_sin_tc:
        columnas.append(ft.DataColumn(ft.Text("USD sin tipo de cambio", size=11), numeric=True))

    filas = []
    for i, (etiqueta, mxn, usd, convertido, sin_tc) in enumerate(items):
        color = color_slot(i, dark)
        total_final = mxn + convertido
        celdas = [
            ft.DataCell(
                ft.Row(
                    [
                        ft.Container(width=8, height=8, bgcolor=color, border_radius=4),
                        ft.Container(
                            ft.Text(etiqueta, size=11, max_lines=1, overflow=ft.TextOverflow.ELLIPSIS,
                                     tooltip=etiqueta),
                            width=160,
                        ),
                    ],
                    spacing=6,
                )
            ),
            ft.DataCell(ft.Text(f"${mxn:,.2f}", size=11)),
            ft.DataCell(ft.Text(f"US${usd:,.2f}" if usd else "—", size=11)),
            ft.DataCell(ft.Text(f"${convertido:,.2f}" if convertido else "—", size=11)),
            ft.DataCell(ft.Text(f"${total_final:,.2f}", size=11, weight=ft.FontWeight.W_600)),
        ]
        if hay_sin_tc:
            celdas.append(
                ft.DataCell(
                    ft.Text(f"US${sin_tc:,.2f}" if sin_tc else "—", size=11,
                             color=ft.Colors.RED_600 if sin_tc else ft.Colors.ON_SURFACE_VARIANT)
                )
            )
        filas.append(ft.DataRow(cells=celdas))

    total_mxn = sum(mxn for _et, mxn, _usd, _convertido, _sin_tc in items)
    total_usd = sum(usd for _et, _mxn, usd, _convertido, _sin_tc in items)
    total_convertido = sum(convertido for _et, _mxn, _usd, convertido, _sin_tc in items)
    total_final = total_mxn + total_convertido
    total_sin_tc = sum(sin_tc for _et, _mxn, _usd, _convertido, sin_tc in items)
    estilo_total = {"size": 11, "weight": ft.FontWeight.W_700, "color": ft.Colors.ON_SURFACE}
    celdas_total = [
        ft.DataCell(ft.Text("TOTAL", **estilo_total)),
        ft.DataCell(ft.Text(f"${total_mxn:,.2f}", **estilo_total)),
        ft.DataCell(ft.Text(f"US${total_usd:,.2f}" if total_usd else "—", **estilo_total)),
        ft.DataCell(ft.Text(f"${total_convertido:,.2f}" if total_convertido else "—", **estilo_total)),
        ft.DataCell(ft.Text(f"${total_final:,.2f} MXN", **estilo_total)),
    ]
    if hay_sin_tc:
        celdas_total.append(
            ft.DataCell(
                ft.Text(
                    f"US${total_sin_tc:,.2f}" if total_sin_tc else "—",
                    **estilo_total,
                )
            )
        )
    filas.append(
        ft.DataRow(
            cells=celdas_total,
            color=ft.Colors.with_opacity(0.08, color_slot(0, dark)),
        )
    )

    tabla = ft.DataTable(
        columns=columnas,
        rows=filas,
        data_row_max_height=32,
        heading_row_height=32,
        column_spacing=16,
    )
    return ft.Column([tabla], scroll=ft.ScrollMode.AUTO, height=200)


# --- Timeline (explorador) ----------------------------------------------------
# NOTA: se descartó fc.LineChart (flet_charts) para esta vista — embebido en el
# árbol completo de la app el widget se pintaba gris en algunos estados de
# layout. Una barra vertical simple con Container es robusta y mantiene el
# criterio de "una sola serie, color de marca" de construir_ranked_list.

_ALTO_BARRA_TIMELINE = 180  # px del carril vertical, a valor máximo


def _etiqueta_periodo(periodo: date, modo: str) -> str:
    return periodo.strftime("%b %Y") if modo == "mensual" else f"Sem {periodo.strftime('%d %b')}"


def construir_timeline(items: list[tuple[date, float]], dark: bool, periodo: str) -> ft.Control:
    """Serie temporal (una sola serie, color de marca) mensual o semanal:
    una barra vertical por periodo, alineadas sobre una base común, con el
    monto arriba y la etiqueta del periodo abajo. Scroll horizontal si hay
    muchos periodos."""
    if not items:
        return estado_vacio()
    max_total = max((v for _, v in items), default=0) or 1
    color = color_slot(0, dark)

    barras = []
    for periodo_fecha, valor in items:
        alto = max(3, round(_ALTO_BARRA_TIMELINE * (valor / max_total)))
        barras.append(
            ft.Column(
                [
                    ft.Text(formato_compacto(valor), size=10, weight=ft.FontWeight.W_500, color=ft.Colors.ON_SURFACE),
                    ft.Container(
                        width=32, height=_ALTO_BARRA_TIMELINE,
                        alignment=ft.Alignment.BOTTOM_CENTER,
                        content=ft.Container(
                            width=32, height=alto, bgcolor=color,
                            border_radius=ft.BorderRadius(top_left=4, top_right=4, bottom_left=0, bottom_right=0),
                            tooltip=formato_compacto(valor),
                        ),
                    ),
                    ft.Text(_etiqueta_periodo(periodo_fecha, periodo), size=10, color=ft.Colors.ON_SURFACE_VARIANT),
                ],
                spacing=6,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            )
        )
    return ft.Row(barras, spacing=18, scroll=ft.ScrollMode.AUTO, vertical_alignment=ft.CrossAxisAlignment.END)
