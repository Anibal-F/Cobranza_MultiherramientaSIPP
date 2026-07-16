"""Sub-pestañas 'Timeline' y 'Detalle': explorador ABIERTO sobre la misma
tabla — sin la restricción de "segmento principal" de Segmentado — con un
panel de filtros compartido (rango de fechas, empresa, sucursal, tipo de
negocio y filial/sn_PagoFilial) y un modal de transparencia que muestra los
filtros/transformaciones SQL activos.

La UI de estas sub-pestañas se RECONSTRUYE DESDE CERO cada vez que se entra
(y tras cada carga de datos): remontar instancias de controles ya usadas en
este árbol dejaba zonas pintadas como un recuadro gris sin excepción visible
(bug de render de Flutter web). Solo el ESTADO (filtros, resultados, catálogo)
persiste entre reconstrucciones, en los atributos de la clase Explorador.

Ojo con `ft.Row(wrap=True)`: NO admite hijos con `expand=True` (Flutter no
permite Expanded dentro de Wrap y el fallo de layout pinta gris TODO lo que
sigue en la página, sin lanzar error a la consola JS) — ese fue el origen del
recuadro gris de la primera versión del panel de filtros.
"""

import asyncio
from datetime import date, datetime

import flet as ft
from flet_datatable2 import DataColumn2, DataColumnSize, DataTable2

from .componentes import (
    chip_total_usd,
    color_slot,
    construir_timeline,
    escribir_hoja_excel,
    estado_vacio,
    guardar_workbook,
    mostrar_dialogo,
    placeholder_carga,
    sombra_tarjeta,
)
from .consultas import (
    LIMITE_FILAS_DETALLE,
    MONEDA_USD,
    consultar_catalogo,
    consultar_detalle_completo_periodo,
    consultar_detalle_movimientos,
    consultar_serie_temporal,
)


def _es_usd(fila: dict) -> bool:
    return (fila.get("nb_Moneda") or "").strip().lower() == MONEDA_USD


ETIQUETAS_FILIAL = {"todos": "Todos", "excluir": "Excluye entre filiales", "solo": "Solo entre filiales"}
_ETIQUETA_TODOS = {"empresa": "Todas", "sucursal": "Todas", "tipo_negocio": "Todos"}
_ATRIBUTO_FILTRO = {"empresa": "empresas", "sucursal": "sucursales", "tipo_negocio": "tipos_negocio"}
# Campo de la fila por índice de columna de la tabla de detalle (para ordenar).
_CAMPOS_ORDEN = [
    "fh_Envio", "nb_Empresa", "nb_sucursal", "tipo_negocio_efectivo",
    "de_RazonSocial", "im_Movimiento", "nb_Moneda", "sn_PagoFilial", "sn_Identificada",
]
# Máximo de filas que se RENDEREAN en la tabla (los datos completos, hasta
# LIMITE_FILAS_DETALLE, sí viven en memoria para filtrar/ordenar). Serializar
# miles de DataRow en un solo update satura el websocket de Flet.
_MAX_FILAS_RENDER = 200


class Explorador:
    """Estado + constructores de UI de las sub-pestañas Timeline y Detalle.

    Uso desde el ensamblado de la pestaña: `construir("timeline"|"detalle")`
    devuelve un control NUEVO cada vez (nada se reutiliza entre entradas); la
    primera entrada dispara la carga de datos. Los cambios de filtro recargan
    y reconstruyen la vista activa.
    """

    def __init__(self, page: ft.Page):
        self.page = page
        # En Flet 0.85 el FilePicker es un servicio: se crea y se usa
        # directamente (NO se agrega a page.overlay; hacerlo provoca "Unknown
        # control: FilePicker").
        self.file_picker = ft.FilePicker()
        hoy = date.today()
        # --- Filtros compartidos por Timeline y Detalle ---
        self.rango: tuple[date, date] = (hoy.replace(day=1), hoy)
        self.empresas: list[str] = []       # [] = todas
        self.sucursales: list[str] = []
        self.tipos_negocio: list[str] = []
        self.filial = "excluir"             # "todos" | "excluir" | "solo" — default = criterio de Segmentado
        self.periodo = "mensual"            # "mensual" | "semanal" (solo Timeline)
        # --- Filtros/orden locales de la tabla de detalle (en memoria) ---
        self.filtros_columna = {"nb_Empresa": "", "nb_sucursal": "", "tipo_negocio_efectivo": "", "de_RazonSocial": ""}
        self.orden: tuple[int, bool] = (0, False)  # (columna, ascendente) — default fecha desc, como la query
        # --- Datos ---
        self.catalogo: dict[str, list[str] | None] = {"empresa": None, "sucursal": None, "tipo_negocio": None}
        self.serie: list[tuple[date, float]] = []  # MXN únicamente — alimenta la gráfica
        self.serie_usd: list[tuple[date, float]] = []  # USD por periodo, aparte — nunca se suma a self.serie
        self.detalle: list[dict] = []
        self.truncado = False
        self.cargado = False
        self.cargando = False
        self.error = ""
        # --- UI ---
        self.activa = "timeline"
        self.raiz: ft.Column | None = None  # el contenedor actualmente montado (se crea en construir())

    # --- Ciclo de vida --------------------------------------------------------

    def construir(self, subtab: str) -> ft.Control:
        """Control NUEVO con panel de filtros + sub-pestaña `subtab`. La
        primera vez dispara la carga de datos (perezosa: no cuesta queries
        mientras el usuario no entre al explorador)."""
        self.activa = subtab
        self.raiz = ft.Column(expand=True, spacing=0, controls=self._contenido())
        if not self.cargado and not self.cargando:
            self.page.run_task(self.cargar)
        return self.raiz

    def _contenido(self) -> list[ft.Control]:
        panel = self._panel_timeline() if self.activa == "timeline" else self._panel_detalle()
        return [self._panel_filtros(), panel]

    def _refrescar(self) -> None:
        """Reconstruye la vista activa en el contenedor montado (si lo hay)."""
        if self.raiz is not None:
            self.raiz.controls = self._contenido()
        self.page.update()

    async def cargar(self, _e=None) -> None:
        """Consulta timeline + detalle (+ catálogos pendientes) en paralelo con
        los filtros actuales; un fallo no tumba al resto (return_exceptions)."""
        self.cargando = True
        self.error = ""
        self._refrescar()

        fi, ff = self.rango
        tareas = [
            consultar_serie_temporal(fi, ff, self.periodo, self.empresas, self.sucursales, self.tipos_negocio, self.filial),
            consultar_detalle_movimientos(fi, ff, self.empresas, self.sucursales, self.tipos_negocio, self.filial),
        ]
        faltantes = [c for c, v in self.catalogo.items() if v is None]
        tareas += [consultar_catalogo(c) for c in faltantes]
        resultados = await asyncio.gather(*tareas, return_exceptions=True)

        r_serie, r_detalle, *r_catalogos = resultados
        if not isinstance(r_serie, Exception):
            self.serie = [(fila["periodo"], fila["total"] or 0) for fila in r_serie]
            self.serie_usd = [(fila["periodo"], fila["total_usd"] or 0) for fila in r_serie]
        if not isinstance(r_detalle, Exception):
            self.detalle, self.truncado = r_detalle
        for clave, resultado in zip(faltantes, r_catalogos):
            if not isinstance(resultado, Exception):
                self.catalogo[clave] = resultado
        if isinstance(r_serie, Exception) or isinstance(r_detalle, Exception):
            self.error = "No se pudo consultar BigQuery con estos filtros."

        self.cargado = True
        self.cargando = False
        self._refrescar()

    def _recargar(self) -> None:
        self.page.run_task(self.cargar)

    def _dark(self) -> bool:
        return self.page.theme_mode == ft.ThemeMode.DARK

    # --- Panel de filtros compartido ------------------------------------------

    def _panel_filtros(self) -> ft.Control:
        fi, ff = self.rango
        boton_rango = ft.OutlinedButton(
            content=ft.Row(
                [ft.Icon(ft.Icons.DATE_RANGE, size=16),
                 ft.Text(f"{fi.strftime('%d %b %Y')} – {ff.strftime('%d %b %Y')}", size=13)],
                spacing=8, tight=True,
            ),
            style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
            on_click=self._abrir_rango,
        )
        selector_filial = ft.SegmentedButton(
            segments=[ft.Segment(value=v, label=ft.Text(l, size=12)) for v, l in ETIQUETAS_FILIAL.items()],
            selected=[self.filial],
            on_change=self._on_filial,
        )
        boton_info = ft.IconButton(
            icon=ft.Icons.INFO_OUTLINE, icon_size=18,
            tooltip="Ver filtros y transformaciones SQL aplicados",
            on_click=self._abrir_info,
        )
        progreso = ft.ProgressRing(width=16, height=16, stroke_width=2, visible=self.cargando)
        error_text = ft.Text(self.error, size=12, color=ft.Colors.RED_600, visible=bool(self.error))

        # IMPORTANTE: nada de expand=True dentro de este Row — tiene wrap=True
        # (ver nota en el docstring del módulo). El botón ⓘ va como un chip más.
        fila = ft.Row(
            [
                boton_rango,
                self._boton_multiselect(ft.Icons.BUSINESS_OUTLINED, "Empresa", "empresa"),
                self._boton_multiselect(ft.Icons.STORE_OUTLINED, "Sucursal", "sucursal"),
                self._boton_multiselect(ft.Icons.CATEGORY_OUTLINED, "Tipo de negocio", "tipo_negocio"),
                selector_filial,
                boton_info,
                progreso,
                error_text,
            ],
            spacing=10, wrap=True, vertical_alignment=ft.CrossAxisAlignment.CENTER,
        )
        return ft.Container(
            content=fila,
            padding=ft.Padding(left=16, right=16, top=12, bottom=12),
            margin=ft.Margin(left=20, right=20, top=16, bottom=0),
            bgcolor=ft.Colors.SURFACE_CONTAINER_LOW,
            border_radius=12,
            shadow=sombra_tarjeta(),
        )

    def _abrir_rango(self, _e) -> None:
        fi, ff = self.rango
        picker = ft.DateRangePicker(
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2035, 12, 31),
            start_value=datetime.combine(fi, datetime.min.time()),
            end_value=datetime.combine(ff, datetime.min.time()),
            entry_mode=ft.DatePickerEntryMode.INPUT,  # diálogo compacto de texto, no el calendario grande
            on_change=self._on_rango,
        )
        self.page.show_dialog(picker)

    def _on_rango(self, e) -> None:
        if not e.control.start_value or not e.control.end_value:
            return
        self.rango = (e.control.start_value.date(), e.control.end_value.date())
        self._recargar()

    def _on_filial(self, e) -> None:
        self.filial = e.control.selected[0] if e.control.selected else "todos"
        self._recargar()

    # --- Selectores multi-valor (empresa / sucursal / tipo de negocio) --------
    # Dropdown anclado (PopupMenuButton), NO modal: un PopupMenuItem con un
    # Checkbox propio por valor del catálogo. Un Checkbox (o TextButton)
    # anidado dentro del content de un PopupMenuItem consume su propio tap y
    # NO cierra el menú — solo un PopupMenuItem con on_click en el ítem mismo
    # (el de "Aplicar" aquí abajo) lo cierra. Eso permite marcar varios
    # checkboxes seguidos sin que el menú se cierre en cada click, y solo
    # "Aplicar" dispara la recarga (comportamiento verificado a mano).
    #
    # Sin campo de búsqueda: se probó un TextField con on_change filtrando
    # (fila.visible = ...) y, aunque el checkbox SÍ refleja cambios de valor
    # en un menú ya abierto, un cambio de VISIBILIDAD de un PopupMenuItem no
    # se vuelve a pintar — el menú abierto es una foto fija de qué filas
    # existen. Mejor no ofrecer una búsqueda que aparenta funcionar y no hace
    # nada, que dejarla a medias.

    def _boton_multiselect(self, icono, etiqueta: str, clave: str) -> ft.PopupMenuButton:
        seleccion: list[str] = getattr(self, _ATRIBUTO_FILTRO[clave])
        resumen = f"{len(seleccion)} sel." if seleccion else _ETIQUETA_TODOS[clave]
        valores = self.catalogo.get(clave)

        if valores is None:
            items: list[ft.PopupMenuItem] = [
                ft.PopupMenuItem(
                    content=ft.Text(
                        "El catálogo aún se está cargando desde BigQuery; intenta de nuevo en unos segundos.",
                        size=12, color=ft.Colors.ON_SURFACE_VARIANT,
                    ),
                )
            ]
        else:
            seleccion_previa = set(seleccion)
            pendiente: dict[str, bool] = {v: (v in seleccion_previa) for v in valores}
            checks: list[ft.Checkbox] = []

            def _marcar_todos(_e) -> None:
                for chk in checks:
                    chk.value = True
                    pendiente[chk.label] = True
                self.page.update()

            def _limpiar(_e) -> None:
                for chk in checks:
                    chk.value = False
                    pendiente[chk.label] = False
                self.page.update()

            def _toggle(v: str):
                def _h(e) -> None:
                    pendiente[v] = e.control.value
                return _h

            def _aplicar(_e) -> None:
                setattr(self, _ATRIBUTO_FILTRO[clave], [v for v, marcado in pendiente.items() if marcado])
                self._recargar()

            filas_valor = []
            for v in valores:
                chk = ft.Checkbox(value=pendiente[v], label=v, on_change=_toggle(v))
                checks.append(chk)
                filas_valor.append(ft.PopupMenuItem(content=ft.Row([chk])))

            items = [
                ft.PopupMenuItem(
                    content=ft.Row([
                        ft.TextButton("Limpiar", on_click=_limpiar),
                        ft.TextButton("Todos", on_click=_marcar_todos),
                    ]),
                ),
                *filas_valor,
                ft.PopupMenuItem(
                    content=ft.Container(
                        content=ft.Text("Aplicar", weight=ft.FontWeight.BOLD, color=ft.Colors.PRIMARY),
                        alignment=ft.Alignment.CENTER,
                    ),
                    on_click=_aplicar,
                ),
            ]

        return ft.PopupMenuButton(
            content=ft.Container(
                content=ft.Row(
                    [ft.Icon(icono, size=16), ft.Text(etiqueta, size=13),
                     ft.Text(resumen, size=12, color=ft.Colors.ON_SURFACE_VARIANT)],
                    spacing=6, tight=True,
                ),
                padding=ft.Padding(left=12, right=12, top=6, bottom=6),
                border=ft.Border.all(1, ft.Colors.OUTLINE),
                border_radius=8,
            ),
            items=items,
        )

    # --- Modal de transparencia de la consulta --------------------------------

    def _abrir_info(self, _e) -> None:
        fi, ff = self.rango
        periodo_txt = "por mes" if self.periodo == "mensual" else "por semana"
        filial_txt = {
            "todos": "Se incluyen todos los movimientos, sean o no pagos entre filiales.",
            "excluir": "Se excluyen los pagos entre filiales (solo movimientos con clientes externos).",
            "solo": "Se muestran únicamente los pagos entre filiales.",
        }[self.filial]
        lineas = [
            f"Se muestran los movimientos del {fi.strftime('%d/%m/%Y')} al {ff.strftime('%d/%m/%Y')}.",
            f"Empresa: {', '.join(self.empresas) if self.empresas else 'todas las empresas'}.",
            f"Sucursal: {', '.join(self.sucursales) if self.sucursales else 'todas las sucursales'}.",
            f"Tipo de negocio: {', '.join(self.tipos_negocio) if self.tipos_negocio else 'todos los tipos de negocio'}.",
            filial_txt,
            f"El Timeline agrupa los totales {periodo_txt}.",
            "El tipo de negocio se reclasifica en tres casos: cuentas bancarias "
            "'Abastecedora SF /AENE' o 'Petroplazas SF' se cuentan como 'SF'; los "
            "clientes de 'Público en general' de Petro Smart se cuentan como "
            "'GasPetroil'; y un cliente específico (id 4359) se cuenta como "
            "'Distribuidora' — sin importar cómo esté registrado originalmente.",
            f"La tabla de Detalle muestra como máximo {LIMITE_FILAS_DETALLE:,} movimientos por consulta"
            + (" y con los filtros actuales se alcanzó ese límite (hay más movimientos de los que se muestran)."
               if self.truncado else "."),
            "Los movimientos en dólares (Moneda = 'Dolar (USD)') nunca se convierten ni se suman a los "
            "de pesos: el Timeline los deja aparte en la pastilla 'USD' (arriba de la gráfica), y en "
            "Detalle cada fila muestra su propia moneda en la columna 'Moneda', con el total MXN/USD del "
            "conjunto filtrado debajo de la tabla.",
            "A diferencia de la vista 'Segmentado', aquí no se excluyen por defecto "
            "los movimientos de GAS, Autotanque ni los de sucursal sin asignar — por "
            "eso los totales de esta vista no son directamente comparables con los de esa vista.",
        ]
        dialogo = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cómo se calculan estos datos"),
            content=ft.Container(
                content=ft.Column(
                    [ft.Text(f"•  {l}", size=12, selectable=True) for l in lineas],
                    spacing=10, scroll=ft.ScrollMode.AUTO,
                ),
                width=520, height=340,
            ),
            actions=[ft.TextButton("Cerrar", on_click=lambda _e: self.page.pop_dialog())],
        )
        mostrar_dialogo(self.page, dialogo)

    # --- Sub-pestaña Timeline --------------------------------------------------

    def _on_periodo(self, e) -> None:
        self.periodo = e.control.selected[0] if e.control.selected else "mensual"
        self._recargar()  # la agrupación (mes/semana) vive en la query

    def _etiqueta_periodo_excel(self, periodo: date) -> str:
        return periodo.strftime("%b %Y") if self.periodo == "mensual" else f"Sem {periodo.strftime('%d %b %Y')}"

    def _texto_filtros_activos(self) -> str:
        fi, ff = self.rango
        return (
            f"Rango: {fi.strftime('%d/%m/%Y')} - {ff.strftime('%d/%m/%Y')} | "
            f"Empresa: {', '.join(self.empresas) if self.empresas else 'todas'} | "
            f"Sucursal: {', '.join(self.sucursales) if self.sucursales else 'todas'} | "
            f"Tipo de negocio: {', '.join(self.tipos_negocio) if self.tipos_negocio else 'todos'} | "
            f"Filial: {ETIQUETAS_FILIAL[self.filial]}"
        )

    def _panel_timeline(self) -> ft.Control:
        selector_periodo = ft.SegmentedButton(
            segments=[
                ft.Segment(value="mensual", label=ft.Text("Mensual", size=12)),
                ft.Segment(value="semanal", label=ft.Text("Semanal", size=12)),
            ],
            selected=[self.periodo],
            on_change=self._on_periodo,
        )
        if self.cargando and not self.serie:
            cuerpo = placeholder_carga()
        else:
            cuerpo = construir_timeline(self.serie, self._dark(), self.periodo)

        estado_exportar = ft.Text("", size=11, color=ft.Colors.ON_SURFACE_VARIANT)

        async def _exportar(_e) -> None:
            boton_exportar.disabled = True
            estado_exportar.value = "Generando Excel…"
            self.page.update()

            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Timeline"
            usd_por_periodo = dict(self.serie_usd)
            escribir_hoja_excel(
                ws,
                ["Periodo", "Total MXN", "Total USD"],
                [
                    [self._etiqueta_periodo_excel(periodo), round(total, 2), round(usd_por_periodo.get(periodo, 0), 2)]
                    for periodo, total in self.serie
                ],
                fila_inicio=3,
            )
            ws.cell(row=1, column=1, value=self._texto_filtros_activos())
            ws.cell(row=2, column=1, value=f"Agrupación: {'mensual' if self.periodo == 'mensual' else 'semanal'}")

            fi, ff = self.rango
            nombre_def = f"dashboard_timeline_{fi:%Y%m%d}_{ff:%Y%m%d}.xlsx"
            ok, mensaje = await guardar_workbook(self.page, self.file_picker, wb, nombre_def)
            boton_exportar.disabled = False
            estado_exportar.value = mensaje
            self.page.update()

        boton_exportar = ft.IconButton(
            icon=ft.Icons.DOWNLOAD,
            icon_size=18,
            tooltip="Descargar Excel del timeline (con los filtros actuales)",
            disabled=not self.serie,
            on_click=lambda e: self.page.run_task(_exportar, e),
        )

        total_usd_periodo = sum(v for _p, v in self.serie_usd)

        return ft.Container(
            content=ft.Column(
                [
                    ft.Row(
                        [
                            ft.Container(
                                ft.Icon(ft.Icons.SHOW_CHART, color=color_slot(0, self._dark()), size=16),
                                width=30, height=30, border_radius=9,
                                bgcolor=ft.Colors.with_opacity(0.14, color_slot(0, self._dark())),
                                alignment=ft.Alignment.CENTER,
                            ),
                            ft.Text("Timeline de ingresos", size=16, weight=ft.FontWeight.W_600,
                                    color=ft.Colors.ON_SURFACE),
                            ft.Container(expand=True),  # Row sin wrap: aquí expand sí es válido
                            chip_total_usd(total_usd_periodo),
                            selector_periodo,
                            boton_exportar,
                        ],
                        spacing=10,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                    ft.Text(
                        "Explorador abierto: sin la restricción de segmento principal de 'Segmentado' "
                        "— ver el botón ⓘ arriba para el detalle de filtros y transformaciones. La línea es "
                        "en pesos; el total en dólares del periodo (si hay) se muestra en la pastilla 'USD'.",
                        size=11, color=ft.Colors.ON_SURFACE_VARIANT,
                    ),
                    estado_exportar,
                    cuerpo,
                ],
                spacing=12, expand=True, scroll=ft.ScrollMode.AUTO,
            ),
            padding=20, expand=True,
        )

    # --- Sub-pestaña Detalle ----------------------------------------------------

    def _filas_filtradas(self) -> list[dict]:
        filas = self.detalle
        for campo, texto in self.filtros_columna.items():
            texto = (texto or "").strip().lower()
            if texto:
                filas = [f for f in filas if texto in (f[campo] or "").lower()]
        indice, ascendente = self.orden
        campo_orden = _CAMPOS_ORDEN[indice]
        return sorted(filas, key=lambda f: (f[campo_orden] is None, f[campo_orden]), reverse=not ascendente)

    @staticmethod
    def _fila(f: dict) -> ft.DataRow:
        fecha = f["fh_Envio"]
        moneda_usd = _es_usd(f)
        etiqueta_moneda = "USD" if moneda_usd else "MXN"
        signo = "US$" if moneda_usd else "$"
        return ft.DataRow(cells=[
            ft.DataCell(ft.Text(fecha.strftime("%d/%m/%Y") if fecha else "—", size=11)),
            ft.DataCell(ft.Text(f["nb_Empresa"] or "—", size=11)),
            ft.DataCell(ft.Text(f["nb_sucursal"] or "—", size=11)),
            ft.DataCell(ft.Text(f["tipo_negocio_efectivo"] or "—", size=11)),
            ft.DataCell(ft.Text(f["de_RazonSocial"] or "—", size=11)),
            ft.DataCell(ft.Text(f"{signo}{(f['im_Movimiento'] or 0):,.2f}", size=11,
                                 color="#eda100" if moneda_usd else None)),
            ft.DataCell(ft.Text(etiqueta_moneda, size=11)),
            ft.DataCell(ft.Text(f["sn_PagoFilial"] or "—", size=11)),
            ft.DataCell(ft.Text(f["sn_Identificada"] or "—", size=11)),
        ])

    def _texto_banner(self, n_filtradas: int) -> str:
        """Aviso combinado de truncamientos: el de BigQuery (LIMITE_FILAS_DETALLE)
        y el de render (_MAX_FILAS_RENDER). Cadena vacía = sin aviso."""
        partes = []
        if self.truncado:
            partes.append(
                f"La consulta trajo solo las primeras {LIMITE_FILAS_DETALLE:,} filas "
                "(ordenadas por fecha desc.) — acota el rango de fechas o los filtros del panel."
            )
        if n_filtradas > _MAX_FILAS_RENDER:
            partes.append(
                f"Mostrando las primeras {_MAX_FILAS_RENDER:,} de {n_filtradas:,} filas "
                "— usa los filtros por columna para acotar."
            )
        return " ".join(partes)

    def _panel_detalle(self) -> ft.Control:
        if self.cargando and not self.detalle:
            return ft.Container(
                content=ft.Column(
                    [ft.Text("Detalle de movimientos", size=16, weight=ft.FontWeight.W_600,
                             color=ft.Colors.ON_SURFACE), placeholder_carga()],
                    spacing=12, expand=True,
                ),
                padding=20, expand=True,
            )

        tabla = DataTable2(
            rows=[],
            fixed_top_rows=1, column_spacing=16, min_width=1050, expand=True,
            heading_row_height=36, data_row_height=32,
            empty=estado_vacio(),
        )
        texto_banner = ft.Text("", size=11, color=ft.Colors.ON_SURFACE_VARIANT)
        banner = ft.Container(
            content=texto_banner,
            bgcolor=ft.Colors.with_opacity(0.12, "#eda100"), padding=8, border_radius=8,
            visible=False,
        )
        texto_totales = ft.Text("", size=11, color=ft.Colors.ON_SURFACE_VARIANT)

        def _aplicar_estado_tabla() -> None:
            """Repone filas/orden/banner/totales de la tabla montada. IMPORTANTE:
            solo se mandan al cliente las primeras _MAX_FILAS_RENDER filas —
            serializar las 5,000 completas en un mensaje tumba el canal de
            Flet (la vista queda pintada pero todos los eventos posteriores
            mueren, sin error visible). Los totales sí se calculan sobre TODAS
            las filas filtradas (no solo las renderizadas), separando MXN de
            USD — nunca se suman entre sí."""
            filas = self._filas_filtradas()
            tabla.rows = [self._fila(f) for f in filas[:_MAX_FILAS_RENDER]]
            tabla.sort_column_index, tabla.sort_ascending = self.orden
            texto_banner.value = self._texto_banner(len(filas))
            banner.visible = bool(texto_banner.value)
            total_mxn = sum((f["im_Movimiento"] or 0) for f in filas if not _es_usd(f))
            total_usd = sum((f["im_Movimiento"] or 0) for f in filas if _es_usd(f))
            texto_totales.value = (
                f"{len(filas):,} movimiento(s) · Total MXN: ${total_mxn:,.2f}"
                + (f" · Total USD: ${total_usd:,.2f}" if total_usd else "")
            )

        def _on_orden(indice: int, ascendente: bool) -> None:
            self.orden = (indice, ascendente)
            _aplicar_estado_tabla()
            self.page.update()

        def _on_filtro_columna(campo: str, valor: str) -> None:
            # Filtrado 100% en memoria (el dataset ya está acotado por
            # LIMITE_FILAS_DETALLE); no se reconstruye el panel para no robar
            # el foco del TextField.
            self.filtros_columna[campo] = valor
            _aplicar_estado_tabla()
            self.page.update()

        def _columna(texto: str, indice: int, numeric: bool = False, fixed_width=None, size=None) -> DataColumn2:
            return DataColumn2(
                ft.Text(texto, weight=ft.FontWeight.BOLD, size=12, color=ft.Colors.ON_SURFACE),
                numeric=numeric, fixed_width=fixed_width, size=size,
                on_sort=lambda e, i=indice: _on_orden(i, e.ascending),
            )

        tabla.columns = [
            _columna("Fecha", 0, fixed_width=100),
            _columna("Empresa", 1, fixed_width=160),
            _columna("Sucursal", 2, fixed_width=160),
            _columna("Tipo de negocio", 3, fixed_width=140),
            _columna("Razón social", 4, size=DataColumnSize.L),
            _columna("Monto", 5, numeric=True, fixed_width=120),
            _columna("Moneda", 6, fixed_width=80),
            _columna("Pago filial", 7, fixed_width=90),
            _columna("Identificada", 8, fixed_width=90),
        ]
        _aplicar_estado_tabla()

        def _campo_filtro(etiqueta: str, campo: str, ancho: int) -> ft.TextField:
            # El filtro se aplica con Enter o al salir del campo (no en cada
            # tecla: repoblar la tabla por keystroke genera tráfico inútil).
            aplicar = lambda e, c=campo: _on_filtro_columna(c, e.control.value or "")
            return ft.TextField(
                label=etiqueta, dense=True, width=ancho, value=self.filtros_columna[campo],
                tooltip="Presiona Enter para aplicar",
                on_submit=aplicar, on_blur=aplicar,
            )

        estado_exportar = ft.Text("", size=11, color=ft.Colors.ON_SURFACE_VARIANT)

        async def _exportar(_e) -> None:
            """A diferencia de la tabla en pantalla (que respeta los filtros
            del panel y el tope de LIMITE_FILAS_DETALLE), la descarga es un
            volcado completo del periodo: solo el filtro de fecha, sin
            empresa/sucursal/tipo de negocio/filial y sin límite de filas."""
            boton_exportar.disabled = True
            estado_exportar.value = "Consultando el periodo completo (sin filtros de empresa/sucursal/tipo/filial)…"
            self.page.update()

            fi, ff = self.rango
            try:
                filas_export = await consultar_detalle_completo_periodo(fi, ff)
            except Exception as error:  # noqa: BLE001 - se muestra en estado_exportar
                boton_exportar.disabled = False
                estado_exportar.value = f"No se pudo consultar BigQuery: {error}"
                self.page.update()
                return

            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Detalle"
            ws.cell(row=1, column=1,
                    value=f"Periodo: {fi.strftime('%d/%m/%Y')} - {ff.strftime('%d/%m/%Y')} "
                          "(sin filtros de empresa/sucursal/tipo de negocio/filial)")
            total_mxn_export = sum((f.get("im_Movimiento") or 0) for f in filas_export if not _es_usd(f))
            total_usd_export = sum((f.get("im_Movimiento") or 0) for f in filas_export if _es_usd(f))
            ws.cell(row=2, column=1,
                    value=f"Total MXN: ${total_mxn_export:,.2f} · Total USD: ${total_usd_export:,.2f} (aparte, sin convertir)")
            encabezados = list(filas_export[0].keys()) if filas_export else []
            escribir_hoja_excel(
                ws, encabezados,
                [[fila.get(col) for col in encabezados] for fila in filas_export],
                fila_inicio=3,
            )
            if "fh_Envio" in encabezados:
                col_fecha = encabezados.index("fh_Envio") + 1
                for fila_celdas in ws.iter_rows(min_row=4, min_col=col_fecha, max_col=col_fecha):
                    for celda in fila_celdas:
                        celda.number_format = "dd/mm/yyyy"

            nombre_def = f"dashboard_detalle_{fi:%Y%m%d}_{ff:%Y%m%d}.xlsx"
            ok, mensaje = await guardar_workbook(self.page, self.file_picker, wb, nombre_def)
            boton_exportar.disabled = False
            estado_exportar.value = mensaje or f"{len(filas_export)} registro(s) exportados."
            self.page.update()

        boton_exportar = ft.IconButton(
            icon=ft.Icons.DOWNLOAD,
            icon_size=18,
            tooltip="Descargar Excel del periodo completo (solo filtro de fecha, sin los demás filtros)",
            on_click=lambda e: self.page.run_task(_exportar, e),
        )

        return ft.Container(
            content=ft.Column(
                [
                    ft.Row(
                        [
                            ft.Container(
                                ft.Icon(ft.Icons.TABLE_ROWS_OUTLINED, color=color_slot(1, self._dark()), size=16),
                                width=30, height=30, border_radius=9,
                                bgcolor=ft.Colors.with_opacity(0.14, color_slot(1, self._dark())),
                                alignment=ft.Alignment.CENTER,
                            ),
                            ft.Text("Detalle de movimientos", size=16, weight=ft.FontWeight.W_600,
                                    color=ft.Colors.ON_SURFACE),
                            ft.Container(expand=True),
                            boton_exportar,
                        ],
                        spacing=10,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                    estado_exportar,
                    banner,
                    texto_totales,
                    ft.Row(
                        [
                            _campo_filtro("Empresa", "nb_Empresa", 160),
                            _campo_filtro("Sucursal", "nb_sucursal", 160),
                            _campo_filtro("Tipo de negocio", "tipo_negocio_efectivo", 140),
                            _campo_filtro("Razón social", "de_RazonSocial", 220),
                        ],
                        spacing=10, wrap=True,  # sin hijos expand (ver nota del módulo)
                    ),
                    ft.Container(content=tabla, expand=True),
                ],
                spacing=12, expand=True,
            ),
            padding=20, expand=True,
        )
