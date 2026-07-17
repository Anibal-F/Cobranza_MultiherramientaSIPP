"""Sub-pestaña 'Segmentado': la vista agregada original del dashboard — todas
las vistas del segmento principal visibles a la vez (Empresa, Tipo de negocio,
Sucursal, Sucursal Gasolineras, SF, Otras empresas y la banda hero de KPIs),
sobre un rango de fechas propio, con toggle gráfica/tabla."""

import asyncio
from datetime import date, datetime

import flet as ft

from .componentes import (
    chip_total,
    chip_total_usd,
    color_slot,
    construir_donut,
    construir_ranked_list,
    construir_tabla_moneda,
    encabezado_seccion,
    escribir_hoja_excel,
    estado_vacio,
    guardar_workbook,
    hero_tile,
    mostrar_dialogo,
    nombre_hoja_valido,
    sombra_tarjeta,
)
from ..services.dashboard_repository import DashboardRepository

# El repositorio se crea perezosamente (necesita credenciales de BigQuery); así
# no falla al importar este módulo si aún no hay credenciales configuradas.
_repo_holder: list[DashboardRepository | None] = [None]


def _repo() -> DashboardRepository:
    if _repo_holder[0] is None:
        _repo_holder[0] = DashboardRepository()
    return _repo_holder[0]


async def _consultar_segmento(
    fecha_inicio: date, fecha_fin: date, columna: str
) -> list[tuple[str, float, float, float, float]]:
    filas = await asyncio.to_thread(_repo().agregado_segmento_principal, fecha_inicio, fecha_fin, columna)
    return [
        (fila["etiqueta"], fila["total"] or 0, fila["total_usd"] or 0,
         fila["total_usd_convertido"] or 0, fila["total_usd_sin_tc"] or 0)
        for fila in filas
    ]


async def _consultar_sucursal_gas(fecha_inicio: date, fecha_fin: date) -> list[tuple[str, float, float, float, float]]:
    filas = await asyncio.to_thread(_repo().agregado_sucursal_gas, fecha_inicio, fecha_fin)
    return [
        (fila["etiqueta"], fila["total"] or 0, fila["total_usd"] or 0,
         fila["total_usd_convertido"] or 0, fila["total_usd_sin_tc"] or 0)
        for fila in filas
    ]


async def _consultar_sf(fecha_inicio: date, fecha_fin: date) -> list[tuple[str, float, float, float, float]]:
    filas = await asyncio.to_thread(_repo().agregado_sf, fecha_inicio, fecha_fin)
    return [
        (fila["etiqueta"], fila["total"] or 0, fila["total_usd"] or 0,
         fila["total_usd_convertido"] or 0, fila["total_usd_sin_tc"] or 0)
        for fila in filas
    ]


async def _consultar_otras_empresas(fecha_inicio: date, fecha_fin: date) -> list[dict]:
    """Una fila POR EMPRESA (no por empresa+filial): la consulta SQL trae 2
    filas por empresa (una por sn_PagoFilial), y aquí se pivotea en Python
    para que la UI muestre una sola fila por empresa con las columnas
    no_mxn/no_usd_convertido/no_usd_sin_tc y si_mxn/si_usd_convertido/
    si_usd_sin_tc lado a lado — evita la vista anterior, que amontonaba 2
    filas casi idénticas por empresa en el mismo leaderboard."""
    filas = await asyncio.to_thread(_repo().agregado_otras_empresas, fecha_inicio, fecha_fin)
    por_empresa: dict[str, dict] = {}
    for fila in filas:
        entrada = por_empresa.setdefault(fila["empresa"], {
            "empresa": fila["empresa"],
            "no_mxn": 0.0, "no_usd_convertido": 0.0, "no_usd_sin_tc": 0.0,
            "si_mxn": 0.0, "si_usd_convertido": 0.0, "si_usd_sin_tc": 0.0,
        })
        prefijo = "no" if fila["filial"] == "NO" else "si"
        entrada[f"{prefijo}_mxn"] = fila["total"] or 0
        entrada[f"{prefijo}_usd_convertido"] = fila["total_usd_convertido"] or 0
        entrada[f"{prefijo}_usd_sin_tc"] = fila["total_usd_sin_tc"] or 0
    items = list(por_empresa.values())
    items.sort(
        key=lambda it: it["no_mxn"] + it["no_usd_convertido"] + it["si_mxn"] + it["si_usd_convertido"],
        reverse=True,
    )
    return items


async def _consultar_no_identificado(fecha_inicio: date, fecha_fin: date) -> tuple[float, float]:
    """(total_mxn, total_usd)."""
    resultado = await asyncio.to_thread(_repo().total_no_identificado, fecha_inicio, fecha_fin)
    return resultado["total"], resultado["total_usd"]


# (titulo, subtitulo, consulta, vista, icono) — vista: "donut" (composición
# parte-todo) o "ranked" (leaderboard con muchas categorías). El color de cada
# tarjeta sale de su posición en esta lista (color_slot(indice, dark)) — así
# cada sección tiene un acento distinto, igual que las tarjetas KPI del hero.
_SECCIONES = [
    (
        "Empresa",
        "Asociados y Distribuidora · excluye pagos entre filiales, GAS, Autotanque, Corporativo y sucursal sin asignar",
        lambda fi, ff: _consultar_segmento(fi, ff, "nb_Empresa"),
        "donut",
        ft.Icons.BUSINESS_OUTLINED,
    ),
    (
        "Tipo de negocio",
        "Asociados y Distribuidora · excluye pagos entre filiales, GAS, Autotanque, Corporativo y sucursal sin asignar",
        lambda fi, ff: _consultar_segmento(fi, ff, "nb_TipoDeNegocio"),
        "donut",
        ft.Icons.CATEGORY_OUTLINED,
    ),
    (
        "Sucursal",
        "Asociados y Distribuidora · excluye pagos entre filiales, GAS, Autotanque, Corporativo y sucursal sin asignar",
        lambda fi, ff: _consultar_segmento(fi, ff, "nb_sucursal"),
        "ranked",
        ft.Icons.STORE_OUTLINED,
    ),
    (
        "Sucursal (Gaseras)",
        "Segmento GasPetroil · excluye pagos entre filiales",
        _consultar_sucursal_gas,
        "ranked",
        ft.Icons.LOCAL_GAS_STATION_OUTLINED,
    ),
    (
        "SF",
        "Segmento SF (de_CuentaBancaria = 'Abastecedora SF /AENE' o 'Petroplazas SF') · "
        "excluye pagos entre filiales · todas las sucursales",
        _consultar_sf,
        "ranked",
        ft.Icons.ACCOUNT_BALANCE_OUTLINED,
    ),
]
# 'Otras empresas' no vive en _SECCIONES: su forma de datos (una fila por
# empresa con columnas Filial NO / Filial SI) no encaja con el contrato
# (etiqueta, mxn, usd, usd_convertido, usd_sin_tc) que usa _construir_seccion
# — tiene su propia consulta, cabecera y tabla (ver más abajo), igual que
# 'Sin identificar' (_consultar_no_identificado) ya vivía fuera de esta lista.


def _construir_seccion(
    titulo: str, subtitulo: str, resultado, dark: bool, en_tabla: bool, vista: str, icono, color: str
) -> ft.Container:
    """Una sección = cabecera (ícono + título + subtítulo + pastillas Total/USD)
    + cuerpo (dona / ranking / tabla). `resultado` es una lista de tuplas
    (etiqueta, mxn, usd, usd_convertido, usd_sin_tc) por categoría si la
    consulta tuvo éxito (puede venir vacía), o una Exception si falló — cada
    sección se degrada de forma independiente, un fallo no tira el resto del
    dashboard. La gráfica (dona/ranking) usa SIEMPRE solo la columna MXN; el
    desglose USD/convertido/sin-TC por categoría solo se ve en modo tabla
    (ver `construir_tabla_moneda`), y el total USD agregado se muestra aparte
    en su propia pastilla, nunca sumado al total en pesos."""
    chips: list[ft.Control] = []
    if isinstance(resultado, Exception):
        cuerpo: ft.Control = ft.Container(
            content=ft.Text(f"No se pudo consultar: {resultado}", size=11, color=ft.Colors.RED_600),
            height=120,
            alignment=ft.Alignment.CENTER,
        )
    else:
        items = resultado
        if not items:
            cuerpo = estado_vacio()
        else:
            total = sum(mxn for _et, mxn, _usd, _conv, _sin_tc in items)
            total_usd = sum(usd for _et, _mxn, usd, _conv, _sin_tc in items)
            chips.append(chip_total(total))
            if en_tabla:
                cuerpo = construir_tabla_moneda(items, dark)
            elif vista == "donut":
                cuerpo = construir_donut([(et, mxn) for et, mxn, *_resto in items], dark)
            else:
                cuerpo = construir_ranked_list([(et, mxn) for et, mxn, *_resto in items], dark)
            if total_usd:
                chips.append(chip_total_usd(total_usd))

    cabecera = encabezado_seccion(icono, color, titulo, subtitulo, chips)
    return ft.Container(
        content=ft.Column([cabecera, ft.Divider(height=1), cuerpo], spacing=10),
        padding=16,
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
        border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
        border_radius=12,
        shadow=sombra_tarjeta(),
        col={"xs": 12, "lg": 6},  # ancho responsivo: 2 secciones por fila en pantallas anchas
    )


_TITULO_OTRAS_EMPRESAS = "Otras empresas"
_SUBTITULO_OTRAS_EMPRESAS = (
    "Fuera de Abastecedora / ACP Combustibles / Petro Smart · Petroplazas y GC Motors agrupan todas "
    "sus variantes en una sola fila · Filial NO / Filial SI (sn_PagoFilial) lado a lado por empresa"
)


def _construir_tabla_otras_empresas(items: list[dict], dark: bool) -> ft.Control:
    """Una fila POR EMPRESA con Filial NO y Filial SI como columnas propias
    (en vez de 2 filas casi idénticas por empresa en un leaderboard, que se
    veía amontonado). La columna Empresa es más ancha que el resto de las
    secciones (`construir_ranked_list` usa 104px) y admite 2 líneas — los
    nombres largos (ej. 'GC Motors de Occidente') ya no se cortaban bien en
    una sola línea angosta; el tooltip conserva el nombre completo."""
    hay_sin_tc = any((it["no_usd_sin_tc"] or 0) + (it["si_usd_sin_tc"] or 0) for it in items)
    columnas = [
        ft.DataColumn(ft.Text("Empresa", size=11)),
        ft.DataColumn(ft.Text("Filial NO", size=11), numeric=True),
        ft.DataColumn(ft.Text("Filial SI", size=11), numeric=True),
        ft.DataColumn(ft.Text("Total", size=11), numeric=True),
    ]
    if hay_sin_tc:
        columnas.append(ft.DataColumn(ft.Text("USD sin TC", size=11), numeric=True))

    filas = []
    for i, it in enumerate(items):
        no_total = it["no_mxn"] + it["no_usd_convertido"]
        si_total = it["si_mxn"] + it["si_usd_convertido"]
        sin_tc = (it["no_usd_sin_tc"] or 0) + (it["si_usd_sin_tc"] or 0)
        celdas = [
            ft.DataCell(
                ft.Row(
                    [
                        ft.Container(width=8, height=8, bgcolor=color_slot(i, dark), border_radius=4),
                        ft.Container(
                            ft.Text(it["empresa"], size=11, max_lines=2, overflow=ft.TextOverflow.ELLIPSIS,
                                     tooltip=it["empresa"]),
                            width=220,
                        ),
                    ],
                    spacing=6,
                )
            ),
            ft.DataCell(ft.Text(f"${no_total:,.2f}" if no_total else "—", size=11)),
            ft.DataCell(ft.Text(f"${si_total:,.2f}" if si_total else "—", size=11)),
            ft.DataCell(ft.Text(f"${no_total + si_total:,.2f}", size=11, weight=ft.FontWeight.W_600)),
        ]
        if hay_sin_tc:
            celdas.append(ft.DataCell(ft.Text(f"US${sin_tc:,.2f}" if sin_tc else "—", size=11,
                                               color=ft.Colors.RED_600 if sin_tc else ft.Colors.ON_SURFACE_VARIANT)))
        filas.append(ft.DataRow(cells=celdas))

    total_no = sum(it["no_mxn"] + it["no_usd_convertido"] for it in items)
    total_si = sum(it["si_mxn"] + it["si_usd_convertido"] for it in items)
    total_sin_tc = sum((it["no_usd_sin_tc"] or 0) + (it["si_usd_sin_tc"] or 0) for it in items)
    estilo_total = {"size": 11, "weight": ft.FontWeight.W_700, "color": ft.Colors.ON_SURFACE}
    celdas_total = [
        ft.DataCell(ft.Text("Total", **estilo_total)),
        ft.DataCell(ft.Text(f"${total_no:,.2f}", **estilo_total)),
        ft.DataCell(ft.Text(f"${total_si:,.2f}", **estilo_total)),
        ft.DataCell(ft.Text(f"${total_no + total_si:,.2f}", **estilo_total)),
    ]
    if hay_sin_tc:
        celdas_total.append(
            ft.DataCell(ft.Text(f"US${total_sin_tc:,.2f}" if total_sin_tc else "—", **estilo_total))
        )
    filas.append(ft.DataRow(cells=celdas_total, color=ft.Colors.with_opacity(0.08, color_slot(0, dark))))

    tabla = ft.DataTable(columns=columnas, rows=filas, data_row_max_height=48, heading_row_height=32,
                         column_spacing=16)
    return ft.Column([tabla], scroll=ft.ScrollMode.AUTO, height=260)


def _construir_seccion_otras_empresas(resultado, dark: bool) -> ft.Container:
    """Card de 'Otras empresas': siempre en tabla (ver `_construir_tabla_otras_empresas`)
    sin importar el toggle gráfica/tabla del resto de secciones — con 2
    métricas por empresa (Filial NO/SI) una tabla se lee mejor que un
    leaderboard de una sola barra."""
    chips: list[ft.Control] = []
    if isinstance(resultado, Exception):
        cuerpo: ft.Control = ft.Container(
            content=ft.Text(f"No se pudo consultar: {resultado}", size=11, color=ft.Colors.RED_600),
            height=120,
            alignment=ft.Alignment.CENTER,
        )
    elif not resultado:
        cuerpo = estado_vacio()
    else:
        total = sum(it["no_mxn"] + it["no_usd_convertido"] + it["si_mxn"] + it["si_usd_convertido"]
                    for it in resultado)
        chips.append(chip_total(total))
        cuerpo = _construir_tabla_otras_empresas(resultado, dark)

    cabecera = encabezado_seccion(
        ft.Icons.APARTMENT_OUTLINED, color_slot(5, dark), _TITULO_OTRAS_EMPRESAS, _SUBTITULO_OTRAS_EMPRESAS, chips
    )
    return ft.Container(
        content=ft.Column([cabecera, ft.Divider(height=1), cuerpo], spacing=10),
        padding=16,
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
        border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
        border_radius=12,
        shadow=sombra_tarjeta(),
        col={"xs": 12, "lg": 6},
    )


def construir_subtab_segmentado(page: ft.Page) -> ft.Control:
    """Contenido de la sub-pestaña 'Segmentado'. Se construye UNA sola vez al
    armar la pestaña (dispara sus consultas al inicio); sus refrescos internos
    reemplazan `.controls` de los contenedores, igual que siempre."""
    hoy = date.today()
    primer_dia_mes = hoy.replace(day=1)
    rango_sel: list[tuple[date, date]] = [(primer_dia_mes, hoy)]
    resultados_actuales: list[list] = [[[] for _ in _SECCIONES] + [0, []]]
    en_tabla = [False]  # False = gráfica, True = tabla, aplica a todas las secciones a la vez

    def _dark() -> bool:
        return page.theme_mode == ft.ThemeMode.DARK

    def _texto_rango(inicio: date, fin: date) -> str:
        return f"{inicio.strftime('%d %b %Y')} – {fin.strftime('%d %b %Y')}"

    titulo = ft.Text("Ingresos Diversos Mensuales", size=20, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE)
    subtitulo = ft.Text(
        "Todas las vistas usan el mismo rango de fechas seleccionado abajo.",
        size=12,
        color=ft.Colors.ON_SURFACE_VARIANT,
    )

    estado_text = ft.Text("", size=12, color=ft.Colors.RED_600)
    progress = ft.ProgressRing(width=16, height=16, visible=False, stroke_width=2)

    # En Flet 0.85 el FilePicker es un servicio: se crea y se usa directamente
    # (NO se agrega a page.overlay; hacerlo provoca "Unknown control: FilePicker").
    file_picker = ft.FilePicker()

    # Banda superior de KPIs (hero) y grid de secciones. ResponsiveRow reparte el
    # ancho disponible en columnas (col por hijo) → los componentes se
    # redimensionan para ocupar TODO el ancho de la ventana.
    hero_contenedor = ft.ResponsiveRow(spacing=16, run_spacing=16)
    secciones_contenedor = ft.ResponsiveRow(
        spacing=16,
        run_spacing=16,
        vertical_alignment=ft.CrossAxisAlignment.START,
    )
    cuerpo = ft.Column([hero_contenedor, secciones_contenedor], spacing=20, opacity=1.0, animate_opacity=200)

    def _total_seguro(resultado) -> float:
        """MXN. `resultado` es una lista de (etiqueta, mxn, usd, usd_convertido,
        usd_sin_tc) por categoría — nunca se mezcla con el USD."""
        if isinstance(resultado, Exception) or not resultado:
            return 0
        return sum(mxn for _et, mxn, _usd, _conv, _sin_tc in resultado)

    def _usd_seguro(resultado) -> float:
        if isinstance(resultado, Exception) or not resultado:
            return 0
        return sum(usd for _et, _mxn, usd, _conv, _sin_tc in resultado)

    def _refrescar_todo() -> None:
        dark = _dark()
        *resultados_secciones, total_no_identificado, resultado_otras = resultados_actuales[0]
        if isinstance(total_no_identificado, tuple):
            total_no_id_mxn, total_no_id_usd = total_no_identificado
        else:
            total_no_id_mxn, total_no_id_usd = 0, 0

        # Banda hero: los grandes indicadores del periodo, en MXN únicamente.
        # No hay una tarjeta "USD total" aquí porque las secciones NO son
        # conjuntos disjuntos (ej. "Sin identificar" se cruza con "Ingresos
        # identificados") — sumar su USD daría un número que cuenta filas dos
        # veces. El USD de cada sección se muestra en su propia pastilla,
        # junto a su tarjeta (ver _construir_seccion), donde sí es correcto.
        res_empresa = resultados_secciones[0] if resultados_secciones else []
        res_gas = resultados_secciones[3] if len(resultados_secciones) > 3 else []
        subtexto_no_id = "sn_Identificada = NO en el rango"
        if total_no_id_usd:
            subtexto_no_id += f" · USD ${total_no_id_usd:,.2f} aparte"
        hero_contenedor.controls = [
            hero_tile("Ingresos identificados", _total_seguro(res_empresa), color_slot(0, dark),
                      ft.Icons.ACCOUNT_BALANCE_WALLET_OUTLINED,
                      "Asociados y Distribuidora (segmento principal)"),
            hero_tile("Gaseras", _total_seguro(res_gas), color_slot(1, dark),
                      ft.Icons.LOCAL_GAS_STATION_OUTLINED, "Segmento GasPetroil"),
            hero_tile("Sin identificar", total_no_id_mxn, "#e34948",
                      ft.Icons.HELP_OUTLINE, subtexto_no_id),
        ]

        secciones_contenedor.controls = [
            _construir_seccion(titulo_s, subtitulo_s, resultado, dark, en_tabla[0], vista,
                                icono, color_slot(i, dark))
            for i, ((titulo_s, subtitulo_s, _consulta, vista, icono), resultado)
            in enumerate(zip(_SECCIONES, resultados_secciones))
        ] + [_construir_seccion_otras_empresas(resultado_otras, dark)]

    async def cargar(_e=None) -> None:
        cuerpo.opacity = 0.5  # mantiene el render anterior visible (sin salto de layout) mientras carga
        progress.visible = True
        estado_text.value = ""
        boton_rango.disabled = True
        page.update()

        fecha_inicio, fecha_fin = rango_sel[0]
        resultados = await asyncio.gather(
            *(consulta(fecha_inicio, fecha_fin) for _titulo, _subtitulo, consulta, _vista, _icono in _SECCIONES),
            _consultar_no_identificado(fecha_inicio, fecha_fin),
            _consultar_otras_empresas(fecha_inicio, fecha_fin),
            return_exceptions=True,
        )
        resultados_actuales[0] = resultados
        _refrescar_todo()

        progress.visible = False
        boton_rango.disabled = False
        boton_exportar.disabled = False
        cuerpo.opacity = 1.0
        if any(isinstance(r, Exception) for r in resultados):
            estado_text.value = "Algunas secciones no se pudieron consultar (ver detalle en cada tarjeta)."
        page.update()

    async def exportar_excel(_e) -> None:
        """Descarga un Excel con una hoja 'Resumen' (KPIs del periodo) + una
        hoja por sección (Empresa, Tipo de negocio, Sucursal, Sucursal
        Gas, SF, Otras empresas), con los mismos filtros que ya aplica
        esta vista (Asociados/Distribuidora, excluye pagos entre filiales,
        GAS/Autotanque/sin sucursal — ver botón ⓘ)."""
        boton_exportar.disabled = True
        page.update()

        fecha_inicio, fecha_fin = rango_sel[0]
        *resultados_secciones, total_no_identificado, resultado_otras = resultados_actuales[0]
        if isinstance(total_no_identificado, tuple):
            total_no_id_mxn, total_no_id_usd = total_no_identificado
        else:
            total_no_id_mxn, total_no_id_usd = 0, 0

        import openpyxl

        wb = openpyxl.Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        res_empresa = resultados_secciones[0] if resultados_secciones else []
        res_gas = resultados_secciones[3] if len(resultados_secciones) > 3 else []
        escribir_hoja_excel(
            ws_resumen,
            ["Indicador", "Total MXN", "Total USD"],
            [
                ["Periodo", f"{fecha_inicio.strftime('%d/%m/%Y')} – {fecha_fin.strftime('%d/%m/%Y')}", ""],
                ["Ingresos identificados", round(_total_seguro(res_empresa), 2), round(_usd_seguro(res_empresa), 2)],
                ["Gaseras", round(_total_seguro(res_gas), 2), round(_usd_seguro(res_gas), 2)],
                ["Sin identificar", round(total_no_id_mxn, 2), round(total_no_id_usd, 2)],
            ],
        )

        usados: set = {"Resumen"}
        for (titulo_s, _subtitulo_s, _consulta, _vista, _icono), resultado in zip(_SECCIONES, resultados_secciones):
            ws = wb.create_sheet(nombre_hoja_valido(titulo_s, usados))
            if isinstance(resultado, Exception):
                escribir_hoja_excel(ws, ["Error"], [[f"No se pudo consultar: {resultado}"]])
                continue
            items = resultado
            filas_hoja = [
                [et, round(mxn, 2), round(usd, 2), round(conv, 2), round(mxn + conv, 2), round(sin_tc, 2)]
                for et, mxn, usd, conv, sin_tc in items
            ]
            escribir_hoja_excel(
                ws,
                ["Categoría", "MXN (sin USD)", "USD", "MXN convertido", "Total final", "USD sin tipo de cambio"],
                filas_hoja,
            )

        ws_otras = wb.create_sheet(nombre_hoja_valido(_TITULO_OTRAS_EMPRESAS, usados))
        if isinstance(resultado_otras, Exception):
            escribir_hoja_excel(ws_otras, ["Error"], [[f"No se pudo consultar: {resultado_otras}"]])
        else:
            filas_otras = [
                [
                    it["empresa"],
                    round(it["no_mxn"] + it["no_usd_convertido"], 2),
                    round(it["si_mxn"] + it["si_usd_convertido"], 2),
                    round(it["no_mxn"] + it["no_usd_convertido"] + it["si_mxn"] + it["si_usd_convertido"], 2),
                    round((it["no_usd_sin_tc"] or 0) + (it["si_usd_sin_tc"] or 0), 2),
                ]
                for it in resultado_otras
            ]
            escribir_hoja_excel(
                ws_otras,
                ["Empresa", "Filial NO", "Filial SI", "Total", "USD sin tipo de cambio"],
                filas_otras,
            )

        nombre_def = f"dashboard_segmentado_{fecha_inicio:%Y%m%d}_{fecha_fin:%Y%m%d}.xlsx"
        ok, mensaje = await guardar_workbook(page, file_picker, wb, nombre_def)
        boton_exportar.disabled = False
        estado_text.value = mensaje
        page.update()

    def on_cambiar_rango(e) -> None:
        picker = e.control
        if not picker.start_value or not picker.end_value:
            return
        inicio = picker.start_value.date()
        fin = picker.end_value.date()
        rango_sel[0] = (inicio, fin)
        boton_rango.content.controls[1].value = _texto_rango(inicio, fin)
        page.update()
        page.run_task(cargar)

    date_range_picker = ft.DateRangePicker(
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2035, 12, 31),
        start_value=datetime.combine(primer_dia_mes, datetime.min.time()),
        end_value=datetime.combine(hoy, datetime.min.time()),
        entry_mode=ft.DatePickerEntryMode.INPUT,  # diálogo compacto de texto, no el calendario grande
        on_change=on_cambiar_rango,
    )

    # Botón-chip discreto: icono + rango actual, en vez de un botón + texto suelto.
    boton_rango = ft.OutlinedButton(
        content=ft.Row(
            [ft.Icon(ft.Icons.DATE_RANGE, size=16), ft.Text(_texto_rango(primer_dia_mes, hoy), size=13)],
            spacing=8,
            tight=True,
        ),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        on_click=lambda _e: page.show_dialog(date_range_picker),
    )

    def _alternar_vista_render(_e) -> None:
        en_tabla[0] = not en_tabla[0]
        boton_vista.icon = ft.Icons.BAR_CHART_OUTLINED if en_tabla[0] else ft.Icons.TABLE_CHART_OUTLINED
        boton_vista.tooltip = "Ver todo como gráfica" if en_tabla[0] else "Ver todo como tabla"
        _refrescar_todo()
        page.update()

    boton_vista = ft.IconButton(
        icon=ft.Icons.TABLE_CHART_OUTLINED,
        icon_size=18,
        tooltip="Ver todo como tabla",
        on_click=_alternar_vista_render,
    )

    def _abrir_info(_e) -> None:
        lineas = [
            "Ingresos identificados, Tipo de negocio y Sucursal: solo cuentan "
            "movimientos de Asociados y Distribuidora en las 3 empresas "
            "principales (Abastecedora, ACP Combustibles y Petro Smart). No "
            "incluyen pagos entre filiales ni movimientos de sucursales de GAS, "
            "Autotanque, Corporativo o sin sucursal asignada.",
            "Gaseras (Sucursal Gasolineras): mismas 3 empresas principales, pero "
            "solo el segmento GasPetroil. Aquí sí se incluyen las sucursales de "
            "GAS y Autotanque, porque son precisamente el objeto de esta vista "
            "(solo se excluyen las filas sin ninguna sucursal asignada). "
            "Tampoco incluye pagos entre filiales.",
            "SF: mismas 3 empresas principales, pero solo el segmento SF (cuentas "
            "bancarias de_CuentaBancaria = 'Abastecedora SF /AENE' o 'Petroplazas SF', "
            "reclasificadas como tipo de negocio 'SF'). Excluye pagos entre filiales, pero a "
            "diferencia de Sucursal e Ingresos identificados, aquí SÍ se incluyen "
            "todas las sucursales (no se excluyen GAS, Autotanque ni Corporativo).",
            "Otras empresas: todo lo que NO sea Abastecedora, ACP Combustibles ni "
            "Petro Smart, sin filtro de tipo de negocio ni de sucursal. Cualquier "
            "nb_Empresa que contenga 'Petroplazas' se agrupa como una sola 'Petroplazas' "
            "(junta todas sus variantes), y lo mismo para 'GC Motors' -> 'GC Motors de "
            "Occidente'; el resto conserva su nombre de empresa tal cual. A diferencia de "
            "las demás secciones, esta siempre se muestra en tabla (no tiene versión "
            "gráfica) con una sola fila por empresa y sus columnas 'Filial NO' / 'Filial "
            "SI' (sn_PagoFilial) una junto a la otra.",
            "Sin identificar: suma de todos los movimientos marcados como no "
            "identificados en el periodo, sin importar empresa, sucursal o tipo "
            "de negocio.",
            "El tipo de negocio se reclasifica antes de agrupar, en este orden: "
            "de_CuentaBancaria = 'Abastecedora SF /AENE' o 'Petroplazas SF' se cuenta "
            "como 'SF'; los "
            "clientes de 'Público en general' de Petro Smart se cuentan como "
            "'GasPetroil'; y un cliente específico (id 4359) se cuenta como "
            "'Distribuidora' — sin importar cómo esté registrado originalmente.",
            "En TODAS las vistas y consultas de esta pestaña (Segmentado, Timeline "
            "y Detalle) se descartan por completo los movimientos cuya "
            "de_CuentaBancaria sea 'Gastos No Deducibles' o 'Petroplazas "
            "Monederos' — son cuentas de control interno, no ingresos reales.",
            "En TODAS las vistas de esta pestaña, los movimientos en dólares "
            "(Moneda = 'Dolar (USD)') se separan de los montos en pesos: nunca se "
            "convierten ni se suman juntos en la gráfica ni en la pastilla. Cada "
            "tarjeta con USD muestra una pastilla 'USD' aparte de su 'Total' en "
            "pesos; en Detalle, la moneda de cada movimiento es una columna más "
            "de la tabla.",
            "En modo tabla, las 5 secciones con toggle gráfica/tabla (Empresa, Tipo de "
            "negocio, Sucursal, Sucursal Gaseras, SF) sí muestran el desglose "
            "completo por categoría: MXN (sin USD), USD, MXN convertido y Total "
            "final. El tipo de cambio sale de Tableros.DocumentosClientesCobranza "
            "(columna im_TipoCambio), promediado por día (fh_Deposito_Mostrar). Se "
            "busca primero el tipo de cambio del día EXACTO de cada movimiento en "
            "USD; si ese día no tiene registro, se usa el del día MÁS CERCANO que "
            "sí tenga (antes o después, el de menor diferencia en días) — nunca se "
            "inventa un valor. La columna 'USD sin tipo de cambio' (en rojo) solo "
            "muestra saldo si Cobranza no tiene NINGÚN tipo de cambio registrado en "
            "toda su historia. 'Total final' = MXN (sin USD) + MXN convertido; el "
            "USD sin tipo de cambio queda fuera de esa suma porque no hay con qué "
            "convertirlo.",
            "A diferencia de las sub-pestañas 'Timeline' y 'Detalle', aquí sí se "
            "excluyen por defecto GAS, Autotanque, sucursal sin asignar y pagos "
            "entre filiales — por eso los totales no son directamente "
            "comparables con los de esas vistas.",
        ]
        dialogo = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cómo se calculan estos datos"),
            content=ft.Container(
                content=ft.Column(
                    [ft.Text(f"•  {l}", size=12, selectable=True) for l in lineas],
                    spacing=10, scroll=ft.ScrollMode.AUTO,
                ),
                width=520, height=340,
            ),
            actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
        )
        mostrar_dialogo(page, dialogo)

    boton_info = ft.IconButton(
        icon=ft.Icons.INFO_OUTLINE,
        icon_size=18,
        tooltip="Ver cómo se calculan estos datos",
        on_click=_abrir_info,
    )

    boton_exportar = ft.IconButton(
        icon=ft.Icons.DOWNLOAD,
        icon_size=18,
        tooltip="Descargar Excel (resumen + una hoja por sección)",
        disabled=True,
        on_click=lambda e: page.run_task(exportar_excel, e),
    )

    barra_herramientas = ft.Container(
        content=ft.Row(
            [boton_rango, progress, estado_text, ft.Container(expand=True), boton_exportar, boton_info, boton_vista],
            spacing=12,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        padding=ft.Padding(left=14, right=14, top=10, bottom=10),
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOW,
        border_radius=12,
    )

    contenido = ft.Container(
        content=ft.Column(
            [
                ft.Column([titulo, subtitulo], spacing=2),
                barra_herramientas,
                ft.Container(content=cuerpo, expand=True),
            ],
            spacing=16,
            scroll=ft.ScrollMode.AUTO,
            expand=True,
        ),
        padding=20,
        expand=True,
    )

    page.run_task(cargar)

    return contenido
