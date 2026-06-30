"""Parsea el reporte "Estado de Cuenta" de SIPP (.xlsx) y, a partir del cliente,
el monto del abono y la EMPRESA seleccionada, SUGIERE la sucursal a la que aplica.

El reporte es jerárquico y mezcla varias empresas del grupo:
  - fila encabezado de cliente: col A = nombre del cliente, col B vacía.
  - filas de detalle (una por factura pendiente): col A = EMPRESA (Abastecedora /
    Petro Smart / ACP Combustibles / ...), col B = Sucursal, col C = Folio,
    col P (índice 15) = Saldo Pendiente. Detectar detalle por col B con valor.
  - filas "Total Sucursal/Empresa/Cliente" (col A empieza con "Total") → ignorar.

Como una empresa solo opera en sus propias sucursales, todo se filtra por la
empresa seleccionada en la app. El monto NO es llave dura (un abono puede exceder
una factura y el resto irse como anticipo), así que la sucursal es una SUGERENCIA
editable.
"""

from dataclasses import dataclass, field
from typing import Optional

from .textutils import normalizar

_COL_EMPRESA = 0      # A
_COL_SUCURSAL = 1     # B
_COL_FOLIO = 2        # C
_COL_SALDO = 15       # P (Saldo Pendiente)

_TOLERANCIA = 1.0     # pesos


def _emp_key(empresa: str) -> str:
    return (empresa or "").strip().lower()


@dataclass
class EstadoCuenta:
    # cliente_norm -> { empresa_key -> { sucursal -> [(folio, saldo), ...] } }
    por_cliente: dict[str, dict[str, dict[str, list[tuple[str, float]]]]] = field(default_factory=dict)
    nombre_original: dict[str, str] = field(default_factory=dict)

    @property
    def num_clientes(self) -> int:
        return len(self.por_cliente)

    def sucursales(self, empresa_reporte: str) -> list[str]:
        """Sucursales distintas de una empresa en todo el reporte (ordenadas)."""
        ek = _emp_key(empresa_reporte)
        sucs: set[str] = set()
        for por_emp in self.por_cliente.values():
            sucs.update(por_emp.get(ek, {}).keys())
        return sorted(sucs)

    def sucursales_de_cliente(self, cliente: str, empresa_reporte: str) -> list[str]:
        clave = _resolver_cliente(self, cliente)
        if clave is None:
            return []
        return list(self.por_cliente[clave].get(_emp_key(empresa_reporte), {}).keys())


def cargar_estado_cuenta(ruta: str) -> EstadoCuenta:
    """Carga el .xlsx a un índice cliente→empresa→sucursal→facturas."""
    import openpyxl

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    estado = EstadoCuenta()
    cliente_norm: Optional[str] = None
    try:
        for fila in ws.iter_rows(min_row=6, values_only=True):
            col_a = fila[_COL_EMPRESA]
            sucursal = fila[_COL_SUCURSAL]
            if isinstance(col_a, str) and col_a.strip().startswith("Total"):
                continue
            if sucursal not in (None, ""):
                # Fila de detalle: col A = empresa, col B = sucursal.
                if cliente_norm is None:
                    continue
                empresa = _emp_key(str(col_a))
                suc = str(sucursal).strip()
                saldo = fila[_COL_SALDO]
                try:
                    saldo_f = float(saldo) if saldo not in (None, "") else 0.0
                except (TypeError, ValueError):
                    saldo_f = 0.0
                folio = str(fila[_COL_FOLIO] or "").strip()
                estado.por_cliente[cliente_norm].setdefault(empresa, {}).setdefault(
                    suc, []
                ).append((folio, saldo_f))
            elif isinstance(col_a, str) and col_a.strip():
                nombre = col_a.strip()
                cliente_norm = normalizar(nombre)
                estado.por_cliente.setdefault(cliente_norm, {})
                estado.nombre_original.setdefault(cliente_norm, nombre)
    finally:
        wb.close()

    estado.por_cliente = {k: v for k, v in estado.por_cliente.items() if v}
    return estado


def _resolver_cliente(estado: EstadoCuenta, cliente: str) -> Optional[str]:
    """Clave del cliente: exacto normalizado, o por contención única."""
    if not cliente:
        return None
    objetivo = normalizar(cliente)
    if objetivo in estado.por_cliente:
        return objetivo
    candidatos = [k for k in estado.por_cliente if k and (k in objetivo or objetivo in k)]
    if len(candidatos) == 1:
        return candidatos[0]
    return None


def _existe_subconjunto(saldos: list[float], objetivo: float, tol: float, max_n: int = 3) -> bool:
    from itertools import combinations

    for n in range(1, min(max_n, len(saldos)) + 1):
        for combo in combinations(saldos, n):
            if abs(sum(combo) - objetivo) <= tol:
                return True
    return False


def sugerir_sucursal(
    estado: EstadoCuenta, cliente: str, abono: Optional[float], empresa_reporte: str
) -> Optional[tuple[str, str]]:
    """(sucursal, motivo) sugerida para (cliente, abono) dentro de la empresa, o
    None si el cliente no tiene facturas de esa empresa en el reporte."""
    clave = _resolver_cliente(estado, cliente)
    if clave is None:
        return None
    sucursales = estado.por_cliente[clave].get(_emp_key(empresa_reporte), {})
    if not sucursales:
        return None

    if len(sucursales) == 1:
        return (next(iter(sucursales)), "única")

    if abono is None:
        return None

    for suc, facturas in sucursales.items():
        if any(abs(saldo - abono) <= _TOLERANCIA for _folio, saldo in facturas):
            return (suc, "factura exacta")

    for suc, facturas in sucursales.items():
        saldos = [saldo for _folio, saldo in facturas]
        if _existe_subconjunto(saldos, abono, _TOLERANCIA):
            return (suc, "suma de facturas")

    mejor_suc = None
    mejor_dif = None
    for suc, facturas in sucursales.items():
        for _folio, saldo in facturas:
            dif = abs(saldo - abono)
            if mejor_dif is None or dif < mejor_dif:
                mejor_dif = dif
                mejor_suc = suc
    if mejor_suc is not None:
        return (mejor_suc, "aproximado")
    return None


def sugerir_sucursal_detalle(
    estado: EstadoCuenta, cliente: str, abono: Optional[float], empresa_reporte: str
) -> Optional[tuple[Optional[str], Optional[str], list[str]]]:
    """Para el grid: (sucursal_sugerida, motivo, [sucursales del cliente en la
    empresa]) o None si el cliente no está en el reporte para esa empresa."""
    clave = _resolver_cliente(estado, cliente)
    if clave is None:
        return None
    todas = list(estado.por_cliente[clave].get(_emp_key(empresa_reporte), {}).keys())
    if not todas:
        return None
    sug = sugerir_sucursal(estado, cliente, abono, empresa_reporte)
    if sug is None:
        return (None, None, todas)
    return (sug[0], sug[1], todas)
