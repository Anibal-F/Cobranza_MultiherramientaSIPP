"""Sub-pestaña 'Detalle' de Cumplimiento de Cobro: los mismos registros que
'Resumen' (mismo filtro de fecha y mismos filtros de calidad de datos,
heredados de `CumplimientoRepository.detalle_periodo`), fila por fila, con
paginación y filtros multi-selección (dropdown con checkboxes) por columna
— Cliente, Sucursal, Empresa, Tipo de negocio y Filial.

Los catálogos de cada filtro se arman con los valores DEL PERIODO CARGADO
(no un catálogo global de toda la tabla): cambiar el rango de fechas
recarga los datos y reinicia los 5 filtros a "todos", porque un valor
seleccionado podría no existir ya en el nuevo periodo.

Los montos son im_Cartera ("Cartera esperada") e im_CarteraVencida
("Pendiente hoy") — NO im_CarteraVigente, que es una foto del día de hoy y
sale en 0 para cualquier fila cuyo fh_Vencimiento ya pasó (ver
app/services/cumplimiento_repository.py)."""

import asyncio
from datetime import date, datetime, timedelta

import flet as ft
from flet_datatable2 import DataColumn2, DataColumnSize, DataTable2

from ..dashboard.componentes import (
    escribir_hoja_excel,
    estado_vacio,
    guardar_workbook,
    mostrar_dialogo,
    sombra_tarjeta,
)
from ..services.cumplimiento_repository import CumplimientoRepository

# El repositorio se crea perezosamente (necesita credenciales de BigQuery); así
# no falla al importar este módulo si aún no hay credenciales configuradas.
_repo_holder: list[CumplimientoRepository | None] = [None]


def _repo() -> CumplimientoRepository:
    if _repo_holder[0] is None:
        _repo_holder[0] = CumplimientoRepository()
    return _repo_holder[0]

_SIN_ASIGNAR = "Sin asignar"
_TAMANO_PAGINA = 50

# (campo de la fila, etiqueta del filtro/columna, ícono del botón multiselect)
_FILTROS_COLUMNA = [
    ("nb_Cliente", "Cliente", ft.Icons.PERSON_OUTLINE),
    ("nb_Sucursal", "Sucursal", ft.Icons.STORE_OUTLINED),
    ("nb_Empresa", "Empresa", ft.Icons.BUSINESS_OUTLINED),
    ("nb_TipoDeNegocio", "Tipo de negocio", ft.Icons.CATEGORY_OUTLINED),
    ("sn_filial", "Filial", ft.Icons.SYNC_ALT_OUTLINED),
]


def _valor_columna(fila: dict, campo: str) -> str:
    return (fila.get(campo) or "").strip() or _SIN_ASIGNAR


def _fmt_fecha(valor) -> str:
    if valor is None:
        return "—"
    if hasattr(valor, "strftime"):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def construir_subtab_detalle(page: ft.Page) -> ft.Control:
    """Contenido de la sub-pestaña 'Detalle'. Se construye UNA sola vez al
    armar la pestaña (dispara la consulta al inicio); filtros y paginación se
    resuelven 100% en memoria sobre el dataset del periodo ya cargado."""
    hoy = date.today()
    hace_una_semana = hoy - timedelta(days=7)

    estado = {
        "rango": (hace_una_semana, hoy),
        "filtros": {campo: [] for campo, _et, _ic in _FILTROS_COLUMNA},  # [] == todos
        "catalogos": {campo: [] for campo, _et, _ic in _FILTROS_COLUMNA},
        "pagina": 0,
        "filas": [],
    }

    def _dark() -> bool:
        return page.theme_mode == ft.ThemeMode.DARK

    def _texto_rango(inicio: date, fin: date) -> str:
        return f"{inicio.strftime('%d %b %Y')} – {fin.strftime('%d %b %Y')}"

    titulo = ft.Text("Detalle de facturas", size=20, weight=ft.FontWeight.W_600, color=ft.Colors.ON_SURFACE)
    subtitulo = ft.Text(
        "Mismos registros que Resumen, fila por fila. Los filtros de Cliente, Sucursal, Empresa, "
        "Tipo de negocio y Filial se aplican sobre el periodo ya cargado.",
        size=12,
        color=ft.Colors.ON_SURFACE_VARIANT,
    )

    estado_text = ft.Text("", size=12, color=ft.Colors.RED_600)
    progress = ft.ProgressRing(width=16, height=16, visible=False, stroke_width=2)

    # En Flet 0.85 el FilePicker es un servicio: se crea y se usa directamente
    # (NO se agrega a page.overlay; hacerlo provoca "Unknown control: FilePicker").
    file_picker = ft.FilePicker()

    tabla = DataTable2(
        rows=[],
        fixed_top_rows=1, column_spacing=16, min_width=1100, expand=True,
        heading_row_height=36, data_row_height=32,
        empty=estado_vacio(),
        columns=[
            DataColumn2(ft.Text("Cliente", weight=ft.FontWeight.BOLD, size=12), size=DataColumnSize.L),
            DataColumn2(ft.Text("Sucursal", weight=ft.FontWeight.BOLD, size=12), fixed_width=150),
            DataColumn2(ft.Text("Folio", weight=ft.FontWeight.BOLD, size=12), fixed_width=110),
            DataColumn2(ft.Text("Vencimiento", weight=ft.FontWeight.BOLD, size=12), fixed_width=100),
            DataColumn2(ft.Text("Cartera esperada", weight=ft.FontWeight.BOLD, size=12), numeric=True, fixed_width=130),
            DataColumn2(ft.Text("Pendiente hoy", weight=ft.FontWeight.BOLD, size=12), numeric=True, fixed_width=120),
            DataColumn2(ft.Text("Empresa", weight=ft.FontWeight.BOLD, size=12), fixed_width=170),
            DataColumn2(ft.Text("Tipo de negocio", weight=ft.FontWeight.BOLD, size=12), fixed_width=140),
            DataColumn2(ft.Text("Filial", weight=ft.FontWeight.BOLD, size=12), fixed_width=80),
        ],
    )
    texto_totales = ft.Text("", size=11, color=ft.Colors.ON_SURFACE_VARIANT)
    texto_pagina = ft.Text("", size=12, color=ft.Colors.ON_SURFACE)
    boton_pagina_ant = ft.IconButton(icon=ft.Icons.CHEVRON_LEFT, icon_size=20, tooltip="Página anterior")
    boton_pagina_sig = ft.IconButton(icon=ft.Icons.CHEVRON_RIGHT, icon_size=20, tooltip="Página siguiente")
    fila_paginacion = ft.Row(
        [boton_pagina_ant, texto_pagina, boton_pagina_sig],
        spacing=4, alignment=ft.MainAxisAlignment.CENTER, visible=False,
    )

    def _filas_filtradas() -> list[dict]:
        filas = estado["filas"]
        for campo, seleccion in estado["filtros"].items():
            if seleccion:
                filas = [f for f in filas if _valor_columna(f, campo) in seleccion]
        return filas

    def _fila_tabla(f: dict) -> ft.DataRow:
        return ft.DataRow(cells=[
            ft.DataCell(ft.Text(f.get("nb_Cliente") or "—", size=11)),
            ft.DataCell(ft.Text(f.get("nb_Sucursal") or "—", size=11)),
            ft.DataCell(ft.Text(f.get("fl_FolioDocumento") or "—", size=11)),
            ft.DataCell(ft.Text(_fmt_fecha(f.get("fh_Vencimiento")), size=11)),
            ft.DataCell(ft.Text(f"${(f.get('im_Cartera') or 0):,.2f}", size=11)),
            ft.DataCell(ft.Text(f"${(f.get('im_CarteraVencida') or 0):,.2f}", size=11)),
            ft.DataCell(ft.Text(f.get("nb_Empresa") or "—", size=11)),
            ft.DataCell(ft.Text(f.get("nb_TipoDeNegocio") or "—", size=11)),
            ft.DataCell(ft.Text(f.get("sn_filial") or "—", size=11)),
        ])

    def _aplicar_estado_tabla() -> None:
        filtradas = _filas_filtradas()
        total = len(filtradas)
        num_paginas = max(1, -(-total // _TAMANO_PAGINA))  # ceil
        estado["pagina"] = min(estado["pagina"], num_paginas - 1)
        inicio = estado["pagina"] * _TAMANO_PAGINA
        pagina_filas = filtradas[inicio:inicio + _TAMANO_PAGINA]

        tabla.rows = [_fila_tabla(f) for f in pagina_filas]
        total_esperado = sum(f.get("im_Cartera") or 0 for f in filtradas)
        total_vencido = sum(f.get("im_CarteraVencida") or 0 for f in filtradas)
        texto_totales.value = (
            f"{total:,} factura(s) filtrada(s) · Cartera esperada: ${total_esperado:,.2f} "
            f"· Pendiente a hoy: ${total_vencido:,.2f}"
        )

        fila_paginacion.visible = num_paginas > 1
        texto_pagina.value = f"Página {estado['pagina'] + 1} de {num_paginas}"
        boton_pagina_ant.disabled = estado["pagina"] <= 0
        boton_pagina_sig.disabled = estado["pagina"] >= num_paginas - 1

    def _refrescar_catalogos() -> None:
        for campo, _et, _ic in _FILTROS_COLUMNA:
            valores = {_valor_columna(f, campo) for f in estado["filas"]}
            estado["catalogos"][campo] = sorted(valores, key=lambda v: (v == _SIN_ASIGNAR, v))

    def _on_pagina_ant(_e) -> None:
        estado["pagina"] = max(0, estado["pagina"] - 1)
        _aplicar_estado_tabla()
        page.update()

    def _on_pagina_sig(_e) -> None:
        estado["pagina"] += 1
        _aplicar_estado_tabla()
        page.update()

    boton_pagina_ant.on_click = _on_pagina_ant
    boton_pagina_sig.on_click = _on_pagina_sig

    async def cargar(_e=None) -> None:
        progress.visible = True
        estado_text.value = ""
        boton_rango.disabled = True
        page.update()

        fecha_inicio, fecha_fin = estado["rango"]
        try:
            filas = await asyncio.to_thread(_repo().detalle_periodo, fecha_inicio, fecha_fin)
        except Exception as error:  # noqa: BLE001 - se muestra en estado_text, igual que Resumen
            progress.visible = False
            boton_rango.disabled = False
            estado_text.value = f"No se pudo consultar BigQuery: {error}"
            page.update()
            return

        estado["filas"] = filas
        # Un valor seleccionado en el periodo anterior podría no existir en el
        # nuevo rango — se reinician los 5 filtros a "todos" para no dejar un
        # filtro fantasma que oculte todo silenciosamente.
        estado["filtros"] = {campo: [] for campo, _et, _ic in _FILTROS_COLUMNA}
        estado["pagina"] = 0
        _refrescar_catalogos()
        # _refrescar() (definida más abajo) reconstruye los botones multiselect
        # con el catálogo recién cargado Y vuelve a aplicar filtro/paginación —
        # llamarla aquí evita que los botones se queden con el catálogo vacío
        # con el que se armaron en el primer render (antes de esta carga).
        _refrescar()

        progress.visible = False
        boton_rango.disabled = False
        boton_exportar.disabled = not filas
        page.update()

    def on_cambiar_rango(e) -> None:
        picker = e.control
        if not picker.start_value or not picker.end_value:
            return
        inicio = picker.start_value.date()
        fin = picker.end_value.date()
        estado["rango"] = (inicio, fin)
        boton_rango.content.controls[1].value = _texto_rango(inicio, fin)
        page.update()
        page.run_task(cargar)

    date_range_picker = ft.DateRangePicker(
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2035, 12, 31),
        start_value=datetime.combine(hace_una_semana, datetime.min.time()),
        end_value=datetime.combine(hoy, datetime.min.time()),
        entry_mode=ft.DatePickerEntryMode.CALENDAR_ONLY,
        on_change=on_cambiar_rango,
    )

    boton_rango = ft.OutlinedButton(
        content=ft.Row(
            [ft.Icon(ft.Icons.DATE_RANGE, size=16), ft.Text(_texto_rango(hace_una_semana, hoy), size=13)],
            spacing=8,
            tight=True,
        ),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        on_click=lambda _e: page.show_dialog(date_range_picker),
    )

    # --- Filtros multi-selección por columna -----------------------------------
    # Dropdown anclado (PopupMenuButton), NO modal: un PopupMenuItem con un
    # Checkbox propio por valor del catálogo — mismo patrón que el explorador
    # del Dashboard de Ingresos (ver app/dashboard/explorador.py). Un Checkbox
    # anidado en el content de un PopupMenuItem consume su propio tap y NO
    # cierra el menú — solo "Aplicar" (un PopupMenuItem con on_click en el
    # ítem mismo) lo cierra, permitiendo marcar varios checkboxes seguidos.

    def _boton_multiselect(campo: str, etiqueta: str, icono) -> ft.PopupMenuButton:
        seleccion_actual: list[str] = estado["filtros"][campo]
        valores = estado["catalogos"][campo]
        resumen = f"{len(seleccion_actual)} sel." if seleccion_actual else "Todos"

        if not valores:
            items: list[ft.PopupMenuItem] = [
                ft.PopupMenuItem(
                    content=ft.Text("Sin valores en el periodo cargado.", size=12,
                                     color=ft.Colors.ON_SURFACE_VARIANT),
                )
            ]
        else:
            seleccion_previa = set(seleccion_actual)
            pendiente: dict[str, bool] = {v: (v in seleccion_previa) for v in valores}
            checks: list[ft.Checkbox] = []

            def _marcar_todos(_e) -> None:
                for chk in checks:
                    chk.value = True
                    pendiente[chk.label] = True
                page.update()

            def _limpiar(_e) -> None:
                for chk in checks:
                    chk.value = False
                    pendiente[chk.label] = False
                page.update()

            def _toggle(v: str):
                def _h(e) -> None:
                    pendiente[v] = e.control.value
                return _h

            def _aplicar(_e) -> None:
                estado["filtros"][campo] = [v for v, marcado in pendiente.items() if marcado]
                estado["pagina"] = 0
                _refrescar()

            filas_valor = []
            for v in valores:
                chk = ft.Checkbox(value=pendiente[v], label=v, on_change=_toggle(v))
                checks.append(chk)
                filas_valor.append(ft.PopupMenuItem(content=ft.Row([chk])))

            items = [
                ft.PopupMenuItem(
                    content=ft.Row([
                        ft.TextButton("Limpiar", on_click=_limpiar),
                        ft.TextButton("Todos", on_click=_marcar_todos),
                    ]),
                ),
                *filas_valor,
                ft.PopupMenuItem(
                    content=ft.Container(
                        content=ft.Text("Aplicar", weight=ft.FontWeight.BOLD, color=ft.Colors.PRIMARY),
                        alignment=ft.Alignment.CENTER,
                    ),
                    on_click=_aplicar,
                ),
            ]

        return ft.PopupMenuButton(
            content=ft.Container(
                content=ft.Row(
                    [ft.Icon(icono, size=16), ft.Text(etiqueta, size=13),
                     ft.Text(resumen, size=12, color=ft.Colors.ON_SURFACE_VARIANT)],
                    spacing=6, tight=True,
                ),
                padding=ft.Padding(left=12, right=12, top=6, bottom=6),
                border=ft.Border.all(1, ft.Colors.OUTLINE),
                border_radius=8,
            ),
            items=items,
        )

    fila_filtros = ft.Row(spacing=10, wrap=True, vertical_alignment=ft.CrossAxisAlignment.CENTER)

    def _refrescar() -> None:
        """Reconstruye los botones multiselect (sus checks reflejan el estado
        actual) y vuelve a aplicar filtro + paginación a la tabla."""
        fila_filtros.controls = [_boton_multiselect(campo, et, ic) for campo, et, ic in _FILTROS_COLUMNA]
        _aplicar_estado_tabla()
        page.update()

    async def exportar_excel(_e) -> None:
        """Exporta TODAS las filas que cumplen los filtros actuales (no solo
        la página en pantalla)."""
        filtradas = _filas_filtradas()
        if not filtradas:
            return
        boton_exportar.disabled = True
        page.update()

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detalle"
        fecha_inicio, fecha_fin = estado["rango"]
        filtros_txt = " · ".join(
            f"{etiqueta}: {', '.join(estado['filtros'][campo])}"
            for campo, etiqueta, _ic in _FILTROS_COLUMNA if estado["filtros"][campo]
        ) or "sin filtros adicionales"
        ws.cell(row=1, column=1,
                value=f"Periodo (fh_Vencimiento): {fecha_inicio.strftime('%d/%m/%Y')} - "
                      f"{fecha_fin.strftime('%d/%m/%Y')} · {filtros_txt}")

        encabezados = ["nb_Cliente", "nb_Sucursal", "fl_FolioDocumento", "fh_Vencimiento",
                       "im_Cartera", "im_CarteraVencida", "nb_Empresa", "nb_TipoDeNegocio", "sn_filial"]
        escribir_hoja_excel(
            ws, encabezados,
            [[fila.get(col) for col in encabezados] for fila in filtradas],
            fila_inicio=3,
        )
        col_fecha = encabezados.index("fh_Vencimiento") + 1
        for fila_celdas in ws.iter_rows(min_row=4, min_col=col_fecha, max_col=col_fecha):
            for celda in fila_celdas:
                celda.number_format = "dd/mm/yyyy"

        nombre_def = f"cumplimiento_detalle_{fecha_inicio:%Y%m%d}_{fecha_fin:%Y%m%d}.xlsx"
        ok, mensaje = await guardar_workbook(page, file_picker, wb, nombre_def)
        boton_exportar.disabled = False
        estado_text.value = mensaje or f"{len(filtradas)} registro(s) exportados."
        page.update()

    boton_exportar = ft.IconButton(
        icon=ft.Icons.DOWNLOAD,
        icon_size=18,
        tooltip="Descargar Excel de los registros filtrados (todas las páginas)",
        disabled=True,
        on_click=lambda e: page.run_task(exportar_excel, e),
    )

    def _abrir_info(_e) -> None:
        lineas = [
            "Mismo dataset que la sub-pestaña 'Resumen': se filtra por fh_Vencimiento dentro del rango "
            "seleccionado y se excluyen cliente/folio vacíos, cliente 'ICV', filas 'Totales' y folios FCOR.",
            "Los catálogos de Cliente, Sucursal, Empresa, Tipo de negocio y Filial se arman con los "
            "valores presentes en el periodo cargado (no un catálogo global de toda la tabla) — cambiar "
            "el rango de fechas reinicia los 5 filtros a 'todos'.",
            "'Sin asignar' agrupa las filas donde esa columna viene vacía en BigQuery.",
            "Cartera esperada = im_Cartera (saldo total original del documento). Pendiente hoy = "
            "im_CarteraVencida (de ese total, lo que sigue sin cobrarse al momento de consultar). No se "
            "usa im_CarteraVigente porque es una foto del día de hoy que sale en 0 para cualquier "
            "factura cuyo vencimiento ya pasó — el caso normal en este reporte.",
            f"La tabla pagina de {_TAMANO_PAGINA} en {_TAMANO_PAGINA} registros; los totales que se "
            "muestran arriba consideran TODAS las filas filtradas, no solo la página visible.",
            "El Excel exportado incluye todas las filas que cumplen los filtros activos (todas las "
            "páginas), no solo la página en pantalla.",
        ]
        dialogo = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cómo se calculan estos datos"),
            content=ft.Container(
                content=ft.Column(
                    [ft.Text(f"•  {l}", size=12, selectable=True) for l in lineas],
                    spacing=10, scroll=ft.ScrollMode.AUTO,
                ),
                width=520, height=320,
            ),
            actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
        )
        mostrar_dialogo(page, dialogo)

    boton_info = ft.IconButton(
        icon=ft.Icons.INFO_OUTLINE,
        icon_size=18,
        tooltip="Ver cómo se calculan estos datos",
        on_click=_abrir_info,
    )

    barra_herramientas = ft.Container(
        content=ft.Row(
            [boton_rango, progress, estado_text, ft.Container(expand=True), boton_exportar, boton_info],
            spacing=12,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        padding=ft.Padding(left=14, right=14, top=10, bottom=10),
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOW,
        border_radius=12,
    )

    tarjeta_tabla = ft.Container(
        content=ft.Column(
            [
                fila_filtros,
                texto_totales,
                ft.Container(content=tabla, expand=True),
                fila_paginacion,
            ],
            spacing=10,
            expand=True,
        ),
        padding=16,
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
        border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
        border_radius=12,
        shadow=sombra_tarjeta(),
        expand=True,
    )

    contenido = ft.Container(
        content=ft.Column(
            [
                ft.Column([titulo, subtitulo], spacing=2),
                barra_herramientas,
                tarjeta_tabla,
            ],
            spacing=16,
            expand=True,
        ),
        padding=20,
        expand=True,
    )

    _refrescar()
    page.run_task(cargar)

    return contenido
