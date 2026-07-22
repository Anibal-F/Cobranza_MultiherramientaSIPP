"""Sub-pestaña 'Resumen' de Cumplimiento de Cobro: facturas cuya fecha de
vencimiento (fh_Vencimiento) cae en el periodo seleccionado (por default, la
semana pasada), sobre la misma tabla que usa Proyección (RDC) —
`documentosClientes_AntiguedadSaldosVencidoPorClienteDetalle`. KPIs del
periodo, mini dashboards por tipo de negocio y un top 15 de clientes.

Usa im_Cartera (saldo total original del documento) para "lo que se
esperaba cobrar", NO im_CarteraVigente — esa columna es una foto del día de
HOY (la parte del saldo que aún no vence al momento de consultar) y sale en
0 para cualquier fila cuyo fh_Vencimiento ya pasó, que es el caso normal en
este reporte (rango default = semana ANTERIOR). im_CarteraVencida se suma
aparte para comparar lo esperado contra lo que sigue pendiente HOY."""

import asyncio
from datetime import date, datetime, timedelta

import flet as ft

from ..dashboard.componentes import (
    color_slot,
    construir_ranked_list,
    encabezado_seccion,
    escribir_hoja_excel,
    estado_vacio,
    guardar_workbook,
    hero_tile,
    mostrar_dialogo,
    nombre_hoja_valido,
    sombra_tarjeta,
    tile_compacta,
)
from ..services.cumplimiento_repository import COLUMNAS_DETALLE, CumplimientoRepository

# El repositorio se crea perezosamente (necesita credenciales de BigQuery); así
# no falla al importar este módulo si aún no hay credenciales configuradas.
_repo_holder: list[CumplimientoRepository | None] = [None]


def _repo() -> CumplimientoRepository:
    if _repo_holder[0] is None:
        _repo_holder[0] = CumplimientoRepository()
    return _repo_holder[0]

_SIN_ASIGNAR = "Sin asignar"


def _agregar_por_tipo_negocio(filas: list[dict]) -> list[tuple[str, float, float, int]]:
    """(tipo_negocio, suma im_Cartera, suma im_CarteraVencida, folios distintos),
    orden desc por im_Cartera (lo esperado)."""
    por_tipo: dict[str, dict] = {}
    for fila in filas:
        tipo = (fila.get("nb_TipoDeNegocio") or "").strip() or _SIN_ASIGNAR
        entrada = por_tipo.setdefault(tipo, {"esperado": 0.0, "vencido": 0.0, "folios": set()})
        entrada["esperado"] += fila.get("im_Cartera") or 0
        entrada["vencido"] += fila.get("im_CarteraVencida") or 0
        folio = fila.get("fl_FolioDocumento")
        if folio:
            entrada["folios"].add(folio)
    items = [
        (tipo, datos["esperado"], datos["vencido"], len(datos["folios"]))
        for tipo, datos in por_tipo.items()
    ]
    items.sort(key=lambda it: it[1], reverse=True)
    return items


def _top_clientes(filas: list[dict], top: int = 15) -> list[tuple[str, float]]:
    """(cliente, suma im_Cartera), orden desc, recortado a `top`."""
    por_cliente: dict[str, float] = {}
    for fila in filas:
        cliente = (fila.get("nb_Cliente") or "").strip() or _SIN_ASIGNAR
        por_cliente[cliente] = por_cliente.get(cliente, 0.0) + (fila.get("im_Cartera") or 0)
    items = sorted(por_cliente.items(), key=lambda it: it[1], reverse=True)
    return items[:top]


def _construir_workbook_reporte(
    filas: list[dict],
    fecha_inicio: date,
    fecha_fin: date,
    total_facturas: int,
    total_esperado: float,
    total_vencido: float,
    por_tipo: list[tuple[str, float, float, int]],
    top_clientes: list[tuple[str, float]],
):
    """Excel de descarga con 4 hojas: Resumen (KPIs), Tipo de negocio, Top
    clientes y el Detalle crudo del periodo con las columnas que consume la UI."""
    import openpyxl

    wb = openpyxl.Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    escribir_hoja_excel(
        ws_resumen,
        ["Indicador", "Valor"],
        [
            ["Periodo (fh_Vencimiento)", f"{fecha_inicio.strftime('%d/%m/%Y')} – {fecha_fin.strftime('%d/%m/%Y')}"],
            ["Total de facturas del periodo", total_facturas],
            ["Total esperado a cobrar (im_Cartera)", round(total_esperado, 2)],
            ["Pendiente de cobro a hoy (im_CarteraVencida)", round(total_vencido, 2)],
        ],
    )

    usados: set = {"Resumen"}

    ws_tipo = wb.create_sheet(nombre_hoja_valido("Tipo de negocio", usados))
    escribir_hoja_excel(
        ws_tipo,
        ["Tipo de negocio", "Facturas", "Cartera esperada", "Pendiente a hoy"],
        [[tipo, folios, round(esperado, 2), round(vencido, 2)] for tipo, esperado, vencido, folios in por_tipo],
    )

    ws_top = wb.create_sheet(nombre_hoja_valido("Top clientes", usados))
    escribir_hoja_excel(
        ws_top,
        ["Cliente", "Cartera esperada"],
        [[cliente, round(monto, 2)] for cliente, monto in top_clientes],
    )

    ws_detalle = wb.create_sheet(nombre_hoja_valido("Detalle periodo", usados))
    ws_detalle.append(COLUMNAS_DETALLE)
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    for celda in ws_detalle[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1B3A5B")
    for fila in filas:
        ws_detalle.append([fila.get(col) for col in COLUMNAS_DETALLE])
    for i, col in enumerate(COLUMNAS_DETALLE, start=1):
        letra = get_column_letter(i)
        ws_detalle.column_dimensions[letra].width = max(12, min(30, len(col) + 2))
        if col == "fh_Vencimiento":
            for fila_celdas in ws_detalle.iter_rows(min_row=2, min_col=i, max_col=i):
                for celda in fila_celdas:
                    celda.number_format = "dd/mm/yyyy"
    ws_detalle.freeze_panes = "A2"

    return wb


def construir_subtab_resumen(page: ft.Page) -> ft.Control:
    """Contenido de la sub-pestaña 'Resumen'. Se construye UNA sola vez al
    armar la pestaña (dispara sus consultas al inicio); sus refrescos internos
    reemplazan `.controls` de los contenedores, igual que Segmentado."""
    hoy = date.today()
    hace_una_semana = hoy - timedelta(days=7)
    rango_sel: list[tuple[date, date]] = [(hace_una_semana, hoy)]

    def _dark() -> bool:
        return page.theme_mode == ft.ThemeMode.DARK

    def _texto_rango(inicio: date, fin: date) -> str:
        return f"{inicio.strftime('%d %b %Y')} – {fin.strftime('%d %b %Y')}"

    titulo = ft.Text("Reporte de Cumplimiento de Cobro", size=20, weight=ft.FontWeight.W_600,
                      color=ft.Colors.ON_SURFACE)
    subtitulo = ft.Text(
        "Facturas cuya fecha de vencimiento (fh_Vencimiento) cae en el rango seleccionado abajo "
        "— por default, la semana pasada.",
        size=12,
        color=ft.Colors.ON_SURFACE_VARIANT,
    )

    estado_text = ft.Text("", size=12, color=ft.Colors.RED_600)
    progress = ft.ProgressRing(width=16, height=16, visible=False, stroke_width=2)

    hero_contenedor = ft.ResponsiveRow(spacing=16, run_spacing=16)
    seccion_tipos = ft.Container()
    seccion_top_clientes = ft.Container()
    cuerpo = ft.Column([hero_contenedor, seccion_tipos, seccion_top_clientes], spacing=20,
                        opacity=1.0, animate_opacity=200)

    # En Flet 0.85 el FilePicker es un servicio: se crea y se usa directamente
    # (NO se agrega a page.overlay; hacerlo provoca "Unknown control: FilePicker").
    file_picker = ft.FilePicker()

    ultimo_detalle: list[list[dict]] = [[]]  # filas crudas de la última consulta exitosa, para exportar

    def _refrescar(resultado) -> None:
        dark = _dark()
        if isinstance(resultado, Exception):
            hero_contenedor.controls = []
            mensaje = ft.Container(
                content=ft.Text(f"No se pudo consultar: {resultado}", size=12, color=ft.Colors.RED_600),
                height=160,
                alignment=ft.Alignment.CENTER,
            )
            seccion_tipos.content = mensaje
            seccion_top_clientes.content = ft.Container()
            boton_exportar.disabled = True
            return

        filas = resultado
        ultimo_detalle[0] = filas
        boton_exportar.disabled = False

        total_facturas = len({f["fl_FolioDocumento"] for f in filas if f.get("fl_FolioDocumento")})
        total_esperado = sum(f.get("im_Cartera") or 0 for f in filas)
        total_vencido = sum(f.get("im_CarteraVencida") or 0 for f in filas)
        por_tipo = _agregar_por_tipo_negocio(filas)
        top_clientes = _top_clientes(filas)

        hero_contenedor.controls = [
            hero_tile(
                "Total de facturas del periodo", total_facturas, color_slot(0, dark),
                ft.Icons.RECEIPT_LONG_OUTLINED, "Folios distintos con vencimiento en el rango seleccionado",
                formatear=lambda v: f"{v:,.0f}",
            ),
            hero_tile(
                "Total esperado a cobrar", total_esperado, color_slot(1, dark),
                ft.Icons.ACCOUNT_BALANCE_WALLET_OUTLINED, "Saldo total de las facturas del periodo (im_Cartera)",
            ),
            hero_tile(
                "Pendiente de cobro a hoy", total_vencido, color_slot(5, dark),
                ft.Icons.WARNING_AMBER_OUTLINED, "De lo esperado, lo que sigue sin cobrarse a la fecha (im_CarteraVencida)",
            ),
        ]

        if not por_tipo:
            cuerpo_tipos: ft.Control = estado_vacio()
        else:
            tarjetas_tipo = [
                tile_compacta(
                    tipo, esperado, color_slot(i, dark), ft.Icons.CATEGORY_OUTLINED,
                    f"{folios:,} factura(s) · Pendiente hoy: ${vencido:,.0f}",
                    col={"xs": 12, "sm": 6, "lg": 4},
                )
                for i, (tipo, esperado, vencido, folios) in enumerate(por_tipo)
            ]
            cuerpo_tipos = ft.ResponsiveRow(tarjetas_tipo, spacing=12, run_spacing=12)
        seccion_tipos.content = ft.Container(
            content=ft.Column(
                [
                    encabezado_seccion(
                        ft.Icons.DONUT_SMALL, color_slot(2, dark),
                        "Cartera esperada por tipo de negocio",
                        "Suma de im_Cartera y folios distintos, por nb_TipoDeNegocio · el subtexto muestra "
                        "lo aún pendiente a hoy (im_CarteraVencida)",
                    ),
                    ft.Divider(height=1),
                    cuerpo_tipos,
                ],
                spacing=10,
            ),
            padding=16,
            bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
            border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
            border_radius=12,
            shadow=sombra_tarjeta(),
        )

        cuerpo_top = estado_vacio() if not top_clientes else construir_ranked_list(top_clientes, dark)
        seccion_top_clientes.content = ft.Container(
            content=ft.Column(
                [
                    encabezado_seccion(
                        ft.Icons.PEOPLE_ALT_OUTLINED, color_slot(3, dark),
                        "Top 15 clientes por cartera esperada",
                        "Agregado por nb_Cliente (im_Cartera), ordenado de mayor a menor",
                    ),
                    ft.Divider(height=1),
                    cuerpo_top,
                ],
                spacing=10,
            ),
            padding=16,
            bgcolor=ft.Colors.SURFACE_CONTAINER_LOWEST,
            border=ft.Border.all(1, ft.Colors.OUTLINE_VARIANT),
            border_radius=12,
            shadow=sombra_tarjeta(),
        )

    async def cargar(_e=None) -> None:
        cuerpo.opacity = 0.5
        progress.visible = True
        estado_text.value = ""
        boton_rango.disabled = True
        page.update()

        fecha_inicio, fecha_fin = rango_sel[0]
        try:
            resultado = await asyncio.to_thread(_repo().detalle_periodo, fecha_inicio, fecha_fin)
        except Exception as error:  # noqa: BLE001 - se muestra en la sección, igual que el resto del dashboard
            resultado = error

        _refrescar(resultado)

        progress.visible = False
        boton_rango.disabled = False
        cuerpo.opacity = 1.0
        if isinstance(resultado, Exception):
            estado_text.value = "No se pudo consultar BigQuery (ver detalle abajo)."
        page.update()

    async def exportar_excel(_e) -> None:
        if not ultimo_detalle[0]:
            return
        boton_exportar.disabled = True
        page.update()

        filas = ultimo_detalle[0]
        fecha_inicio, fecha_fin = rango_sel[0]
        total_facturas = len({f["fl_FolioDocumento"] for f in filas if f.get("fl_FolioDocumento")})
        total_esperado = sum(f.get("im_Cartera") or 0 for f in filas)
        total_vencido = sum(f.get("im_CarteraVencida") or 0 for f in filas)
        por_tipo = _agregar_por_tipo_negocio(filas)
        top_clientes = _top_clientes(filas)

        wb = _construir_workbook_reporte(
            filas, fecha_inicio, fecha_fin, total_facturas, total_esperado, total_vencido, por_tipo, top_clientes
        )
        nombre_def = f"cumplimiento_cobro_{fecha_inicio:%Y%m%d}_{fecha_fin:%Y%m%d}.xlsx"
        ok, mensaje = await guardar_workbook(page, file_picker, wb, nombre_def)
        boton_exportar.disabled = False
        estado_text.value = mensaje
        page.update()

    def on_cambiar_rango(e) -> None:
        picker = e.control
        if not picker.start_value or not picker.end_value:
            return
        inicio = picker.start_value.date()
        fin = picker.end_value.date()
        rango_sel[0] = (inicio, fin)
        boton_rango.content.controls[1].value = _texto_rango(inicio, fin)
        page.update()
        page.run_task(cargar)

    date_range_picker = ft.DateRangePicker(
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2035, 12, 31),
        start_value=datetime.combine(hace_una_semana, datetime.min.time()),
        end_value=datetime.combine(hoy, datetime.min.time()),
        entry_mode=ft.DatePickerEntryMode.CALENDAR_ONLY,
        on_change=on_cambiar_rango,
    )

    boton_rango = ft.OutlinedButton(
        content=ft.Row(
            [ft.Icon(ft.Icons.DATE_RANGE, size=16), ft.Text(_texto_rango(hace_una_semana, hoy), size=13)],
            spacing=8,
            tight=True,
        ),
        style=ft.ButtonStyle(padding=ft.Padding(left=12, right=12, top=6, bottom=6)),
        on_click=lambda _e: page.show_dialog(date_range_picker),
    )

    def _abrir_info(_e) -> None:
        lineas = [
            "Se filtra por fh_Vencimiento dentro del rango seleccionado (por default, la semana pasada: "
            "hoy - 7 días a hoy).",
            "Igual que en Proyección (RDC), que usa esta misma tabla, se excluyen las filas sin cliente o "
            "sin factura, el cliente 'ICV', las filas cuyo nombre de cliente contenga la palabra 'totales' "
            "(subtotales del reporte, no clientes reales) y los folios que empiezan con 'FCOR'.",
            "Total de facturas del periodo: folios distintos (fl_FolioDocumento) dentro del rango.",
            "Total esperado a cobrar: suma de im_Cartera (saldo total original de cada documento) de "
            "todas las filas del periodo — lo que se esperaba cobrar en ese rango, sin importar si ya "
            "se cobró o no.",
            "Pendiente de cobro a hoy: suma de im_CarteraVencida — de ese total esperado, lo que sigue "
            "sin cobrarse AL MOMENTO DE CONSULTAR. OJO: no se usa im_CarteraVigente para 'lo esperado' "
            "porque esa columna es una foto del día de hoy (la parte del saldo que aún no vence al "
            "momento de consultar) y sale en 0 para cualquier factura cuyo vencimiento ya pasó — el caso "
            "normal en este reporte, que por default mira la semana ANTERIOR.",
            "Cartera esperada por tipo de negocio: suma de im_Cartera, im_CarteraVencida y folios "
            "distintos, agrupado por nb_TipoDeNegocio. Las filas sin tipo de negocio asignado se agrupan "
            "como 'Sin asignar'.",
            "Top de clientes: suma de im_Cartera agrupada por nb_Cliente, de mayor a menor, recortado a "
            "las 15 primeras.",
            "El Excel exportado incluye además el detalle crudo del periodo con nb_Cliente, nb_Sucursal, "
            "fl_FolioDocumento, fh_Vencimiento, im_Cartera, im_CarteraVencida, nb_Empresa, "
            "nb_TipoDeNegocio y sn_filial.",
            "La sub-pestaña 'Detalle' (arriba) muestra estos mismos registros fila por fila, con "
            "paginación y filtros multi-selección por columna.",
        ]
        dialogo = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cómo se calculan estos datos"),
            content=ft.Container(
                content=ft.Column(
                    [ft.Text(f"•  {l}", size=12, selectable=True) for l in lineas],
                    spacing=10, scroll=ft.ScrollMode.AUTO,
                ),
                width=520, height=340,
            ),
            actions=[ft.TextButton("Cerrar", on_click=lambda _e: page.pop_dialog())],
        )
        mostrar_dialogo(page, dialogo)

    boton_info = ft.IconButton(
        icon=ft.Icons.INFO_OUTLINE,
        icon_size=18,
        tooltip="Ver cómo se calculan estos datos",
        on_click=_abrir_info,
    )

    boton_exportar = ft.IconButton(
        icon=ft.Icons.DOWNLOAD,
        icon_size=18,
        tooltip="Descargar Excel (resumen + tipo de negocio + top clientes + detalle)",
        disabled=True,
        on_click=lambda e: page.run_task(exportar_excel, e),
    )

    barra_herramientas = ft.Container(
        content=ft.Row(
            [boton_rango, progress, estado_text, ft.Container(expand=True), boton_exportar, boton_info],
            spacing=12,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        padding=ft.Padding(left=14, right=14, top=10, bottom=10),
        bgcolor=ft.Colors.SURFACE_CONTAINER_LOW,
        border_radius=12,
    )

    contenido = ft.Container(
        content=ft.Column(
            [
                ft.Column([titulo, subtitulo], spacing=2),
                barra_herramientas,
                ft.Container(content=cuerpo, expand=True),
            ],
            spacing=16,
            scroll=ft.ScrollMode.AUTO,
            expand=True,
        ),
        padding=20,
        expand=True,
    )

    page.run_task(cargar)

    return contenido
