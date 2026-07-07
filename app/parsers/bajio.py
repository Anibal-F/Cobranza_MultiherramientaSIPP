"""Parser del estado de cuenta de BanBajío (Conecta BanBajío), en formato .xlsx.

A diferencia de Santander/BanRegio/Banorte (que llegan como .csv), BanBajío se
descarga como Excel. La hoja 'ConsultaMovimientos' trae unas filas de "Datos
Generales" y, más abajo, una tabla de movimientos con encabezado:

    # | Fecha Movimiento | Hora | Recibo | Descripción | Cargos | Abonos | Saldo

Se procesan solo los ABONOS (entregas de recursos / cobros); los cargos
(comisiones, IVA, traspasos/SPEI enviados) se ignoran.
"""

import csv as _csv
import os as _os
import tempfile as _tempfile
from datetime import date, datetime

from ..models import Movimiento
from ..textutils import normalizar
from .base import parse_date_dmy_slash, parse_money

BANCO = "BANBAJIO"

# Marcadores del encabezado de la tabla (normalizados): identifican la fila de
# columnas y, a la vez, distinguen el archivo de BanBajío de otros .xlsx.
_HEADERS_TABLA = {"RECIBO", "DESCRIPCION", "ABONOS"}


def _key(valor) -> str:
    """Clave robusta de encabezado: sin acentos, mayúsculas, solo alfanumérico."""
    return normalizar(str(valor if valor is not None else "")).upper()


def _monto(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    return parse_money(str(valor))


def _fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return parse_date_dmy_slash(str(valor)) if valor else None


def _buscar_fila_encabezado(ws) -> int | None:
    """Devuelve el nº de fila (1-based) donde está el encabezado de la tabla de
    movimientos, o None si la hoja no corresponde a BanBajío."""
    max_r = min(ws.max_row or 0, 40)
    max_c = min(ws.max_column or 0, 12)
    for r in range(1, max_r + 1):
        valores = {_key(ws.cell(row=r, column=c).value) for c in range(1, max_c + 1)}
        if _HEADERS_TABLA.issubset(valores):
            return r
    return None


def _mapa_columnas(ws, fila: int) -> dict[str, int]:
    mapa: dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        clave = _key(ws.cell(row=fila, column=c).value)
        if clave and clave not in mapa:
            mapa[clave] = c
    return mapa


def detect(path: str) -> bool:
    if not path.lower().endswith((".xlsx", ".xlsm")):
        return False
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return False
    try:
        return any(_buscar_fila_encabezado(ws) is not None for ws in wb.worksheets)
    finally:
        wb.close()


def _fmt_celda_csv(valor) -> str:
    """Formatea una celda para el CSV de salida, imitando la vista del reporte."""
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        # Fecha a medianoche → solo fecha; con hora → solo hora (columna 'Hora').
        if (valor.hour, valor.minute, valor.second) == (0, 0, 0):
            return valor.strftime("%d/%m/%Y")
        return valor.strftime("%H:%M:%S")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def convertir_a_csv(xlsx_path: str, destino: str | None = None) -> str:
    """Vuelca la hoja de movimientos de BanBajío a un CSV (equivalente a 'Guardar
    como .csv' en Excel), para probar si el importador de SIPP acepta el formato
    CSV. Devuelve la ruta del CSV generado (temporal si no se indica `destino`)."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if destino is None:
        fd, destino = _tempfile.mkstemp(suffix=".csv", prefix="bajio_")
        _os.close(fd)
    try:
        ws = next((w for w in wb.worksheets if _buscar_fila_encabezado(w) is not None), wb.active)
        with open(destino, "w", encoding="utf-8-sig", newline="") as f:
            escritor = _csv.writer(f)
            for fila in ws.iter_rows(values_only=True):
                # Omite filas totalmente vacías para no meter renglones en blanco.
                if all(v is None or str(v).strip() == "" for v in fila):
                    continue
                escritor.writerow([_fmt_celda_csv(v) for v in fila])
    finally:
        wb.close()
    return destino


def parse(path: str) -> list[Movimiento]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    movimientos: list[Movimiento] = []
    try:
        for ws in wb.worksheets:
            fila_enc = _buscar_fila_encabezado(ws)
            if fila_enc is None:
                continue
            mapa = _mapa_columnas(ws, fila_enc)
            col_fecha = mapa.get("FECHA MOVIMIENTO")
            col_recibo = mapa.get("RECIBO")
            col_desc = mapa.get("DESCRIPCION")
            col_cargo = mapa.get("CARGOS")
            col_abono = mapa.get("ABONOS")
            col_saldo = mapa.get("SALDO")

            for r in range(fila_enc + 1, (ws.max_row or fila_enc) + 1):
                descripcion = ws.cell(row=r, column=col_desc).value if col_desc else None
                descripcion = (str(descripcion).strip() if descripcion is not None else "")
                # Fin de la tabla: fila sin descripción o la nota informativa final.
                if not descripcion or _key(descripcion).startswith("NOTA"):
                    break

                # Solo abonos (cobros). Cargos (comisiones, IVA, SPEI enviado /
                # traspasos) se ignoran.
                abono = _monto(ws.cell(row=r, column=col_abono).value) if col_abono else 0.0
                if abono <= 0:
                    continue

                # Compensaciones por desfase de SPEI: ajustes internos del banco.
                if "COMPENSACION" in normalizar(descripcion).upper():
                    continue

                recibo = ws.cell(row=r, column=col_recibo).value if col_recibo else None
                recibo = str(recibo).strip() if recibo is not None else ""

                movimientos.append(
                    Movimiento(
                        banco=BANCO,
                        fecha=_fecha(ws.cell(row=r, column=col_fecha).value) if col_fecha else None,
                        descripcion=descripcion,
                        referencia=recibo,
                        concepto="",
                        cargo=_monto(ws.cell(row=r, column=col_cargo).value) if col_cargo else 0.0,
                        abono=abono,
                        saldo=_monto(ws.cell(row=r, column=col_saldo).value) if col_saldo else None,
                        # La descripción trae la cuenta del ordenante, el nombre
                        # ("Ordenante | AGA COMBUSTIBLES SA DE CV") y el folio de
                        # pago: suficiente para el match por cuenta/nombre/folio.
                        texto_busqueda=descripcion,
                    )
                )
    finally:
        wb.close()
    return movimientos
